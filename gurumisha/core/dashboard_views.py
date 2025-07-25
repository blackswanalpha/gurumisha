"""
Dashboard-specific views for Gurumisha Motors
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, DetailView
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Sum, Avg, F, Case, When, IntegerField
from django.db import models
from django.utils import timezone
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.urls import reverse
from django.views.decorators.http import require_http_methods
from datetime import datetime, timedelta
import csv
import json
from io import StringIO

from .models import (
    Car, CarBrand, CarModel, CarImage, VehicleCondition, SparePart, ImportRequest, ImportOrder, ImportOrderStatusHistory,
    Inquiry, Testimonial, BlogPost, Vendor, User,
    Supplier, PurchaseOrder, StockMovement, InventoryAlert,
    Order, OrderItem, Payment, SparePartCategory, Notification, SystemSetting,
    ActivityLog, AuditLog, NotificationPreference, NotificationQueue,
    HotDeal, CarRating, PromotionAnalytics, VendorSubscription, FeaturedCarTier,
    ProfileView, VendorAnalytics, UserActivityLog, ContentCategory, ContentTag,
    ContentSeries, ContentView, ContentLike, ContentComment, ContentBookmark,
    ContentAnalytics, ContentPerformanceReport
)
from .dashboard_forms import (
    UserProfileForm, VendorProfileForm, PasswordChangeForm,
    InquiryResponseForm, CarApprovalForm, VendorApprovalForm,
    UserSearchForm, CarSearchForm, VendorSearchForm, AdminCarEditForm,
    UserPreferencesForm, VendorPreferencesForm, BusinessHoursForm
)
from .forms import SparePartForm, VendorSparePartForm
from .utils.image_utils import default_image_handler
from .utils.analytics_utils import (
    track_profile_view, log_user_activity, get_analytics_dashboard_data,
    get_vendor_analytics_summary, get_user_analytics_summary
)


def apply_import_request_filters(request, queryset):
    """Apply comprehensive filters to import request queryset"""

    # Status filter
    status_filter = request.GET.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    # Search filter (across multiple fields)
    search_query = request.GET.get('search')
    if search_query:
        queryset = queryset.filter(
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query) |
            Q(customer__email__icontains=search_query) |
            Q(customer__username__icontains=search_query) |
            Q(brand__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(tracking_number__icontains=search_query) |
            Q(id__icontains=search_query)
        )

    # Brand filter
    brand_filter = request.GET.get('brand')
    if brand_filter:
        queryset = queryset.filter(brand=brand_filter)

    # Country filter
    country_filter = request.GET.get('country')
    if country_filter:
        queryset = queryset.filter(origin_country=country_filter)

    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if date_from:
        try:
            date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=date_from_parsed)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=date_to_parsed)
        except ValueError:
            pass

    return queryset


@login_required
def user_profile_view(request):
    """Enhanced user profile management view with comprehensive form handling"""
    user = request.user
    vendor = None

    if user.role == 'vendor':
        try:
            vendor = user.vendor
        except Vendor.DoesNotExist:
            # Create vendor profile if it doesn't exist (but not for admin users)
            if user.role != 'admin':
                vendor = Vendor.objects.create(
                    user=user,
                    company_name=f"{user.first_name} {user.last_name}".strip() or user.username,
                    is_approved=False
                )

    if request.method == 'POST':
        # Handle HTMX auto-save requests
        if request.headers.get('X-Auto-Save') == 'true':
            field_name = request.POST.get('field_name')
            field_value = request.POST.get('field_value')

            try:
                # Update specific field
                if hasattr(user, field_name):
                    setattr(user, field_name, field_value)
                    user.save(update_fields=[field_name])

                    # Log the activity
                    log_user_activity(user, 'profile_update', f'Auto-saved {field_name}', request)

                    return JsonResponse({'success': True, 'message': f'{field_name.title()} saved'})
                else:
                    return JsonResponse({'success': False, 'message': 'Invalid field'})

            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)})

        # Handle HTMX validation requests
        if request.headers.get('X-Validate-Only') == 'true':
            field_name = request.POST.get('field_name')
            field_value = request.POST.get('field_value')

            # Create a temporary form instance for validation
            form_data = {field_name: field_value}
            temp_form = UserProfileForm(form_data, instance=user)

            if temp_form.is_valid():
                return JsonResponse({'valid': True})
            else:
                errors = temp_form.errors.get(field_name, [])
                error_message = errors[0] if errors else 'Invalid input'
                return JsonResponse({'valid': False, 'message': error_message})

        # Handle regular form submission
        user_form = UserProfileForm(request.POST, request.FILES, instance=user)
        vendor_form = VendorProfileForm(request.POST, request.FILES, instance=vendor) if vendor else None

        if user_form.is_valid() and (vendor_form is None or vendor_form.is_valid()):
            try:
                # Handle profile picture processing
                if 'profile_picture' in request.FILES:
                    profile_picture = request.FILES['profile_picture']
                    try:
                        # Process the image
                        processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                            profile_picture, 'profile'
                        )

                        # Delete old profile picture if exists
                        if user.profile_picture:
                            default_image_handler.cleanup(user.profile_picture.path)

                        # Save the processed image
                        user.profile_picture.save(filename, processed_file, save=False)

                    except ValueError as e:
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                            return JsonResponse({'success': False, 'message': str(e)})
                        messages.error(request, str(e))
                        return render(request, 'core/dashboard/profile.html', {
                            'user': user, 'vendor': vendor, 'user_form': user_form, 'vendor_form': vendor_form
                        })

                user_form.save()

                # Handle vendor form if exists
                if vendor_form:
                    vendor = vendor_form.save(commit=False)

                    # Handle company logo processing
                    if 'company_logo' in request.FILES:
                        company_logo = request.FILES['company_logo']
                        try:
                            processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                                company_logo, 'logo'
                            )

                            # Delete old logo if exists
                            if vendor.company_logo:
                                default_image_handler.cleanup(vendor.company_logo.path)

                            vendor.company_logo.save(filename, processed_file, save=False)

                        except ValueError as e:
                            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                return JsonResponse({'success': False, 'message': str(e)})
                            messages.error(request, str(e))
                            return render(request, 'core/dashboard/profile.html', {
                                'user': user, 'vendor': vendor, 'user_form': user_form, 'vendor_form': vendor_form
                            })

                    # Handle cover image processing
                    if 'cover_image' in request.FILES:
                        cover_image = request.FILES['cover_image']
                        try:
                            processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                                cover_image, 'cover'
                            )

                            # Delete old cover image if exists
                            if vendor.cover_image:
                                default_image_handler.cleanup(vendor.cover_image.path)

                            vendor.cover_image.save(filename, processed_file, save=False)

                        except ValueError as e:
                            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                                return JsonResponse({'success': False, 'message': str(e)})
                            messages.error(request, str(e))
                            return render(request, 'core/dashboard/profile.html', {
                                'user': user, 'vendor': vendor, 'user_form': user_form, 'vendor_form': vendor_form
                            })

                    vendor.save()

                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Profile updated successfully!'})

                messages.success(request, 'Profile updated successfully!')
                return redirect('core:profile')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Error updating profile. Please try again.'})
                messages.error(request, 'Error updating profile. Please try again.')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                errors = {}
                if user_form.errors:
                    errors.update(user_form.errors)
                if vendor_form and vendor_form.errors:
                    errors.update(vendor_form.errors)
                return JsonResponse({'success': False, 'errors': errors})
            messages.error(request, 'Please correct the errors below.')
    else:
        user_form = UserProfileForm(instance=user)
        vendor_form = VendorProfileForm(instance=vendor) if vendor else None

    # Track profile view
    track_profile_view(request, user)

    # Log profile view activity
    log_user_activity(request.user, 'profile_view', f'Viewed profile page', request)

    # Get analytics data
    analytics_data = get_analytics_dashboard_data(user)

    context = {
        'user': user,
        'vendor': vendor,
        'user_form': user_form,
        'vendor_form': vendor_form,
        'analytics_data': analytics_data,
    }

    return render(request, 'core/dashboard/profile.html', context)


@login_required
def vendor_profile_view(request):
    """Dedicated vendor profile management view"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied. Vendor account required.')
        return redirect('core:dashboard')

    # Redirect admin users to admin dashboard
    if request.user.role == 'admin':
        messages.info(request, 'As an admin, please use the admin dashboard.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        # Create vendor profile if it doesn't exist (but not for admin users)
        if request.user.role != 'admin':
            vendor = Vendor.objects.create(
                user=request.user,
                company_name=f"{request.user.first_name} {request.user.last_name}".strip() or request.user.username,
                is_approved=False
            )
        else:
            messages.error(request, 'Admin users cannot have vendor profiles.')
            return redirect('core:dashboard')

    if request.method == 'POST':
        vendor_form = VendorProfileForm(request.POST, request.FILES, instance=vendor)

        if vendor_form.is_valid():
            try:
                vendor_form.save()

                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': 'Business profile updated successfully!'})

                messages.success(request, 'Business profile updated successfully!')
                return redirect('core:vendor_profile')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'message': 'Error updating business profile. Please try again.'})
                messages.error(request, 'Error updating business profile. Please try again.')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'errors': vendor_form.errors})
            messages.error(request, 'Please correct the errors below.')
    else:
        vendor_form = VendorProfileForm(instance=vendor)

    context = {
        'vendor': vendor,
        'vendor_form': vendor_form,
        'user': request.user,
    }

    return render(request, 'core/dashboard/vendor_profile.html', context)


@login_required
def change_password_view(request):
    """Handle password change requests"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)

        if form.is_valid():
            try:
                form.save()

                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Password changed successfully! Please log in again.'
                    })

                messages.success(request, 'Password changed successfully! Please log in again.')
                from django.contrib.auth import logout
                logout(request)
                return redirect('core:login')

            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Error changing password. Please try again.'
                    })
                messages.error(request, 'Error changing password. Please try again.')
        else:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'errors': form.errors,
                    'message': 'Please correct the errors below.'
                })
            messages.error(request, 'Please correct the errors below.')

    # For GET requests or form errors, return to profile page
    return redirect('core:profile')


@login_required
def user_settings_view(request):
    """Enhanced user settings and preferences management"""
    user = request.user
    vendor = None

    if user.role == 'vendor':
        try:
            vendor = user.vendor
        except Vendor.DoesNotExist:
            vendor = Vendor.objects.create(
                user=user,
                company_name=f"{user.first_name} {user.last_name}".strip() or user.username,
                is_approved=False
            )

    if request.method == 'POST':
        # Handle different form types based on the submitted data
        form_type = request.POST.get('form_type', 'preferences')

        if form_type == 'preferences':
            user_prefs_form = UserPreferencesForm(request.POST, instance=user)
            vendor_prefs_form = VendorPreferencesForm(request.POST, instance=vendor) if vendor else None

            if user_prefs_form.is_valid() and (vendor_prefs_form is None or vendor_prefs_form.is_valid()):
                try:
                    user_prefs_form.save()
                    if vendor_prefs_form:
                        vendor_prefs_form.save()

                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': True, 'message': 'Preferences updated successfully!'})

                    messages.success(request, 'Preferences updated successfully!')
                    return redirect('core:user_settings')

                except Exception as e:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'Error updating preferences.'})
                    messages.error(request, 'Error updating preferences.')
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    errors = {}
                    if user_prefs_form.errors:
                        errors.update(user_prefs_form.errors)
                    if vendor_prefs_form and vendor_prefs_form.errors:
                        errors.update(vendor_prefs_form.errors)
                    return JsonResponse({'success': False, 'errors': errors})
                messages.error(request, 'Please correct the errors below.')

        elif form_type == 'business_hours' and vendor:
            business_hours_form = BusinessHoursForm(request.POST, vendor=vendor)

            if business_hours_form.is_valid():
                try:
                    # Update vendor operating days
                    for day_code, day_name in BusinessHoursForm.DAYS_OF_WEEK:
                        operates_field = f'operates_{day_code}'
                        operates = business_hours_form.cleaned_data.get(f'operates_{day_code}', False)
                        setattr(vendor, operates_field, operates)

                    # Save business hours as JSON
                    business_hours = {}
                    for day_code, day_name in BusinessHoursForm.DAYS_OF_WEEK:
                        if business_hours_form.cleaned_data.get(f'operates_{day_code}'):
                            open_time = business_hours_form.cleaned_data.get(f'{day_code}_open')
                            close_time = business_hours_form.cleaned_data.get(f'{day_code}_close')
                            if open_time and close_time:
                                business_hours[day_code] = {
                                    'open': open_time.strftime('%H:%M'),
                                    'close': close_time.strftime('%H:%M')
                                }

                    import json
                    vendor.business_hours = json.dumps(business_hours)
                    vendor.save()

                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': True, 'message': 'Business hours updated successfully!'})

                    messages.success(request, 'Business hours updated successfully!')
                    return redirect('core:user_settings')

                except Exception as e:
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({'success': False, 'message': 'Error updating business hours.'})
                    messages.error(request, 'Error updating business hours.')
            else:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'errors': business_hours_form.errors})
                messages.error(request, 'Please correct the errors below.')

    # For GET requests, initialize forms
    user_prefs_form = UserPreferencesForm(instance=user)
    vendor_prefs_form = VendorPreferencesForm(instance=vendor) if vendor else None
    business_hours_form = BusinessHoursForm(vendor=vendor) if vendor else None

    context = {
        'user': user,
        'vendor': vendor,
        'user_prefs_form': user_prefs_form,
        'vendor_prefs_form': vendor_prefs_form,
        'business_hours_form': business_hours_form,
    }

    return render(request, 'core/dashboard/settings.html', context)


@login_required
def profile_analytics_view(request):
    """Profile analytics dashboard"""
    user = request.user

    # Get comprehensive analytics data
    analytics_data = get_analytics_dashboard_data(user)

    if not analytics_data:
        messages.error(request, 'Unable to load analytics data.')
        return redirect('core:profile')

    # Handle AJAX requests for chart data
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        chart_type = request.GET.get('chart_type')

        if chart_type == 'profile_views' and user.role == 'vendor':
            # Return profile views chart data
            daily_views = analytics_data.get('daily_views', [])
            chart_data = {
                'labels': [item['day'] for item in daily_views],
                'data': [item['views'] for item in daily_views],
                'title': 'Profile Views (Last 30 Days)'
            }
            return JsonResponse(chart_data)

        elif chart_type == 'activity_breakdown':
            # Return activity breakdown chart data
            activity_breakdown = analytics_data.get('activity_breakdown', [])
            chart_data = {
                'labels': [item['action'].replace('_', ' ').title() for item in activity_breakdown],
                'data': [item['count'] for item in activity_breakdown],
                'title': 'Activity Breakdown'
            }
            return JsonResponse(chart_data)

    context = {
        'user': user,
        'analytics_data': analytics_data,
    }

    return render(request, 'core/dashboard/analytics.html', context)


@login_required
def admin_user_detail_view(request, user_id):
    """Admin user detail and management view"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('core:dashboard')

    # Get the user
    user = get_object_or_404(User, id=user_id)
    vendor = None

    # Get vendor profile if user is a vendor
    if user.role == 'vendor':
        try:
            vendor = user.vendor
        except Vendor.DoesNotExist:
            vendor = None

    # Handle form submissions
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_profile':
            # Handle profile update
            user_form = UserProfileForm(request.POST, request.FILES, instance=user)
            vendor_form = VendorProfileForm(request.POST, request.FILES, instance=vendor) if vendor else None

            if user_form.is_valid() and (vendor_form is None or vendor_form.is_valid()):
                try:
                    # Handle profile picture processing
                    if 'profile_picture' in request.FILES:
                        profile_picture = request.FILES['profile_picture']
                        try:
                            processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                                profile_picture, 'profile'
                            )

                            if user.profile_picture:
                                default_image_handler.cleanup(user.profile_picture.path)

                            user.profile_picture.save(filename, processed_file, save=False)

                        except ValueError as e:
                            messages.error(request, f'Image processing error: {str(e)}')
                            return redirect('core:admin_user_detail', user_id=user_id)

                    user_form.save()

                    # Handle vendor form if exists
                    if vendor_form:
                        vendor = vendor_form.save(commit=False)

                        # Handle company logo processing
                        if 'company_logo' in request.FILES:
                            company_logo = request.FILES['company_logo']
                            try:
                                processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                                    company_logo, 'logo'
                                )

                                if vendor.company_logo:
                                    default_image_handler.cleanup(vendor.company_logo.path)

                                vendor.company_logo.save(filename, processed_file, save=False)

                            except ValueError as e:
                                messages.error(request, f'Logo processing error: {str(e)}')
                                return redirect('core:admin_user_detail', user_id=user_id)

                        vendor.save()

                    # Log the activity
                    log_user_activity(
                        request.user,
                        'profile_update',
                        f'Admin updated profile for user {user.username}',
                        request,
                        {'target_user_id': user.id}
                    )

                    messages.success(request, f'Profile updated successfully for {user.username}!')
                    return redirect('core:admin_user_detail', user_id=user_id)

                except Exception as e:
                    messages.error(request, f'Error updating profile: {str(e)}')
            else:
                messages.error(request, 'Please correct the errors below.')

        elif action == 'change_role':
            # Handle role change
            new_role = request.POST.get('new_role')
            if new_role in ['customer', 'vendor', 'admin']:
                old_role = user.role
                user.role = new_role
                user.save()

                # Create vendor profile if changing to vendor (but not for admin users)
                if new_role == 'vendor' and not hasattr(user, 'vendor') and old_role != 'admin':
                    Vendor.objects.create(
                        user=user,
                        company_name=f"{user.first_name} {user.last_name}".strip() or user.username,
                        is_approved=False
                    )
                elif new_role == 'vendor' and old_role == 'admin':
                    messages.warning(request, 'Admin users cannot be changed to vendor role. Please create a separate vendor account.')

                # Log the activity
                log_user_activity(
                    request.user,
                    'role_change',
                    f'Admin changed role for {user.username} from {old_role} to {new_role}',
                    request,
                    {'target_user_id': user.id, 'old_role': old_role, 'new_role': new_role}
                )

                messages.success(request, f'Role changed from {old_role} to {new_role} for {user.username}!')
                return redirect('core:admin_user_detail', user_id=user_id)
            else:
                messages.error(request, 'Invalid role selected.')

        elif action == 'toggle_status':
            # Handle account activation/deactivation
            user.is_active = not user.is_active
            user.save()

            status = 'activated' if user.is_active else 'deactivated'
            log_user_activity(
                request.user,
                'account_status_change',
                f'Admin {status} account for {user.username}',
                request,
                {'target_user_id': user.id, 'new_status': user.is_active}
            )

            messages.success(request, f'Account {status} for {user.username}!')
            return redirect('core:admin_user_detail', user_id=user_id)

        elif action == 'toggle_vendor_approval':
            # Handle vendor approval/disapproval
            if vendor:
                vendor.is_approved = not vendor.is_approved
                if vendor.is_approved:
                    vendor.approval_date = timezone.now()
                    vendor.verification_status = 'verified'
                else:
                    vendor.approval_date = None
                    vendor.verification_status = 'pending'
                vendor.save()

                status = 'approved' if vendor.is_approved else 'disapproved'
                log_user_activity(
                    request.user,
                    'vendor_status_change',
                    f'Admin {status} vendor {vendor.company_name}',
                    request,
                    {'vendor_id': vendor.id, 'new_status': vendor.is_approved}
                )

                messages.success(request, f'Vendor {status} successfully!')
                return redirect('core:admin_user_detail', user_id=user_id)
            else:
                messages.error(request, 'User is not a vendor.')

        elif action == 'update_vendor':
            # Handle vendor profile update
            if vendor and vendor_form.is_valid():
                try:
                    vendor_form.save()
                    log_user_activity(
                        request.user,
                        'vendor_profile_update',
                        f'Admin updated vendor profile for {vendor.company_name}',
                        request,
                        {'vendor_id': vendor.id}
                    )
                    messages.success(request, 'Vendor profile updated successfully!')
                    return redirect('core:admin_user_detail', user_id=user_id)
                except Exception as e:
                    messages.error(request, f'Error updating vendor profile: {str(e)}')
            else:
                messages.error(request, 'Please correct the errors in the vendor form.')

        elif action == 'reset_password':
            # Handle password reset
            new_password = User.objects.make_random_password(length=12)
            user.set_password(new_password)
            user.save()

            # Log the activity
            log_user_activity(
                request.user,
                'password_reset',
                f'Admin reset password for {user.username}',
                request,
                {'target_user_id': user.id}
            )

            messages.success(request, f'Password reset for {user.username}! New password: {new_password}')
            return redirect('core:admin_user_detail', user_id=user_id)

    # For GET requests, initialize forms
    user_form = UserProfileForm(instance=user)
    vendor_form = VendorProfileForm(instance=vendor) if vendor else None

    # Get user analytics and activity
    analytics_data = get_analytics_dashboard_data(user)

    # Get recent activity logs
    recent_activities = UserActivityLog.objects.filter(
        user=user
    ).order_by('-timestamp')[:20]

    # Get profile views if vendor
    profile_views = []
    if vendor:
        profile_views = ProfileView.objects.filter(
            profile_user=user
        ).order_by('-viewed_at')[:10]

    # Get user's orders, listings, etc.
    user_orders = []
    user_listings = []
    user_inquiries = []

    if hasattr(user, 'orders'):
        user_orders = user.orders.all().order_by('-created_at')[:5]

    if user.role == 'vendor' and vendor:
        user_listings = Car.objects.filter(vendor=vendor).order_by('-created_at')[:5]

    if hasattr(user, 'inquiries_sent'):
        user_inquiries = user.inquiries_sent.all().order_by('-created_at')[:5]

    context = {
        'user_profile': user,  # Renamed to avoid conflict with request.user
        'vendor': vendor,
        'user_form': user_form,
        'vendor_form': vendor_form,
        'analytics_data': analytics_data,
        'recent_activities': recent_activities,
        'profile_views': profile_views,
        'user_orders': user_orders,
        'user_listings': user_listings,
        'user_inquiries': user_inquiries,
    }

    return render(request, 'core/dashboard/admin_user_detail.html', context)


@login_required
def admin_vendor_user_detail_view(request, user_id):
    """Enhanced admin vendor user detail and management view"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('core:dashboard')

    # Get the user and ensure they are a vendor
    user = get_object_or_404(User, id=user_id)

    if user.role != 'vendor':
        messages.error(request, 'This user is not a vendor.')
        return redirect('core:admin_users')

    # Get vendor profile
    try:
        vendor = user.vendor
    except Vendor.DoesNotExist:
        messages.error(request, 'Vendor profile not found.')
        return redirect('core:admin_users')

    # Handle form submissions
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'update_profile':
            # Handle profile update
            user_form = UserProfileForm(request.POST, request.FILES, instance=user)
            vendor_form = VendorProfileForm(request.POST, request.FILES, instance=vendor)

            if user_form.is_valid() and vendor_form.is_valid():
                try:
                    # Handle profile picture processing
                    if 'profile_picture' in request.FILES:
                        profile_picture = request.FILES['profile_picture']
                        try:
                            processed_file, filename, thumbnail_file, thumbnail_filename = default_image_handler.process(
                                profile_picture, 'profile'
                            )

                            if user.profile_picture:
                                default_image_handler.cleanup(user.profile_picture.path)

                            user.profile_picture.save(filename, processed_file, save=False)

                        except ValueError as e:
                            messages.error(request, f'Image processing error: {str(e)}')
                            return redirect('core:admin_vendor_user_detail', user_id=user_id)

                    user_form.save()
                    vendor_form.save()

                    log_user_activity(
                        request.user,
                        'vendor_profile_update',
                        f'Admin updated vendor profile for {vendor.company_name}',
                        request,
                        {'vendor_id': vendor.id, 'user_id': user.id}
                    )

                    messages.success(request, 'Vendor profile updated successfully!')
                    return redirect('core:admin_vendor_user_detail', user_id=user_id)

                except Exception as e:
                    messages.error(request, f'Error updating profile: {str(e)}')
            else:
                messages.error(request, 'Please correct the errors in the form.')

        elif action == 'approve_vendor':
            # Approve vendor
            vendor.is_approved = True
            vendor.approval_date = timezone.now()
            vendor.save()

            log_user_activity(
                request.user,
                'vendor_approval',
                f'Admin approved vendor {vendor.company_name}',
                request,
                {'vendor_id': vendor.id, 'user_id': user.id}
            )

            messages.success(request, f'Vendor {vendor.company_name} has been approved successfully.')
            return redirect('core:admin_vendor_user_detail', user_id=user_id)

        elif action == 'suspend_vendor':
            # Suspend vendor
            vendor.is_approved = False
            vendor.approval_date = None
            vendor.save()

            log_user_activity(
                request.user,
                'vendor_suspension',
                f'Admin suspended vendor {vendor.company_name}',
                request,
                {'vendor_id': vendor.id, 'user_id': user.id}
            )

            messages.warning(request, f'Vendor {vendor.company_name} has been suspended.')
            return redirect('core:admin_vendor_user_detail', user_id=user_id)

        elif action == 'toggle_status':
            # Handle account activation/deactivation
            user.is_active = not user.is_active
            user.save()

            status = 'activated' if user.is_active else 'deactivated'
            log_user_activity(
                request.user,
                'account_status_change',
                f'Admin {status} vendor account for {user.username}',
                request,
                {'target_user_id': user.id, 'new_status': user.is_active}
            )

            messages.success(request, f'Vendor account {status} for {user.username}!')
            return redirect('core:admin_vendor_user_detail', user_id=user_id)

        elif action == 'reset_password':
            # Handle password reset
            new_password = User.objects.make_random_password(length=12)
            user.set_password(new_password)
            user.save()

            log_user_activity(
                request.user,
                'password_reset',
                f'Admin reset password for vendor {user.username}',
                request,
                {'target_user_id': user.id}
            )

            messages.success(request, f'Password reset for {user.username}! New password: {new_password}')
            return redirect('core:admin_vendor_user_detail', user_id=user_id)

    # For GET requests, initialize forms
    user_form = UserProfileForm(instance=user)
    vendor_form = VendorProfileForm(instance=vendor)

    # Get vendor analytics and activity
    analytics_data = get_analytics_dashboard_data(user)

    # Get recent activity logs
    recent_activities = UserActivityLog.objects.filter(
        user=user
    ).order_by('-timestamp')[:20]

    # Get profile views
    profile_views = ProfileView.objects.filter(
        profile_user=user
    ).order_by('-viewed_at')[:10]

    # Get vendor's listings
    vendor_listings = Car.objects.filter(vendor=vendor).order_by('-created_at')[:10]

    # Get vendor's inquiries received
    vendor_inquiries = Inquiry.objects.filter(
        car__vendor=vendor
    ).select_related('user', 'car').order_by('-created_at')[:10]

    # Get vendor's orders/sales
    vendor_orders = []
    if hasattr(vendor, 'orders'):
        vendor_orders = vendor.orders.all().order_by('-created_at')[:10]

    # Get subscription info
    subscription = None
    try:
        subscription = vendor.subscription
    except VendorSubscription.DoesNotExist:
        pass

    context = {
        'user_profile': user,
        'vendor': vendor,
        'user_form': user_form,
        'vendor_form': vendor_form,
        'analytics_data': analytics_data,
        'recent_activities': recent_activities,
        'profile_views': profile_views,
        'vendor_listings': vendor_listings,
        'vendor_inquiries': vendor_inquiries,
        'vendor_orders': vendor_orders,
        'subscription': subscription,
        'is_vendor_detail': True,  # Flag to indicate this is vendor-specific view
    }

    return render(request, 'core/dashboard/admin_vendor_user_detail.html', context)


@login_required
def user_orders_view(request):
    """User orders dashboard"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get user orders
    orders = Order.objects.filter(customer=request.user).prefetch_related('items', 'payments').order_by('-created_at')

    # Calculate stats
    total_orders = orders.count()
    total_spent = sum(order.total_amount for order in orders if order.payment_status == 'completed')
    pending_orders = orders.filter(status='pending').count()
    completed_orders = orders.filter(status='delivered').count()

    context = {
        'orders': orders[:10],  # Recent 10 orders
        'total_orders': total_orders,
        'total_spent': total_spent,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
    }

    return render(request, 'core/dashboard/orders.html', context)


@login_required
def user_addresses_view(request):
    """User addresses management view"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # This would be implemented when address system is added
    context = {
        'addresses': [],  # Placeholder for addresses
    }

    return render(request, 'core/dashboard/addresses.html', context)


@login_required
def user_import_requests_view(request):
    """User import requests management view"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get user's import requests
    import_requests = ImportRequest.objects.filter(customer=request.user).order_by('-created_at')

    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        import_requests = import_requests.filter(status=status_filter)

    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        import_requests = import_requests.filter(
            Q(brand__icontains=search_query) |
            Q(model__icontains=search_query) |
            Q(origin_country__icontains=search_query)
        )

    # Calculate stats
    total_requests = ImportRequest.objects.filter(customer=request.user).count()
    pending_requests = ImportRequest.objects.filter(customer=request.user, status='pending').count()
    completed_requests = ImportRequest.objects.filter(customer=request.user, status='completed').count()
    in_progress_requests = ImportRequest.objects.filter(
        customer=request.user,
        status__in=['on_quotation', 'processing', 'fee_paid']
    ).count()

    context = {
        'import_requests': import_requests,
        'total_requests': total_requests,
        'pending_requests': pending_requests,
        'completed_requests': completed_requests,
        'in_progress_requests': in_progress_requests,
        'status_filter': status_filter,
        'search_query': search_query,
        'status_choices': ImportRequest.STATUS_CHOICES,
    }

    return render(request, 'core/dashboard/user_import_requests.html', context)


@login_required
def user_inquiries_view(request):
    """User inquiries management view"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get user's inquiries
    inquiries = Inquiry.objects.filter(customer=request.user).order_by('-created_at')

    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        inquiries = inquiries.filter(status=status_filter)

    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        inquiries = inquiries.filter(
            Q(subject__icontains=search_query) |
            Q(message__icontains=search_query)
        )

    # Calculate stats
    total_inquiries = Inquiry.objects.filter(customer=request.user).count()
    open_inquiries = Inquiry.objects.filter(customer=request.user, status='open').count()
    resolved_inquiries = Inquiry.objects.filter(customer=request.user, status='resolved').count()
    in_progress_inquiries = Inquiry.objects.filter(customer=request.user, status='in_progress').count()

    context = {
        'inquiries': inquiries,
        'total_inquiries': total_inquiries,
        'open_inquiries': open_inquiries,
        'resolved_inquiries': resolved_inquiries,
        'in_progress_inquiries': in_progress_inquiries,
        'status_filter': status_filter,
        'search_query': search_query,
    }

    return render(request, 'core/dashboard/user_inquiries.html', context)


@login_required
def user_wishlist_view(request):
    """User wishlist management view"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # This would be implemented when wishlist system is added
    context = {
        'wishlist_items': [],  # Placeholder for wishlist items
        'wishlist_count': 0,
    }

    return render(request, 'core/dashboard/user_wishlist.html', context)


@login_required
def user_settings_view(request):
    """User settings management view"""
    if request.user.role != 'customer':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    if request.method == 'POST':
        # Handle settings updates
        pass

    context = {
        'user': request.user,
    }

    return render(request, 'core/dashboard/user_settings.html', context)


@login_required
def user_listings_view(request):
    """User car listings management view - for customers and vendors who submitted cars"""
    # Allow both customers and vendors to access this view
    # Vendors who were originally customers can see their submitted cars here
    if request.user.role not in ['customer', 'vendor']:
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get filter and search parameters
    status_filter = request.GET.get('status')
    search_query = request.GET.get('search')

    # Get user's car listings through their vendor profile
    user_cars = Car.objects.none()
    try:
        # Check if user has a vendor profile (created when they submit a car)
        vendor = request.user.vendor
        user_cars = Car.objects.filter(vendor=vendor).order_by('-created_at')

        # Filter by status if provided
        if status_filter == 'approved':
            user_cars = user_cars.filter(is_approved=True)
        elif status_filter == 'pending':
            user_cars = user_cars.filter(is_approved=False)

        # Search functionality
        if search_query:
            user_cars = user_cars.filter(
                Q(title__icontains=search_query) |
                Q(brand__name__icontains=search_query) |
                Q(model__name__icontains=search_query)
            )

    except Vendor.DoesNotExist:
        # User hasn't submitted any cars yet
        user_cars = Car.objects.none()

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(user_cars, 10)
    page_number = request.GET.get('page')
    cars_page = paginator.get_page(page_number)

    context = {
        'cars': cars_page,
        'total_cars': user_cars.count(),
        'approved_cars': user_cars.filter(is_approved=True).count(),
        'pending_cars': user_cars.filter(is_approved=False).count(),
        'status_filter': status_filter,
        'search_query': search_query,
    }

    return render(request, 'core/dashboard/user_listings.html', context)


@login_required
def vendor_settings_view(request):
    """Vendor settings management view"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')

    if request.method == 'POST':
        # Handle settings updates
        setting_type = request.POST.get('setting_type')

        if setting_type == 'notifications':
            # Handle notification preferences
            vendor.email_notifications = request.POST.get('email_notifications') == 'on'
            vendor.sms_notifications = request.POST.get('sms_notifications') == 'on'
            vendor.inquiry_notifications = request.POST.get('inquiry_notifications') == 'on'
            vendor.order_notifications = request.POST.get('order_notifications') == 'on'
            vendor.save()
            messages.success(request, 'Notification preferences updated successfully.')

        elif setting_type == 'business_hours':
            # Handle business hours
            vendor.business_hours = request.POST.get('business_hours', '')
            vendor.timezone = request.POST.get('timezone', 'Africa/Nairobi')
            vendor.save()
            messages.success(request, 'Business hours updated successfully.')

        elif setting_type == 'payment':
            # Handle payment settings
            vendor.payment_methods = request.POST.get('payment_methods', '')
            vendor.bank_details = request.POST.get('bank_details', '')
            vendor.mpesa_number = request.POST.get('mpesa_number', '')
            vendor.save()
            messages.success(request, 'Payment settings updated successfully.')

        elif setting_type == 'account':
            # Handle account settings
            vendor.auto_approve_inquiries = request.POST.get('auto_approve_inquiries') == 'on'
            vendor.public_profile = request.POST.get('public_profile') == 'on'
            vendor.save()
            messages.success(request, 'Account settings updated successfully.')

        return redirect('core:vendor_settings')

    # Ensure vendor has all required fields with defaults
    if not hasattr(vendor, 'email_notifications'):
        vendor.email_notifications = True
        vendor.sms_notifications = False
        vendor.inquiry_notifications = True
        vendor.order_notifications = True
        vendor.public_profile = True
        vendor.show_contact = True
        vendor.auto_approve_inquiries = False
        vendor.timezone = 'Africa/Nairobi'
        vendor.save()

    context = {
        'vendor': vendor,
        'user': request.user,
    }

    return render(request, 'core/dashboard/vendor_settings.html', context)


@login_required
def vendor_listings_view(request):
    """Enhanced vendor car listings view with comprehensive features"""
    # Check vendor role
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        # Get vendor profile
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')

    # Simple approach to avoid recursion
    try:
        # Get basic car data without complex relationships
        cars = Car.objects.filter(vendor=vendor)

        # Apply filters
        status_filter = request.GET.get('status_filter', '')
        if status_filter == 'approved':
            cars = cars.filter(is_approved=True)
        elif status_filter == 'pending':
            cars = cars.filter(is_approved=False)

        # Apply location filters
        city_filter = request.GET.get('city_filter', '')
        if city_filter:
            cars = cars.filter(city__icontains=city_filter)

        country_filter = request.GET.get('country_filter', '')
        if country_filter:
            cars = cars.filter(country__icontains=country_filter)

        # Apply search
        search = request.GET.get('search', '')
        if search:
            cars = cars.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(area__icontains=search) |
                Q(city__icontains=search) |
                Q(country__icontains=search)
            )

        # Apply sorting
        sort_by = request.GET.get('sort_by', '-created_at')
        cars = cars.order_by(sort_by)

        # Get available cities and countries for filters
        all_vendor_cars = Car.objects.filter(vendor=vendor)
        available_cities = all_vendor_cars.exclude(city__isnull=True).exclude(city='').values_list('city', flat=True).distinct().order_by('city')
        available_countries = all_vendor_cars.exclude(country__isnull=True).exclude(country='').values_list('country', flat=True).distinct().order_by('country')

        # Basic statistics
        total_cars = all_vendor_cars.count()
        approved_cars = all_vendor_cars.filter(is_approved=True).count()
        pending_cars = all_vendor_cars.filter(is_approved=False).count()
        featured_cars = all_vendor_cars.filter(is_featured=True).count()
        hot_deals = all_vendor_cars.filter(is_hot_deal=True).count()
        total_views = 0  # Simplified for now
        total_inquiries = 0  # Simplified for now

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(cars, 20)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        context = {
            'vendor': vendor,
            'cars': page_obj,
            'total_cars': total_cars,
            'approved_cars': approved_cars,
            'pending_cars': pending_cars,
            'featured_cars': featured_cars,
            'hot_deals': hot_deals,
            'total_views': total_views,
            'total_inquiries': total_inquiries,
            'status_filter': request.GET.get('status_filter', ''),
            'sort_by': request.GET.get('sort_by', '-created_at'),
            'search': request.GET.get('search', ''),
            'city_filter': request.GET.get('city_filter', ''),
            'country_filter': request.GET.get('country_filter', ''),
            'available_cities': available_cities,
            'available_countries': available_countries,
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
        }

        # Handle HTMX requests
        if request.headers.get('HX-Request'):
            return render(request, 'core/dashboard/partials/vendor_car_list.html', context)

        # Use the full-featured template
        return render(request, 'core/dashboard/vendor_listings.html', context)

    except Exception as e:
        # If there's still an error, use the minimal template
        return render(request, 'core/dashboard/vendor_listings_minimal.html', {
            'vendor': {'company_name': 'Unknown'},
            'cars': [],
            'total_cars': 0,
            'approved_cars': 0,
            'pending_cars': 0,
            'featured_cars': 0,
            'hot_deals': 0,
            'total_views': 0,
            'total_inquiries': 0,
            'error_message': 'Error loading car listings. Please try again.'
        })


@login_required
def vendor_car_view(request, car_id):
    """View car listing details for vendor"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        car = get_object_or_404(
            Car.objects.select_related('brand', 'model', 'condition', 'vendor')
                      .prefetch_related('images', 'inquiries'),
            id=car_id,
            vendor__user=request.user
        )

        # Get car analytics
        recent_views = car.views_count or 0
        total_inquiries = car.inquiries.count()
        recent_inquiries = car.inquiries.order_by('-created_at')[:5]

        # Get performance metrics
        performance_data = {
            'views_this_month': recent_views,
            'inquiries_this_month': total_inquiries,
            'conversion_rate': (total_inquiries / max(recent_views, 1)) * 100,
            'avg_response_time': '2 hours',
        }

        context = {
            'car': car,
            'recent_inquiries': recent_inquiries,
            'performance_data': performance_data,
            'total_inquiries': total_inquiries,
        }

        # Handle HTMX requests
        if request.headers.get('HX-Request'):
            return render(request, 'core/dashboard/partials/vendor_car_view_modal.html', context)

        return render(request, 'core/dashboard/vendor_car_view.html', context)

    except Exception as e:
        messages.error(request, 'Error loading car details.')
        return redirect('core:vendor_listings')


@login_required
def vendor_car_edit(request, car_id):
    """Edit car listing"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        car = get_object_or_404(Car, id=car_id, vendor__user=request.user)

        if request.method == 'POST':
            # Handle car update with enhanced validation
            car.title = request.POST.get('title', car.title).strip()
            car.description = request.POST.get('description', car.description).strip()

            # Handle numeric fields with validation
            try:
                car.price = float(request.POST.get('price', car.price))
                car.year = int(request.POST.get('year', car.year))
                car.mileage = int(request.POST.get('mileage', car.mileage or 0))
            except (ValueError, TypeError):
                messages.error(request, 'Please enter valid numeric values.')
                return redirect('core:vendor_car_edit', car_id=car_id)

            car.fuel_type = request.POST.get('fuel_type', car.fuel_type)
            car.transmission = request.POST.get('transmission', car.transmission)
            car.save()

            messages.success(request, 'Car listing updated successfully!')

            # Handle HTMX requests
            if request.headers.get('HX-Request'):
                return redirect('core:vendor_listings')

            return redirect('core:vendor_listings')

        # Get form data for GET requests
        context = {
            'car': car,
            'car_brands': CarBrand.objects.filter(is_active=True).order_by('name'),
            'vehicle_conditions': VehicleCondition.objects.filter(is_active=True).order_by('display_order'),
        }

        # Handle HTMX requests for modal
        if request.headers.get('HX-Request'):
            return render(request, 'core/dashboard/partials/vendor_car_edit_modal.html', context)

        return render(request, 'core/dashboard/vendor_car_edit.html', context)

    except Exception as e:
        messages.error(request, 'Error updating car listing.')
        return redirect('core:vendor_listings')


@login_required
def vendor_car_delete(request, car_id):
    """Delete car listing"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'DELETE':
        try:
            car = get_object_or_404(Car, id=car_id, vendor__user=request.user)
            car_title = car.title
            car.delete()

            messages.success(request, f'Car listing "{car_title}" deleted successfully!')

            if request.headers.get('HX-Request'):
                # Return updated car list for HTMX
                return redirect('core:vendor_listings')

            return JsonResponse({'success': True, 'message': 'Car deleted successfully'})

        except Exception as e:
            return JsonResponse({'error': 'Error deleting car'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def vendor_car_toggle_status(request, car_id):
    """Toggle car listing status (active/inactive)"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            car = get_object_or_404(Car, id=car_id, vendor__user=request.user)
            car.is_active = not car.is_active
            car.save()

            status = 'activated' if car.is_active else 'deactivated'
            messages.success(request, f'Car listing {status} successfully!')

            return JsonResponse({
                'success': True,
                'is_active': car.is_active,
                'message': f'Car {status} successfully'
            })

        except Exception as e:
            return JsonResponse({'error': 'Error updating car status'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def vendor_car_feature_toggle(request, car_id):
    """Toggle car featured status"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            car = get_object_or_404(Car, id=car_id, vendor__user=request.user)

            # Check if vendor can feature cars (based on subscription or limits)
            vendor = request.user.vendor
            featured_count = Car.objects.filter(vendor=vendor, is_featured=True).count()

            if not car.is_featured and featured_count >= 3:  # Limit to 3 featured cars
                return JsonResponse({
                    'error': 'You can only feature up to 3 cars. Please unfeature another car first.'
                }, status=400)

            car.is_featured = not car.is_featured
            car.save()

            status = 'featured' if car.is_featured else 'unfeatured'
            messages.success(request, f'Car listing {status} successfully!')

            return JsonResponse({
                'success': True,
                'is_featured': car.is_featured,
                'message': f'Car {status} successfully'
            })

        except Exception as e:
            return JsonResponse({'error': 'Error updating featured status'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def vendor_car_duplicate(request, car_id):
    """Duplicate car listing"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            original_car = get_object_or_404(Car, id=car_id, vendor__user=request.user)

            # Create a duplicate
            duplicate_car = Car.objects.create(
                vendor=original_car.vendor,
                title=f"{original_car.title} (Copy)",
                description=original_car.description,
                price=original_car.price,
                year=original_car.year,
                mileage=original_car.mileage,
                fuel_type=original_car.fuel_type,
                transmission=original_car.transmission,
                brand=original_car.brand,
                model=original_car.model,
                condition=original_car.condition,
                status='available',
                is_approved=False,  # Requires re-approval
                is_featured=False,  # Not featured by default
                is_hot_deal=False,
            )

            messages.success(request, f'Car listing duplicated successfully! "{duplicate_car.title}" has been created and is pending approval.')

            # Handle HTMX requests
            if request.headers.get('HX-Request'):
                return redirect('core:vendor_listings')

            return JsonResponse({'success': True, 'message': 'Car duplicated successfully', 'new_car_id': duplicate_car.id})

        except Exception as e:
            return JsonResponse({'error': 'Error duplicating car'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def vendor_bulk_actions(request):
    """Handle bulk actions on car listings"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            action = request.POST.get('action')
            car_ids = request.POST.getlist('car_ids')

            if not car_ids:
                return JsonResponse({'error': 'No cars selected'}, status=400)

            cars = Car.objects.filter(id__in=car_ids, vendor__user=request.user)

            if action == 'activate':
                cars.update(is_active=True)
                messages.success(request, f'{cars.count()} cars activated successfully!')

            elif action == 'deactivate':
                cars.update(is_active=False)
                messages.success(request, f'{cars.count()} cars deactivated successfully!')

            elif action == 'delete':
                count = cars.count()
                cars.delete()
                messages.success(request, f'{count} cars deleted successfully!')

            elif action == 'feature':
                # Check feature limits
                vendor = request.user.vendor
                current_featured = Car.objects.filter(vendor=vendor, is_featured=True).count()
                can_feature = min(len(car_ids), 3 - current_featured)

                if can_feature > 0:
                    cars[:can_feature].update(is_featured=True)
                    messages.success(request, f'{can_feature} cars featured successfully!')
                else:
                    messages.warning(request, 'Feature limit reached. You can only feature up to 3 cars.')

            return JsonResponse({'success': True, 'message': 'Bulk action completed'})

        except Exception as e:
            return JsonResponse({'error': 'Error performing bulk action'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
def vendor_export_listings(request):
    """Export vendor car listings to CSV"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        import csv
        from django.http import HttpResponse

        vendor = request.user.vendor
        cars = Car.objects.filter(vendor=vendor).select_related('brand', 'model', 'condition')

        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="car_listings_{vendor.company_name}_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)

        # Write header
        writer.writerow([
            'ID', 'Title', 'Brand', 'Model', 'Year', 'Price (KSH)',
            'Mileage', 'Fuel Type', 'Transmission', 'Status',
            'Approved', 'Featured', 'Views', 'Created Date'
        ])

        # Write car data
        for car in cars:
            writer.writerow([
                car.id,
                car.title,
                car.brand.name if car.brand else '',
                car.model.name if car.model else '',
                car.year,
                car.price,
                car.mileage,
                car.fuel_type,
                car.transmission,
                car.status,
                'Yes' if car.is_approved else 'No',
                'Yes' if car.is_featured else 'No',
                car.views_count or 0,
                car.created_at.strftime('%Y-%m-%d %H:%M')
            ])

        return response

    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error exporting listings: {str(e)}")
        return JsonResponse({'error': 'Error exporting data'}, status=500)


@login_required
def vendor_inquiries_view(request):
    """Vendor inquiries management view with HTMX support"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor
        inquiries = Inquiry.objects.filter(car__vendor=vendor).order_by('-created_at')

        # Apply filters
        status_filter = request.GET.get('status_filter')
        car_filter = request.GET.get('car_filter')
        search = request.GET.get('search')

        if status_filter:
            inquiries = inquiries.filter(status=status_filter)

        if car_filter:
            inquiries = inquiries.filter(car_id=car_filter)

        if search:
            inquiries = inquiries.filter(
                Q(subject__icontains=search) |
                Q(message__icontains=search) |
                Q(customer__first_name__icontains=search) |
                Q(customer__last_name__icontains=search) |
                Q(customer__email__icontains=search)
            )

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(inquiries, 10)
        page_number = request.GET.get('page')
        inquiries_page = paginator.get_page(page_number)

        context = {
            'vendor': vendor,
            'inquiries': inquiries_page,
            'vendor_cars': Car.objects.filter(vendor=vendor),
            'open_inquiries': Inquiry.objects.filter(car__vendor=vendor, status='open').count(),
            'in_progress_inquiries': Inquiry.objects.filter(car__vendor=vendor, status='in_progress').count(),
            'resolved_inquiries': Inquiry.objects.filter(car__vendor=vendor, status='resolved').count(),
        }

        # Return partial template for HTMX requests
        if request.headers.get('HX-Request'):
            return render(request, 'core/dashboard/partials/inquiry_list.html', context)

        return render(request, 'core/dashboard/vendor_inquiries.html', context)

    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
def vendor_analytics_view(request):
    """Vendor analytics and reporting view"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')
    
    try:
        vendor = request.user.vendor
        cars = Car.objects.filter(vendor=vendor)
        
        # Calculate analytics
        total_views = sum(car.views_count for car in cars)
        total_inquiries = Inquiry.objects.filter(car__vendor=vendor).count()
        
        # Monthly data (last 6 months)
        monthly_data = []
        for i in range(6):
            month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
            month_end = month_start + timedelta(days=30)
            
            month_cars = cars.filter(created_at__range=[month_start, month_end])
            month_inquiries = Inquiry.objects.filter(
                car__vendor=vendor,
                created_at__range=[month_start, month_end]
            )
            
            monthly_data.append({
                'month': month_start.strftime('%B'),
                'cars_added': month_cars.count(),
                'inquiries': month_inquiries.count(),
                'views': sum(car.views_count for car in month_cars),
            })
        
        context = {
            'vendor': vendor,
            'total_cars': cars.count(),
            'total_views': total_views,
            'total_inquiries': total_inquiries,
            'monthly_data': monthly_data,
            'top_cars': cars.order_by('-views_count')[:5],
        }
        
        return render(request, 'core/dashboard/vendor_analytics.html', context)
        
    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
def admin_users_view(request):
    """Enhanced admin user management view with pagination, filters, and export"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Base queryset
    users = User.objects.select_related().order_by('-date_joined')

    # Filter by role if specified
    role_filter = request.GET.get('role', '')
    if role_filter and role_filter != 'all':
        users = users.filter(role=role_filter)

    # Filter by status if specified
    status_filter = request.GET.get('status', '')
    if status_filter and status_filter != 'all':
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)
        elif status_filter == 'verified':
            users = users.filter(is_verified=True)
        elif status_filter == 'unverified':
            users = users.filter(is_verified=False)

    # Search functionality
    search = request.GET.get('search', '').strip()
    if search:
        users = users.filter(
            Q(username__icontains=search) |
            Q(first_name__icontains=search) |
            Q(last_name__icontains=search) |
            Q(email__icontains=search) |
            Q(phone__icontains=search)
        )

    # Handle export requests
    export_format = request.GET.get('export')
    if export_format in ['csv', 'excel']:
        return export_users_data(users, export_format)

    # Enhanced statistics
    total_users = User.objects.count()
    active_users = User.objects.filter(is_active=True).count()
    verified_users = User.objects.filter(is_verified=True).count()
    customers = User.objects.filter(role='customer').count()
    vendors = User.objects.filter(role='vendor').count()
    admins = User.objects.filter(role='admin').count()

    # Recent registrations (last 30 days)
    from datetime import datetime, timedelta
    thirty_days_ago = datetime.now() - timedelta(days=30)
    recent_registrations = User.objects.filter(date_joined__gte=thirty_days_ago).count()

    # Pagination - 20 users per page
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    paginator = Paginator(users, 20)
    page = request.GET.get('page', 1)

    try:
        users_page = paginator.page(page)
    except PageNotAnInteger:
        users_page = paginator.page(1)
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)

    context = {
        'users': users_page,
        'total_users': total_users,
        'active_users': active_users,
        'verified_users': verified_users,
        'customers': customers,
        'vendors': vendors,
        'admins': admins,
        'recent_registrations': recent_registrations,
        'current_role_filter': role_filter,
        'current_status_filter': status_filter,
        'search_query': search,
        'paginator': paginator,
        'page_obj': users_page,
    }

    # HTMX partial template for table updates
    if request.headers.get('HX-Request'):
        return render(request, 'core/dashboard/partials/admin_users_table.html', context)

    return render(request, 'core/dashboard/admin_users.html', context)


def export_users_data(users_queryset, format_type):
    """Export users data to CSV or Excel format"""
    import csv
    from django.http import HttpResponse
    from datetime import datetime

    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'users_export_{timestamp}'

    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'ID', 'Username', 'First Name', 'Last Name', 'Email', 'Phone',
            'Role', 'Status', 'Verified', 'Date Joined', 'Last Login'
        ])

        for user in users_queryset:
            writer.writerow([
                user.id,
                user.username,
                user.first_name,
                user.last_name,
                user.email,
                user.phone,
                user.get_role_display(),
                'Active' if user.is_active else 'Inactive',
                'Yes' if user.is_verified else 'No',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never'
            ])

        return response

    elif format_type == 'excel':
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, PatternFill

            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Users Export'

            # Headers
            headers = [
                'ID', 'Username', 'First Name', 'Last Name', 'Email', 'Phone',
                'Role', 'Status', 'Verified', 'Date Joined', 'Last Login'
            ]

            # Style headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill

            # Data rows
            for row, user in enumerate(users_queryset, 2):
                ws.cell(row=row, column=1, value=user.id)
                ws.cell(row=row, column=2, value=user.username)
                ws.cell(row=row, column=3, value=user.first_name)
                ws.cell(row=row, column=4, value=user.last_name)
                ws.cell(row=row, column=5, value=user.email)
                ws.cell(row=row, column=6, value=user.phone)
                ws.cell(row=row, column=7, value=user.get_role_display())
                ws.cell(row=row, column=8, value='Active' if user.is_active else 'Inactive')
                ws.cell(row=row, column=9, value='Yes' if user.is_verified else 'No')
                ws.cell(row=row, column=10, value=user.date_joined.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row=row, column=11, value=user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never')

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].auto_size = True

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            wb.save(response)
            return response

        except ImportError:
            # Fallback to CSV if openpyxl is not available
            return export_users_data(users_queryset, 'csv')


@login_required
def admin_users_search(request):
    """HTMX endpoint for real-time user search"""
    if request.user.role != 'admin':
        return HttpResponse('Unauthorized', status=403)

    search_query = request.GET.get('search', '').strip()
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    users = User.objects.select_related().order_by('-date_joined')

    # Apply filters
    if role_filter and role_filter != 'all':
        users = users.filter(role=role_filter)

    if status_filter and status_filter != 'all':
        if status_filter == 'active':
            users = users.filter(is_active=True)
        elif status_filter == 'inactive':
            users = users.filter(is_active=False)
        elif status_filter == 'verified':
            users = users.filter(is_verified=True)
        elif status_filter == 'unverified':
            users = users.filter(is_verified=False)

    # Apply search
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

    # Pagination
    from django.core.paginator import Paginator
    paginator = Paginator(users, 20)
    page = request.GET.get('page', 1)
    users_page = paginator.page(page)

    context = {
        'users': users_page,
        'search_query': search_query,
        'current_role_filter': role_filter,
        'current_status_filter': status_filter,
    }

    return render(request, 'core/dashboard/partials/admin_users_table.html', context)


@login_required
def admin_users_refresh(request):
    """HTMX endpoint for refreshing user list"""
    if request.user.role != 'admin':
        return HttpResponse('Unauthorized', status=403)

    # Get current filters
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    # Redirect to main view with current filters
    from urllib.parse import urlencode

    params = {}
    if role_filter:
        params['role'] = role_filter
    if status_filter:
        params['status'] = status_filter
    if search_query:
        params['search'] = search_query

    url = reverse('core:admin_users')
    if params:
        url += '?' + urlencode(params)

    response = HttpResponse()
    response['HX-Redirect'] = url
    return response


@login_required
def admin_vendors_view(request):
    """Admin vendor management view"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('core:dashboard')

    vendors = Vendor.objects.select_related('user').order_by('-created_at')

    # Filter by approval status
    status_filter = request.GET.get('status')
    if status_filter == 'approved':
        vendors = vendors.filter(is_approved=True)
    elif status_filter == 'pending':
        vendors = vendors.filter(is_approved=False)
    elif status_filter == 'suspended':
        vendors = vendors.filter(user__is_active=False)

    # Filter by business type
    business_type_filter = request.GET.get('business_type')
    if business_type_filter:
        vendors = vendors.filter(business_type=business_type_filter)

    # Search functionality
    search = request.GET.get('search', '')
    if search:
        vendors = vendors.filter(
            Q(company_name__icontains=search) |
            Q(user__first_name__icontains=search) |
            Q(user__last_name__icontains=search) |
            Q(user__email__icontains=search) |
            Q(business_license__icontains=search)
        )

    # Pagination - 20 vendors per page
    paginator = Paginator(vendors, 20)
    page = request.GET.get('page')
    try:
        vendors = paginator.page(page)
    except PageNotAnInteger:
        vendors = paginator.page(1)
    except EmptyPage:
        vendors = paginator.page(paginator.num_pages)

    context = {
        'vendors': vendors,
        'total_vendors': Vendor.objects.count(),
        'approved_vendors': Vendor.objects.filter(is_approved=True).count(),
        'pending_vendors': Vendor.objects.filter(is_approved=False).count(),
        'suspended_vendors': Vendor.objects.filter(user__is_active=False).count(),
        'current_filter': status_filter,
        'business_type_filter': business_type_filter,
        'search_query': search,
    }

    return render(request, 'core/dashboard/admin_vendors.html', context)


@login_required
def admin_listings_view(request):
    """Admin car listings management view with pagination"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    cars = Car.objects.select_related('brand', 'model', 'vendor', 'vendor__user').order_by('-created_at')

    # Filter by approval status
    status_filter = request.GET.get('status')
    if status_filter == 'approved':
        cars = cars.filter(is_approved=True)
    elif status_filter == 'pending':
        cars = cars.filter(is_approved=False)
    elif status_filter == 'featured':
        cars = cars.filter(is_featured=True)
    elif status_filter == 'hot_deals':
        cars = cars.filter(is_hot_deal=True)
    elif status_filter == 'sold':
        cars = cars.filter(status='sold')

    # Search functionality
    search = request.GET.get('search', '')
    if search:
        cars = cars.filter(
            Q(title__icontains=search) |
            Q(brand__name__icontains=search) |
            Q(model__name__icontains=search) |
            Q(vendor__company_name__icontains=search) |
            Q(vendor__user__first_name__icontains=search) |
            Q(vendor__user__last_name__icontains=search)
        )

    # Pagination - 20 cars per page
    paginator = Paginator(cars, 20)
    page = request.GET.get('page')

    try:
        cars = paginator.page(page)
    except PageNotAnInteger:
        cars = paginator.page(1)
    except EmptyPage:
        cars = paginator.page(paginator.num_pages)

    context = {
        'cars': cars,
        'total_cars': Car.objects.count(),
        'approved_cars': Car.objects.filter(is_approved=True).count(),
        'pending_cars': Car.objects.filter(is_approved=False).count(),
        'featured_cars': Car.get_featured_cars_count(),
        'featured_remaining': Car.get_featured_cars_remaining(),
        'current_filter': status_filter,
        'search': search,
    }

    return render(request, 'core/dashboard/admin_listings.html', context)


@login_required
def admin_car_detail_view(request, car_id):
    """Admin car detail view"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    car = get_object_or_404(Car.objects.select_related('brand', 'model', 'vendor', 'vendor__user'), id=car_id)

    # Get car images
    car_images = car.images.all().order_by('order')

    # Get car inquiries
    car_inquiries = Inquiry.objects.filter(car=car).select_related('user').order_by('-created_at')[:5]

    context = {
        'car': car,
        'car_images': car_images,
        'car_inquiries': car_inquiries,
        'features_list': car.get_features_list(),
    }

    return render(request, 'core/dashboard/admin_car_detail.html', context)


@login_required
def admin_car_edit_view(request, car_id):
    """Enhanced admin car edit modal view with comprehensive field handling"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    car = get_object_or_404(Car.objects.select_related('brand', 'model', 'condition', 'vendor__user'), id=car_id)
    original_approval_status = car.is_approved

    if request.method == 'POST':
        form = AdminCarEditForm(request.POST, request.FILES, instance=car)

        if form.is_valid():
            # Save the main form
            updated_car = form.save(commit=False)

            # Handle approval status change
            if updated_car.is_approved != original_approval_status:
                if updated_car.is_approved:
                    updated_car.approval_date = timezone.now()
                else:
                    updated_car.approval_date = None

            # Handle featured until date
            featured_until = request.POST.get('featured_until')
            if featured_until:
                try:
                    from datetime import datetime
                    updated_car.featured_until = datetime.fromisoformat(featured_until.replace('T', ' '))
                except ValueError:
                    pass

            # Save the car
            updated_car.save()

            # Handle hot deals creation/update with enhanced validation
            hot_deal_errors = []
            try:
                if updated_car.is_hot_deal:
                    hot_deal_discount = request.POST.get('hot_deal_discount', '10')
                    hot_deal_days = request.POST.get('hot_deal_days', '7')

                    # Validate and convert values with proper error handling
                    try:
                        discount_value = float(hot_deal_discount)
                        days_value = int(hot_deal_days)

                        # Enhanced validation with specific error messages
                        if discount_value < 5 or discount_value > 50:
                            hot_deal_errors.append("Discount percentage must be between 5% and 50%")
                            discount_value = max(5, min(50, discount_value))  # Clamp to valid range

                        if days_value < 1 or days_value > 30:
                            hot_deal_errors.append("Hot deal duration must be between 1 and 30 days")
                            days_value = max(1, min(30, days_value))  # Clamp to valid range

                        from datetime import timedelta

                        # Get or create hot deal with better error handling
                        hot_deal, created = HotDeal.objects.get_or_create(
                            car=updated_car,
                            defaults={
                                'title': f'Hot Deal: {updated_car.title}',
                                'description': f'Limited time offer - {discount_value}% off!',
                                'discount_type': 'percentage',
                                'discount_value': discount_value,
                                'original_price': updated_car.price,
                                'start_date': timezone.now(),
                                'end_date': timezone.now() + timedelta(days=days_value),
                                'is_active': True,
                                'auto_activate': True
                            }
                        )

                        if not created:
                            # Update existing hot deal with all relevant fields
                            hot_deal.title = f'Hot Deal: {updated_car.title}'
                            hot_deal.description = f'Limited time offer - {discount_value}% off!'
                            hot_deal.discount_value = discount_value
                            hot_deal.original_price = updated_car.price
                            hot_deal.start_date = timezone.now()
                            hot_deal.end_date = timezone.now() + timedelta(days=days_value)
                            hot_deal.is_active = True
                            hot_deal.auto_activate = True
                            hot_deal.save()

                        # Log successful hot deal creation/update
                        print(f"Hot deal {'created' if created else 'updated'} for car {updated_car.title}: {discount_value}% off for {days_value} days")

                    except (ValueError, TypeError) as e:
                        hot_deal_errors.append(f"Invalid hot deal values: {str(e)}")
                        # Still mark as hot deal but without creating HotDeal object
                        print(f"Hot deal validation error: {e}")
                else:
                    # Remove hot deal if unchecked
                    deactivated_count = HotDeal.objects.filter(car=updated_car, is_active=True).update(is_active=False)
                    if deactivated_count > 0:
                        print(f"Deactivated hot deal for car {updated_car.title}")

            except Exception as e:
                hot_deal_errors.append(f"Hot deal processing error: {str(e)}")
                print(f"Hot deal processing error: {e}")
                # Don't fail the entire form submission

            # If approval status changed, update approval date
            if updated_car.is_approved and not original_approval_status:
                updated_car.approval_date = timezone.now()
                updated_car.save()

            # Create success message
            success_message = f'Car listing "{updated_car.title}" has been updated successfully.'

            # Add approval status change notification
            if updated_car.is_approved and not original_approval_status:
                success_message += ' Car is now approved and visible to customers.'
            elif not updated_car.is_approved and original_approval_status:
                success_message += ' Car approval has been revoked.'

            # Prepare comprehensive JSON response for HTMX
            response_data = {
                'status': 'success',
                'message': success_message,
                'car_id': updated_car.id,
                'car_title': updated_car.title,
                'is_approved': updated_car.is_approved,
                'is_featured': updated_car.is_featured,
                'is_certified': updated_car.is_certified,
                'is_hot_deal': updated_car.is_hot_deal,
                'hot_deal_errors': hot_deal_errors if hot_deal_errors else [],
                'reload_required': True,
                'timestamp': timezone.now().isoformat(),
                'updated_fields': list(form.changed_data) if hasattr(form, 'changed_data') else []
            }

            # Add hot deal details if applicable
            if updated_car.is_hot_deal and hasattr(updated_car, 'hot_deal_details'):
                try:
                    hot_deal = updated_car.hot_deal_details
                    response_data['hot_deal_info'] = {
                        'discount_value': float(hot_deal.discount_value),
                        'discounted_price': float(hot_deal.discounted_price),
                        'original_price': float(hot_deal.original_price),
                        'end_date': hot_deal.end_date.isoformat(),
                        'days_remaining': hot_deal.days_remaining(),
                        'is_active': hot_deal.is_currently_active(),
                        'title': hot_deal.title
                    }
                except Exception as e:
                    response_data['hot_deal_errors'].append(f"Hot deal info error: {str(e)}")

            # Always return JSON response for HTMX
            return JsonResponse(response_data)


        else:
            # Return form with errors for HTMX
            if request.headers.get('HX-Request'):
                # Prepare detailed error information
                error_details = {}
                for field, errors in form.errors.items():
                    error_details[field] = [str(error) for error in errors]

                error_response = {
                    'status': 'error',
                    'message': 'Please correct the validation errors below.',
                    'errors': error_details,
                    'hot_deal_errors': hot_deal_errors if 'hot_deal_errors' in locals() else [],
                    'form_data': dict(request.POST),
                    'timestamp': timezone.now().isoformat()
                }


                return JsonResponse(error_response, status=400)
    else:
        form = AdminCarEditForm(instance=car)

    # Prepare context with all necessary data
    context = {
        'form': form,
        'car': car,
        'car_brands': CarBrand.objects.filter(is_active=True).order_by('name'),
        'car_models': CarModel.objects.filter(is_active=True).order_by('name'),
        'vehicle_conditions': VehicleCondition.objects.filter(is_active=True).order_by('display_order'),
    }

    return render(request, 'core/modals/admin_car_edit.html', context)


@login_required
def admin_feature_car(request, car_id):
    """Enhanced admin feature/unfeature car action with tier support and limits"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    car = get_object_or_404(Car, id=car_id)
    action = request.POST.get('action', 'feature')

    try:
        if action == 'feature':
            success, message = car.feature_car()
            if success:
                # Update calculated rating for better visibility
                if car.calculated_rating < 4.0:
                    car.calculated_rating = 4.0
                    car.save(update_fields=['calculated_rating'])

                return JsonResponse({
                    'status': 'success',
                    'message': message,
                    'is_featured': car.is_featured,
                    'featured_count': Car.get_featured_cars_count(),
                    'remaining_slots': Car.get_featured_cars_remaining()
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': message,
                    'featured_count': Car.get_featured_cars_count(),
                    'remaining_slots': Car.get_featured_cars_remaining()
                }, status=400)

        elif action == 'unfeature':
            success, message = car.unfeature_car()
            if success:
                return JsonResponse({
                    'status': 'success',
                    'message': message,
                    'is_featured': car.is_featured,
                    'featured_count': Car.get_featured_cars_count(),
                    'remaining_slots': Car.get_featured_cars_remaining()
                })
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': message
                }, status=400)

        elif action == 'hot_deal':
            car.is_hot_deal = True
            car.save(update_fields=['is_hot_deal'])
            message = f'Car "{car.title}" has been marked as a hot deal.'
            return JsonResponse({
                'status': 'success',
                'message': message,
                'is_hot_deal': car.is_hot_deal
            })

        elif action == 'remove_hot_deal':
            car.is_hot_deal = False
            car.save(update_fields=['is_hot_deal'])
            message = f'Hot deal status removed from "{car.title}".'
            return JsonResponse({
                'status': 'success',
                'message': message,
                'is_hot_deal': car.is_hot_deal
            })
        else:
            return JsonResponse({'error': 'Invalid action'}, status=400)

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }, status=500)


@login_required
def admin_hot_deal_management(request, car_id):
    """Enhanced admin hot deal management with comprehensive functionality"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    car = get_object_or_404(Car, id=car_id)

    try:
        import json
        from datetime import timedelta
        from django.utils import timezone
        from .models import HotDeal

        data = json.loads(request.body)
        action = data.get('action')

        if action == 'create':
            # Check if car already has an active hot deal
            if HotDeal.objects.filter(car=car, is_active=True).exists():
                return JsonResponse({
                    'status': 'error',
                    'message': 'This car already has an active hot deal.'
                })

            discount_value = float(data.get('discount_value', 10))
            days = int(data.get('days', 7))

            # Validate discount value
            if discount_value < 5 or discount_value > 50:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Discount must be between 5% and 50%.'
                })

            # Create hot deal
            hot_deal = HotDeal.objects.create(
                car=car,
                title=f'Hot Deal: {car.title}',
                description=f'Limited time offer - {discount_value}% off!',
                discount_type='percentage',
                discount_value=discount_value,
                original_price=car.price,
                start_date=timezone.now(),
                end_date=timezone.now() + timedelta(days=days),
                is_active=True,
                auto_activate=True
            )

            # Update car status
            car.is_hot_deal = True
            car.save(update_fields=['is_hot_deal'])

            return JsonResponse({
                'status': 'success',
                'message': f'Hot deal created successfully! {discount_value}% off for {days} days.',
                'deal_id': hot_deal.id
            })

        elif action == 'end':
            try:
                hot_deal = HotDeal.objects.get(car=car, is_active=True)
                hot_deal.is_active = False
                hot_deal.end_date = timezone.now()
                hot_deal.save()

                # Update car status
                car.is_hot_deal = False
                car.price = hot_deal.original_price
                car.save(update_fields=['is_hot_deal', 'price'])

                return JsonResponse({
                    'status': 'success',
                    'message': 'Hot deal ended successfully.'
                })
            except HotDeal.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No active hot deal found for this car.'
                })

        elif action == 'extend':
            try:
                hot_deal = HotDeal.objects.get(car=car, is_active=True)
                additional_days = int(data.get('days', 7))
                hot_deal.end_date = hot_deal.end_date + timedelta(days=additional_days)
                hot_deal.save()

                return JsonResponse({
                    'status': 'success',
                    'message': f'Hot deal extended by {additional_days} days.'
                })
            except HotDeal.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No active hot deal found for this car.'
                })

        elif action == 'edit':
            try:
                hot_deal = HotDeal.objects.get(car=car, is_active=True)
                new_discount = float(data.get('discount_value', hot_deal.discount_value))

                if new_discount < 5 or new_discount > 50:
                    return JsonResponse({
                        'status': 'error',
                        'message': 'Discount must be between 5% and 50%.'
                    })

                hot_deal.discount_value = new_discount
                hot_deal.title = f'Hot Deal: {car.title}'
                hot_deal.description = f'Limited time offer - {new_discount}% off!'
                hot_deal.save()

                return JsonResponse({
                    'status': 'success',
                    'message': f'Hot deal updated to {new_discount}% discount.'
                })
            except HotDeal.DoesNotExist:
                return JsonResponse({
                    'status': 'error',
                    'message': 'No active hot deal found for this car.'
                })

        else:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid action specified.'
            })

    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred: {str(e)}'
        }, status=500)


# Sub-categories are now hardcoded in templates - no dynamic loading endpoint needed


@login_required
def admin_car_delete_view(request, car_id):
    """Admin car deletion view with proper cleanup"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    try:
        car = get_object_or_404(Car, id=car_id)
        car_title = car.title
        was_featured = car.is_featured

        # Store information before deletion
        vendor = car.vendor

        # Delete associated images first
        car_images = car.images.all()
        for image in car_images:
            if image.image and hasattr(image.image, 'path'):
                try:
                    import os
                    if os.path.exists(image.image.path):
                        os.remove(image.image.path)
                except Exception as e:
                    print(f"Error deleting image file: {e}")

        # Delete main image if exists
        if car.main_image and hasattr(car.main_image, 'path'):
            try:
                import os
                if os.path.exists(car.main_image.path):
                    os.remove(car.main_image.path)
            except Exception as e:
                print(f"Error deleting main image file: {e}")

        # Delete the car (this will cascade delete related objects)
        car.delete()

        # Get updated featured car stats
        featured_count = Car.get_featured_cars_count()
        remaining_slots = Car.get_featured_cars_remaining()

        # Log the deletion
        print(f"Admin {request.user.username} deleted car: {car_title} (ID: {car_id})")

        return JsonResponse({
            'status': 'success',
            'message': f'Car "{car_title}" has been permanently deleted',
            'car_id': car_id,
            'car_title': car_title,
            'was_featured': was_featured,
            'featured_count': featured_count,
            'remaining_slots': remaining_slots,
            'vendor_id': vendor.id if vendor else None
        })

    except Car.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'Car not found'
        }, status=404)
    except Exception as e:
        print(f"Error deleting car {car_id}: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': f'An error occurred while deleting the car: {str(e)}'
        }, status=500)


@login_required
def admin_analytics_view(request):
    """Enhanced admin analytics and reporting view with promotion metrics"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from .analytics_utils import PromotionAnalyticsManager

    # Get date range from request
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)

    # Initialize analytics manager
    analytics = PromotionAnalyticsManager()

    # Calculate system-wide analytics
    total_users = User.objects.count()
    total_cars = Car.objects.count()
    total_vendors = Vendor.objects.count()
    total_inquiries = Inquiry.objects.count()

    # Promotion metrics
    featured_cars_count = Car.objects.filter(is_featured=True).count()
    active_hot_deals = HotDeal.objects.filter(is_active=True).count()
    avg_rating = CarRating.objects.filter(
        created_at__date__gte=start_date,
        is_approved=True
    ).aggregate(avg_rating=Avg('rating'))['avg_rating'] or 0

    # Get promotion analytics
    featured_performance = analytics.get_featured_cars_performance(days)
    hot_deals_performance = analytics.get_hot_deals_performance(days)
    rating_distribution = analytics.get_rating_distribution(days)
    featured_comparison = analytics.get_featured_comparison()
    daily_metrics = analytics.get_daily_metrics(days)
    conversion_funnel = analytics.get_conversion_funnel(days)

    # Monthly growth data
    monthly_data = []
    for i in range(6):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)

        month_users = User.objects.filter(date_joined__range=[month_start, month_end])
        month_cars = Car.objects.filter(created_at__range=[month_start, month_end])
        month_inquiries = Inquiry.objects.filter(created_at__range=[month_start, month_end])

        monthly_data.append({
            'month': month_start.strftime('%B'),
            'users': month_users.count(),
            'cars': month_cars.count(),
            'inquiries': month_inquiries.count(),
        })

    # Top performing vendors with promotion metrics
    vendor_performance = Vendor.objects.filter(
        is_approved=True
    ).annotate(
        car_count=Count('cars', filter=Q(cars__is_approved=True)),
        featured_count=Count('cars', filter=Q(cars__is_featured=True)),
        avg_rating=Avg('cars__calculated_rating', filter=Q(cars__is_approved=True))
    ).order_by('-featured_count', '-car_count')[:10]

    context = {
        'total_users': total_users,
        'total_cars': total_cars,
        'total_vendors': total_vendors,
        'total_inquiries': total_inquiries,
        'featured_cars_count': featured_cars_count,
        'active_hot_deals': active_hot_deals,
        'avg_rating': avg_rating,
        'featured_performance': featured_performance,
        'hot_deals_performance': hot_deals_performance,
        'rating_distribution': rating_distribution,
        'featured_comparison': featured_comparison,
        'daily_metrics': daily_metrics,
        'conversion_funnel': conversion_funnel,
        'monthly_data': monthly_data,
        'vendor_performance': vendor_performance,
        'top_brands': CarBrand.objects.annotate(
            car_count=Count('car')
        ).order_by('-car_count')[:5],
        'days': days,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'core/dashboard/admin_analytics.html', context)


# Enhanced Image Gallery Management Views

@login_required
@require_http_methods(["POST"])
def car_image_upload(request, car_id):
    """Upload new images to car gallery via HTMX"""
    car = get_object_or_404(Car, id=car_id)

    # Check permissions
    if request.user.role == 'vendor' and car.vendor.user != request.user:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    elif request.user.role not in ['admin', 'vendor']:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if 'images' not in request.FILES:
        return JsonResponse({'error': 'No images provided'}, status=400)

    uploaded_images = []
    errors = []

    # Get the highest order number for new images
    max_order = car.images.aggregate(models.Max('order'))['order__max'] or 0

    for i, image_file in enumerate(request.FILES.getlist('images')):
        try:
            # Use image optimization validation
            from .utils.image_optimization import validate_car_image

            # Validate image
            is_valid, error_msg = validate_car_image(image_file)
            if not is_valid:
                errors.append(f"Image {image_file.name}: {error_msg}")
                continue

            # Create CarImage instance (optimization happens in model save)
            car_image = CarImage.objects.create(
                car=car,
                image=image_file,
                caption=request.POST.get(f'caption_{i}', ''),
                order=max_order + i + 1
            )

            uploaded_images.append({
                'id': car_image.id,
                'url': car_image.image.url,
                'caption': car_image.caption,
                'order': car_image.order,
                'is_primary': car_image.is_primary
            })

        except Exception as e:
            errors.append(f"Error uploading {image_file.name}: {str(e)}")

    if request.headers.get('HX-Request'):
        # Return updated gallery HTML
        context = {
            'car': car,
            'car_images': car.images.all().order_by('order'),
            'uploaded_count': len(uploaded_images),
            'errors': errors
        }
        return render(request, 'core/partials/car_image_gallery.html', context)

    return JsonResponse({
        'success': True,
        'uploaded_images': uploaded_images,
        'errors': errors
    })


@login_required
@require_http_methods(["POST"])
def car_image_delete(request, car_id, image_id):
    """Delete a car image via HTMX"""
    car = get_object_or_404(Car, id=car_id)
    image = get_object_or_404(CarImage, id=image_id, car=car)

    # Check permissions
    if request.user.role == 'vendor' and car.vendor.user != request.user:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    elif request.user.role not in ['admin', 'vendor']:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        # Delete the image file
        if image.image:
            image.image.delete(save=False)

        # Delete the database record
        image.delete()

        if request.headers.get('HX-Request'):
            # Return updated gallery HTML
            context = {
                'car': car,
                'car_images': car.images.all().order_by('order')
            }
            return render(request, 'core/partials/car_image_gallery.html', context)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def car_image_set_primary(request, car_id, image_id):
    """Set an image as primary via HTMX"""
    car = get_object_or_404(Car, id=car_id)
    image = get_object_or_404(CarImage, id=image_id, car=car)

    # Check permissions
    if request.user.role == 'vendor' and car.vendor.user != request.user:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    elif request.user.role not in ['admin', 'vendor']:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        # Set this image as primary (model handles unsetting others)
        image.is_primary = True
        image.save()

        if request.headers.get('HX-Request'):
            # Return updated gallery HTML
            context = {
                'car': car,
                'car_images': car.images.all().order_by('order')
            }
            return render(request, 'core/partials/car_image_gallery.html', context)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def car_image_reorder(request, car_id):
    """Reorder car images via HTMX"""
    car = get_object_or_404(Car, id=car_id)

    # Check permissions
    if request.user.role == 'vendor' and car.vendor.user != request.user:
        return JsonResponse({'error': 'Permission denied'}, status=403)
    elif request.user.role not in ['admin', 'vendor']:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        # Get the new order from request
        image_orders = request.POST.get('image_orders', '')
        if not image_orders:
            return JsonResponse({'error': 'No order data provided'}, status=400)

        # Parse the order data (expected format: "id1:order1,id2:order2,...")
        orders = {}
        for item in image_orders.split(','):
            if ':' in item:
                img_id, order = item.split(':')
                orders[int(img_id)] = int(order)

        # Update the order for each image
        for image in car.images.all():
            if image.id in orders:
                image.order = orders[image.id]
                image.save(update_fields=['order'])

        if request.headers.get('HX-Request'):
            # Return updated gallery HTML
            context = {
                'car': car,
                'car_images': car.images.all().order_by('order')
            }
            return render(request, 'core/partials/car_image_gallery.html', context)

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def promotion_analytics_api(request):
    """API endpoint for promotion analytics data"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .analytics_utils import PromotionAnalyticsManager

    analytics = PromotionAnalyticsManager()
    metric_type = request.GET.get('metric', 'daily')
    days = int(request.GET.get('days', 30))

    if metric_type == 'daily':
        data = analytics.get_daily_metrics(days)
    elif metric_type == 'featured':
        data = analytics.get_featured_cars_performance(days)
    elif metric_type == 'hot_deals':
        data = analytics.get_hot_deals_performance(days)
    elif metric_type == 'ratings':
        data = analytics.get_rating_distribution(days)
    elif metric_type == 'funnel':
        data = analytics.get_conversion_funnel(days)
    else:
        data = {}

    return JsonResponse(data, safe=False)


# Lazy Loading Views for HTMX
@login_required
def vendor_recent_listings_lazy(request):
    """HTMX endpoint for vendor recent listings"""
    if request.user.role != 'vendor':
        return HttpResponse('Unauthorized', status=403)

    try:
        vendor = request.user.vendor
        vendor_cars = Car.objects.filter(vendor=vendor).order_by('-created_at')[:5]

        context = {
            'vendor_cars': vendor_cars,
        }

        return render(request, 'core/dashboard/partials/recent_listings.html', context)

    except Vendor.DoesNotExist:
        return HttpResponse('<div class="text-center py-8"><p class="text-gray-600">Please complete your vendor profile.</p></div>')


@login_required
def admin_quick_actions_lazy(request):
    """HTMX endpoint for admin quick actions"""
    if request.user.role != 'admin':
        return HttpResponse('Unauthorized', status=403)

    # Get admin stats
    total_users = User.objects.count()
    total_cars = Car.objects.count()
    total_vendors = Vendor.objects.count()
    pending_approvals = Car.objects.filter(is_approved=False).count()

    context = {
        'total_users': total_users,
        'total_cars': total_cars,
        'total_vendors': total_vendors,
        'pending_approvals': pending_approvals,
    }

    return render(request, 'core/dashboard/partials/admin_quick_actions.html', context)


# Spare Parts Dashboard Views


@login_required
def vendor_spare_parts_view(request):
    """Vendor spare parts dashboard"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor
        spare_parts = SparePart.objects.filter(vendor=vendor).select_related('supplier', 'category_new')

        # Apply filters
        search = request.GET.get('search', '')
        category = request.GET.get('category', '')
        stock_status = request.GET.get('stock_status', '')

        if search:
            spare_parts = spare_parts.filter(
                Q(name__icontains=search) |
                Q(sku__icontains=search) |
                Q(part_number__icontains=search)
            )

        if category:
            spare_parts = spare_parts.filter(category_new_id=category)

        if stock_status == 'low':
            spare_parts = [part for part in spare_parts if part.is_low_stock]
        elif stock_status == 'out':
            spare_parts = spare_parts.filter(stock_quantity=0)
        elif stock_status == 'available':
            spare_parts = spare_parts.filter(stock_quantity__gt=0)

        # Calculate stats
        total_parts = spare_parts.count() if isinstance(spare_parts, type(SparePart.objects.all())) else len(spare_parts)
        total_stock_value = sum(part.stock_value for part in spare_parts)
        low_stock_count = len([part for part in spare_parts if hasattr(part, 'is_low_stock') and part.is_low_stock])
        out_of_stock_count = spare_parts.filter(stock_quantity=0).count() if isinstance(spare_parts, type(SparePart.objects.all())) else len([part for part in spare_parts if part.stock_quantity == 0])

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(spare_parts, 20)
        page_number = request.GET.get('page')
        parts_page = paginator.get_page(page_number)

        context = {
            'vendor': vendor,
            'spare_parts': parts_page,
            'total_parts': total_parts,
            'total_stock_value': total_stock_value,
            'low_stock_count': low_stock_count,
            'out_of_stock_count': out_of_stock_count,
            'categories': SparePartCategory.objects.filter(is_active=True),
            'search': search,
            'selected_category': category,
            'selected_stock_status': stock_status,
        }

        return render(request, 'core/dashboard/vendor_spare_parts.html', context)

    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
@require_http_methods(["GET", "POST"])
def vendor_spare_part_add(request):
    """Add new spare part for vendor"""
    print(f"vendor_spare_part_add called with method: {request.method}")
    print(f"User: {request.user}, Role: {getattr(request.user, 'role', 'No role')}")

    if request.user.role != 'vendor':
        print("Access denied - user is not a vendor")
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
        print(f"Vendor found: {vendor}")
    except Vendor.DoesNotExist:
        print("Vendor profile not found")
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    if request.method == 'GET':
        # Return add modal
        categories = SparePartCategory.objects.filter(is_active=True).order_by('name')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        context = {
            'categories': categories,
            'suppliers': suppliers,
        }
        return render(request, 'core/modals/vendor_spare_part_add.html', context)

    elif request.method == 'POST':
        print("POST request received")
        print(f"POST data: {request.POST}")
        print(f"FILES data: {request.FILES}")
        print(f"Content type: {request.content_type}")
        print(f"HX-Request header: {request.headers.get('HX-Request')}")

        form = VendorSparePartForm(request.POST, request.FILES, vendor=vendor)
        print(f"Form created: {form}")
        print(f"Form data: {form.data}")

        if form.is_valid():
            print("Form is valid")
            try:
                spare_part = form.save(commit=False)
                spare_part.vendor = vendor
                spare_part.save()

                # Create initial stock movement
                StockMovement.objects.create(
                    spare_part=spare_part,
                    movement_type='in',
                    reason='initial',
                    quantity=spare_part.stock_quantity,
                    quantity_before=0,
                    quantity_after=spare_part.stock_quantity,
                    notes=f'Initial stock for {spare_part.name}',
                    created_by=request.user
                )

                messages.success(request, f'Spare part "{spare_part.name}" added successfully.')

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': True,
                        'message': f'Spare part "{spare_part.name}" added successfully!',
                        'part_id': spare_part.id
                    })

                return redirect('core:vendor_spare_parts')

            except Exception as e:
                # Handle any database or other errors
                error_message = f'Error saving spare part: {str(e)}'

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': False,
                        'message': error_message,
                        'errors': {'__all__': [error_message]}
                    })

                messages.error(request, error_message)
                return redirect('core:vendor_spare_parts')
        else:
            # Form validation failed
            print("Form validation failed")
            print(f"Form errors: {form.errors}")

            if request.headers.get('HX-Request'):
                # Format errors for JSON response
                formatted_errors = {}
                for field, errors in form.errors.items():
                    if field == '__all__':
                        formatted_errors['__all__'] = [str(error) for error in errors]
                    else:
                        # Get field label for better error messages
                        field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                        formatted_errors[field] = [str(error) for error in errors]

                print(f"Formatted errors: {formatted_errors}")

                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': formatted_errors
                })

            # Non-HTMX request - show errors via messages
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                        messages.error(request, f'{field_label}: {error}')

            return redirect('core:vendor_spare_parts')


@login_required
@require_http_methods(["GET", "POST"])
def vendor_spare_part_edit(request, part_id):
    """Edit spare part for vendor"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
        spare_part = get_object_or_404(SparePart, id=part_id, vendor=vendor)
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    if request.method == 'GET':
        # Return edit modal
        categories = SparePartCategory.objects.filter(is_active=True).order_by('name')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        context = {
            'spare_part': spare_part,
            'categories': categories,
            'suppliers': suppliers,
        }
        return render(request, 'core/modals/vendor_spare_part_edit.html', context)

    elif request.method == 'POST':
        form = VendorSparePartForm(request.POST, request.FILES, instance=spare_part, vendor=vendor)
        if form.is_valid():
            try:
                old_quantity = spare_part.stock_quantity
                spare_part = form.save()

                # Create stock movement if quantity changed
                new_quantity = spare_part.stock_quantity
                if old_quantity != new_quantity:
                    movement_type = 'in' if new_quantity > old_quantity else 'out'
                    quantity_change = abs(new_quantity - old_quantity)

                    StockMovement.objects.create(
                        spare_part=spare_part,
                        movement_type=movement_type,
                        reason='adjustment',
                        quantity=quantity_change,
                        quantity_before=old_quantity,
                        quantity_after=new_quantity,
                        notes=f'Stock adjustment: {old_quantity} → {new_quantity}',
                        created_by=request.user
                    )

                messages.success(request, f'Spare part "{spare_part.name}" updated successfully.')

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': True,
                        'message': f'Spare part "{spare_part.name}" updated successfully!',
                        'part_id': spare_part.id
                    })

                return redirect('core:vendor_spare_parts')

            except Exception as e:
                # Handle any database or other errors
                error_message = f'Error updating spare part: {str(e)}'

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': False,
                        'message': error_message,
                        'errors': {'__all__': [error_message]}
                    })

                messages.error(request, error_message)
                return redirect('core:vendor_spare_parts')
        else:
            # Form validation failed
            if request.headers.get('HX-Request'):
                # Format errors for JSON response
                formatted_errors = {}
                for field, errors in form.errors.items():
                    if field == '__all__':
                        formatted_errors['__all__'] = [str(error) for error in errors]
                    else:
                        # Get field label for better error messages
                        field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                        formatted_errors[field] = [str(error) for error in errors]

                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': formatted_errors
                })

            # Non-HTMX request - show errors via messages
            for field, errors in form.errors.items():
                for error in errors:
                    if field == '__all__':
                        messages.error(request, error)
                    else:
                        field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                        messages.error(request, f'{field_label}: {error}')

            return redirect('core:vendor_spare_parts')


@login_required
@require_http_methods(["GET"])
def vendor_spare_part_view(request, part_id):
    """View spare part details for vendor"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
        spare_part = get_object_or_404(SparePart, id=part_id, vendor=vendor)
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    # Get recent stock movements
    recent_movements = StockMovement.objects.filter(
        spare_part=spare_part
    ).order_by('-created_at')[:10]

    context = {
        'spare_part': spare_part,
        'recent_movements': recent_movements,
    }
    return render(request, 'core/modals/vendor_spare_part_view.html', context)


@login_required
@require_http_methods(["POST"])
def vendor_spare_part_delete(request, part_id):
    """Delete spare part for vendor"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
        spare_part = get_object_or_404(SparePart, id=part_id, vendor=vendor)
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    # Check if part has any orders
    if spare_part.orderitem_set.exists():
        if request.headers.get('HX-Request'):
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete spare part with existing orders.'
            })
        messages.error(request, 'Cannot delete spare part with existing orders.')
        return redirect('core:vendor_spare_parts')

    part_name = spare_part.name
    spare_part.delete()

    messages.success(request, f'Spare part "{part_name}" deleted successfully.')

    if request.headers.get('HX-Request'):
        return JsonResponse({
            'success': True,
            'message': f'Spare part "{part_name}" deleted successfully.'
        })

    return redirect('core:vendor_spare_parts')


@login_required
def vendor_orders_view(request):
    """Vendor orders dashboard"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor
        orders = Order.objects.filter(
            items__vendor=vendor
        ).distinct().prefetch_related('items', 'payments').order_by('-created_at')

        # Calculate stats
        total_orders = orders.count()
        total_revenue = sum(
            sum(item.total_price for item in order.items.filter(vendor=vendor))
            for order in orders if order.payment_status == 'completed'
        )
        pending_orders = orders.filter(status='pending').count()
        processing_orders = orders.filter(status='processing').count()

        context = {
            'vendor': vendor,
            'orders': orders[:20],  # Recent 20 orders
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'pending_orders': pending_orders,
            'processing_orders': processing_orders,
        }

        return render(request, 'core/dashboard/vendor_orders.html', context)

    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
def vendor_import_requests_view(request):
    """Vendor import requests management view - shows import requests submitted by the vendor"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor

        # Get import requests submitted by this vendor (vendor as customer)
        import_requests = ImportRequest.objects.filter(customer=request.user).order_by('-created_at')

        # Filter by status if provided
        status_filter = request.GET.get('status')
        if status_filter:
            import_requests = import_requests.filter(status=status_filter)

        # Search functionality
        search_query = request.GET.get('search')
        if search_query:
            import_requests = import_requests.filter(
                Q(brand__icontains=search_query) |
                Q(model__icontains=search_query) |
                Q(origin_country__icontains=search_query)
            )

        # Pagination
        from django.core.paginator import Paginator
        paginator = Paginator(import_requests, 20)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        # Calculate stats
        total_requests = ImportRequest.objects.filter(customer=request.user).count()
        pending_requests = ImportRequest.objects.filter(customer=request.user, status='pending').count()
        completed_requests = ImportRequest.objects.filter(customer=request.user, status='completed').count()
        in_progress_requests = ImportRequest.objects.filter(
            customer=request.user,
            status__in=['on_quotation', 'processing', 'fee_paid']
        ).count()

        context = {
            'vendor': vendor,
            'import_requests': page_obj,
            'total_requests': total_requests,
            'pending_requests': pending_requests,
            'in_progress_requests': in_progress_requests,
            'completed_requests': completed_requests,
            'status_filter': status_filter,
            'search_query': search_query,
            'is_paginated': page_obj.has_other_pages(),
            'page_obj': page_obj,
        }

        # Handle HTMX requests
        if request.headers.get('HX-Request'):
            return render(request, 'core/dashboard/partials/vendor_import_requests_table.html', context)

        return render(request, 'core/dashboard/vendor_import_requests.html', context)

    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
def vendor_import_request_detail_view(request, request_id):
    """Vendor import request detail view - shows detailed information about a specific import request"""
    if request.user.role != 'vendor':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    try:
        vendor = request.user.vendor

        # Get the import request - ensure it belongs to this vendor
        import_request = get_object_or_404(ImportRequest, id=request_id, customer=request.user)

        # Get related import order if it exists
        import_order = None
        if hasattr(import_request, 'import_order'):
            import_order = import_request.import_order

        context = {
            'vendor': vendor,
            'import_request': import_request,
            'import_order': import_order,
            'status_choices': ImportRequest.STATUS_CHOICES,
        }

        return render(request, 'core/dashboard/vendor_import_request_detail.html', context)

    except Vendor.DoesNotExist:
        messages.warning(request, 'Please complete your vendor profile.')
        return redirect('core:vendor_profile_create')


@login_required
def admin_spare_parts_overview(request):
    """Admin spare parts overview"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Calculate system-wide spare parts analytics
    total_parts = SparePart.objects.count()
    total_suppliers = Supplier.objects.filter(is_active=True).count()
    total_categories = SparePartCategory.objects.filter(is_active=True).count()
    total_stock_value = sum(part.stock_value for part in SparePart.objects.all())

    # Low stock alerts
    low_stock_parts = [part for part in SparePart.objects.all() if part.is_low_stock]
    active_alerts = InventoryAlert.objects.filter(status='active').count()

    # Recent orders
    recent_orders = Order.objects.filter(
        items__spare_part__isnull=False
    ).distinct().order_by('-created_at')[:10]

    # Monthly data
    monthly_data = []
    for i in range(6):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)

        month_orders = Order.objects.filter(
            created_at__range=[month_start, month_end],
            items__spare_part__isnull=False
        ).distinct()

        month_revenue = sum(
            sum(item.total_price for item in order.items.all())
            for order in month_orders if order.payment_status == 'completed'
        )

        monthly_data.append({
            'month': month_start.strftime('%B'),
            'orders': month_orders.count(),
            'revenue': month_revenue,
            'parts_sold': sum(
                sum(item.quantity for item in order.items.all())
                for order in month_orders
            ),
        })

    context = {
        'total_parts': total_parts,
        'total_suppliers': total_suppliers,
        'total_categories': total_categories,
        'total_stock_value': total_stock_value,
        'low_stock_count': len(low_stock_parts),
        'active_alerts': active_alerts,
        'recent_orders': recent_orders,
        'monthly_data': monthly_data,
        'top_selling_parts': SparePart.objects.annotate(
            total_sold=Sum('orderitem__quantity')
        ).order_by('-total_sold')[:10],
    }

    return render(request, 'core/dashboard/admin_spare_parts.html', context)


@login_required
def user_quick_actions_lazy(request):
    """HTMX endpoint for user quick actions"""
    if request.user.role == 'admin' or request.user.role == 'vendor':
        return HttpResponse('Unauthorized', status=403)

    # Get user stats
    import_orders = ImportOrder.objects.filter(customer=request.user).order_by('-created_at')[:5]
    customer_inquiries = Inquiry.objects.filter(customer=request.user)[:5]

    context = {
        'import_orders': import_orders,
        'customer_inquiries': customer_inquiries,
    }

    return render(request, 'core/dashboard/partials/user_quick_actions.html', context)


@login_required
def approve_car_listing(request, car_id):
    """Approve a car listing (Admin only)"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    car = get_object_or_404(Car, id=car_id)
    car.is_approved = True
    car.approval_date = timezone.now()
    car.save()
    
    messages.success(request, f'Car listing "{car.title}" has been approved.')
    
    if request.headers.get('HX-Request'):
        return JsonResponse({'status': 'approved'})
    
    return redirect('core:admin_listings')


@login_required
def reject_car_listing(request, car_id):
    """Reject a car listing (Admin only)"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    car = get_object_or_404(Car, id=car_id)
    # For now, we'll just delete rejected listings
    # In a real system, you might want to keep them with a 'rejected' status
    car.delete()
    
    messages.success(request, 'Car listing has been rejected and removed.')
    
    if request.headers.get('HX-Request'):
        return JsonResponse({'status': 'rejected'})
    
    return redirect('core:admin_listings')


@login_required
def approve_vendor(request, vendor_id):
    """Approve a vendor (Admin only)"""
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    vendor = get_object_or_404(Vendor, id=vendor_id)
    vendor.is_approved = True
    vendor.approval_date = timezone.now()
    vendor.verification_status = 'verified'
    vendor.save()

    # Log the activity
    log_user_activity(
        request.user,
        'vendor_approval',
        f'Admin approved vendor {vendor.company_name}',
        request,
        {'vendor_id': vendor.id}
    )

    messages.success(request, f'Vendor "{vendor.company_name}" has been approved.')

    if request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'approved'})

    return redirect('core:admin_vendors')


@login_required
def disapprove_vendor(request, vendor_id):
    """Disapprove/revoke vendor approval"""
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    vendor = get_object_or_404(Vendor, id=vendor_id)
    vendor.is_approved = False
    vendor.approval_date = None
    vendor.verification_status = 'pending'
    vendor.save()

    # Log the activity
    log_user_activity(
        request.user,
        'vendor_disapproval',
        f'Admin revoked approval for vendor {vendor.company_name}',
        request,
        {'vendor_id': vendor.id}
    )

    messages.success(request, f'Approval revoked for vendor "{vendor.company_name}".')

    if request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'disapproved'})

    return redirect('core:admin_vendors')


@login_required
def suspend_vendor(request, vendor_id):
    """Suspend a vendor account"""
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    vendor = get_object_or_404(Vendor, id=vendor_id)
    vendor.user.is_active = False
    vendor.user.save()

    # Log the activity
    log_user_activity(
        request.user,
        'vendor_suspension',
        f'Admin suspended vendor {vendor.company_name}',
        request,
        {'vendor_id': vendor.id, 'user_id': vendor.user.id}
    )

    messages.success(request, f'Vendor "{vendor.company_name}" has been suspended.')

    if request.headers.get('HX-Request') or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'status': 'suspended'})

    return redirect('core:admin_vendors')


@login_required
def respond_to_inquiry(request):
    """HTMX endpoint for responding to inquiries"""
    if request.method != 'POST' or request.user.role != 'vendor':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    inquiry_id = request.POST.get('inquiry_id')
    response_text = request.POST.get('response')
    status = request.POST.get('status')
    send_email = request.POST.get('send_email') == 'on'

    try:
        vendor = request.user.vendor
        inquiry = get_object_or_404(Inquiry, id=inquiry_id, car__vendor=vendor)

        # Since response is not a model field, we'll handle it differently
        # For now, we'll just update the status and add a note to the message
        inquiry.status = status
        inquiry.save()

        # TODO: Implement a proper response system with a separate model

        # TODO: Send email notification if send_email is True

        messages.success(request, 'Response sent successfully!')

        # Return updated inquiry list
        inquiries = Inquiry.objects.filter(car__vendor=vendor).order_by('-created_at')
        from django.core.paginator import Paginator
        paginator = Paginator(inquiries, 10)
        inquiries_page = paginator.get_page(1)

        context = {
            'vendor': vendor,
            'inquiries': inquiries_page,
            'vendor_cars': Car.objects.filter(vendor=vendor),
        }

        return render(request, 'core/dashboard/partials/inquiry_list.html', context)

    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def update_inquiry_status(request):
    """HTMX endpoint for updating inquiry status"""
    if request.method != 'POST' or request.user.role != 'vendor':
        return JsonResponse({'error': 'Invalid request'}, status=400)

    inquiry_id = request.POST.get('inquiry_id')
    status = request.POST.get('status')

    try:
        vendor = request.user.vendor
        inquiry = get_object_or_404(Inquiry, id=inquiry_id, car__vendor=vendor)

        inquiry.status = status
        if status == 'resolved' and not inquiry.response_date:
            inquiry.response_date = timezone.now()
        inquiry.save()

        messages.success(request, f'Inquiry status updated to {inquiry.get_status_display()}!')

        # Return updated inquiry card
        context = {'inquiry': inquiry}
        return render(request, 'core/dashboard/partials/inquiry_card.html', context)

    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


# ===== EXPORT FUNCTIONALITY =====

@login_required
def export_users_csv(request):
    """Export users data to CSV format"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="users_export.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Username', 'Email', 'First Name', 'Last Name', 'Role', 'Phone', 'Date Joined', 'Is Active', 'Is Verified'])

    users = User.objects.all().order_by('-date_joined')
    for user in users:
        writer.writerow([
            user.id,
            user.username,
            user.email,
            user.first_name,
            user.last_name,
            user.role,
            user.phone,
            user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            user.is_active,
            user.is_verified
        ])

    return response


@login_required
def export_cars_csv(request):
    """Export car listings data to CSV format"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="cars_export.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Brand', 'Model', 'Year', 'Price', 'Mileage', 'Fuel Type', 'Transmission', 'Condition', 'Vendor', 'Is Approved', 'Created At'])

    cars = Car.objects.select_related('brand', 'model', 'vendor').all().order_by('-created_at')
    for car in cars:
        writer.writerow([
            car.id,
            car.brand.name if car.brand else '',
            car.model.name if car.model else '',
            car.year,
            car.price,
            car.mileage,
            car.fuel_type,
            car.transmission,
            car.condition,
            car.vendor.company_name if car.vendor else '',
            car.is_approved,
            car.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


@login_required
def export_vendors_csv(request):
    """Export vendors data to CSV format"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="vendors_export.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Company Name', 'User Email', 'Business License', 'Website', 'Is Approved', 'Approval Date', 'Created At'])

    vendors = Vendor.objects.select_related('user').all().order_by('-created_at')
    for vendor in vendors:
        writer.writerow([
            vendor.id,
            vendor.company_name,
            vendor.user.email,
            vendor.business_license,
            vendor.website,
            vendor.is_approved,
            vendor.approval_date.strftime('%Y-%m-%d %H:%M:%S') if vendor.approval_date else '',
            vendor.created_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


@login_required
def export_analytics_json(request):
    """Export analytics data to JSON format"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Gather analytics data
    total_users = User.objects.count()
    total_cars = Car.objects.count()
    total_vendors = Vendor.objects.count()
    total_inquiries = Inquiry.objects.count()

    # Monthly data for the last 12 months
    monthly_data = []
    for i in range(12):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)

        monthly_data.append({
            'month': month_start.strftime('%Y-%m'),
            'users': User.objects.filter(date_joined__range=[month_start, month_end]).count(),
            'cars': Car.objects.filter(created_at__range=[month_start, month_end]).count(),
            'vendors': Vendor.objects.filter(created_at__range=[month_start, month_end]).count(),
        })

    # Top brands
    top_brands = list(CarBrand.objects.annotate(
        car_count=Count('car')
    ).order_by('-car_count')[:10].values('name', 'car_count'))

    analytics_data = {
        'export_date': timezone.now().isoformat(),
        'summary': {
            'total_users': total_users,
            'total_cars': total_cars,
            'total_vendors': total_vendors,
            'total_inquiries': total_inquiries,
        },
        'monthly_data': monthly_data,
        'top_brands': top_brands,
    }

    response = HttpResponse(
        json.dumps(analytics_data, indent=2),
        content_type='application/json'
    )
    response['Content-Disposition'] = 'attachment; filename="analytics_export.json"'

    return response


# ===== NOTIFICATION SYSTEM =====

@login_required
def notifications_view(request):
    """Display user notifications"""
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    unread_count = notifications.filter(is_read=False).count()

    context = {
        'notifications': notifications,
        'unread_count': unread_count,
    }

    return render(request, 'core/dashboard/notifications.html', context)


@login_required
def mark_notification_read(request, notification_id):
    """Mark a notification as read via HTMX"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, recipient=request.user)
        notification.mark_as_read()

        return JsonResponse({'success': True, 'message': 'Notification marked as read'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required
def mark_all_notifications_read(request):
    """Mark all notifications as read via HTMX"""
    if request.method == 'POST':
        try:
            Notification.objects.filter(recipient=request.user, is_read=False).update(
                is_read=True,
                read_at=timezone.now()
            )

            return JsonResponse({'success': True, 'message': 'All notifications marked as read'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


from .decorators import htmx_login_required

@htmx_login_required
def notifications_count_htmx(request):
    """Get unread notifications count via HTMX"""
    unread_count = Notification.objects.filter(recipient=request.user, is_read=False).count()

    context = {'unread_count': unread_count}
    return render(request, 'core/dashboard/partials/notification_count.html', context)


from .decorators import ajax_login_required

@ajax_login_required
def notification_badges_api(request):
    """API endpoint for real-time notification badge updates"""
    from .context_processors import notification_badges

    # Get badge data from context processor
    badge_data = notification_badges(request)
    badges = badge_data.get('notification_badges', {})

    # Format for JavaScript consumption
    badge_updates = []

    # Map badge types to their identifiers
    badge_mappings = {
        'unread_notifications': {'selector': '[data-badge="notifications"]', 'type': 'primary'},
        'pending_import_requests': {'selector': '[data-badge="import-requests"]', 'type': 'warning'},
        'active_import_orders': {'selector': '[data-badge="import-orders"]', 'type': 'blue'},
        'new_inquiries': {'selector': '[data-badge="inquiries"]', 'type': 'warning'},
        'pending_approvals': {'selector': '[data-badge="approvals"]', 'type': 'danger'},
        'system_alerts': {'selector': '[data-badge="alerts"]', 'type': 'danger'},
        'vendor_orders': {'selector': '[data-badge="vendor-orders"]', 'type': 'blue'},
        'vendor_inquiries': {'selector': '[data-badge="vendor-inquiries"]', 'type': 'warning'},
        'order_updates': {'selector': '[data-badge="order-updates"]', 'type': 'primary'},
        'import_updates': {'selector': '[data-badge="import-updates"]', 'type': 'blue'},
        'inquiry_responses': {'selector': '[data-badge="inquiry-responses"]', 'type': 'warning'},
    }

    for badge_key, count in badges.items():
        if badge_key in badge_mappings and count > 0:
            mapping = badge_mappings[badge_key]
            badge_updates.append({
                'badgeId': badge_key,
                'selector': mapping['selector'],
                'count': count,
                'type': mapping['type'],
                'options': {
                    'pulse': count > 5,  # Pulse for high counts
                    'urgent': badge_key in ['system_alerts', 'pending_approvals']
                }
            })

    return JsonResponse({
        'badges': badge_updates,
        'timestamp': timezone.now().isoformat(),
        'total_count': badges.get('total_badge_count', 0)
    })


def create_notification(recipient, title, message, notification_type='info', action_url='', action_text=''):
    """Helper function to create notifications"""
    return Notification.objects.create(
        recipient=recipient,
        title=title,
        message=message,
        notification_type=notification_type,
        action_url=action_url,
        action_text=action_text
    )


@login_required
def activity_logs_view(request):
    """View user's activity logs"""
    from .activity_manager import ActivityManager

    # Get user's activities
    activities = ActivityManager.get_user_activities(request.user, limit=100)

    # Filter by action if specified
    action_filter = request.GET.get('action')
    if action_filter:
        activities = activities.filter(action=action_filter)

    # Get available actions for filter
    available_actions = ActivityLog.objects.filter(user=request.user).values_list('action', flat=True).distinct()

    context = {
        'activities': activities,
        'available_actions': available_actions,
        'current_filter': action_filter,
    }

    return render(request, 'core/dashboard/activity_logs.html', context)


@login_required
def admin_activity_logs_view(request):
    """Admin view for all activity logs"""
    if not request.user.role == 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from .activity_manager import ActivityManager

    # Get system activities
    activities = ActivityManager.get_system_activities(limit=200)

    # Apply filters
    user_filter = request.GET.get('user')
    action_filter = request.GET.get('action')
    date_filter = request.GET.get('date')

    if user_filter:
        try:
            user = User.objects.get(username=user_filter)
            activities = activities.filter(user=user)
        except User.DoesNotExist:
            pass

    if action_filter:
        activities = activities.filter(action=action_filter)

    if date_filter:
        try:
            filter_date = datetime.strptime(date_filter, '%Y-%m-%d').date()
            activities = activities.filter(timestamp__date=filter_date)
        except ValueError:
            pass

    # Get filter options
    available_actions = ActivityLog.objects.values_list('action', flat=True).distinct()
    recent_users = User.objects.filter(activity_logs__isnull=False).distinct()[:20]

    context = {
        'activities': activities,
        'available_actions': available_actions,
        'recent_users': recent_users,
        'current_filters': {
            'user': user_filter,
            'action': action_filter,
            'date': date_filter,
        }
    }

    return render(request, 'core/dashboard/admin_activity_logs.html', context)


@login_required
def admin_audit_logs_view(request):
    """Admin view for audit logs"""
    if not request.user.role == 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from .activity_manager import AuditManager

    # Get audit logs
    audit_logs = AuditManager.get_audit_trail(limit=200)

    # Apply filters
    action_filter = request.GET.get('action_type')
    severity_filter = request.GET.get('severity')
    table_filter = request.GET.get('table_name')

    if action_filter:
        audit_logs = audit_logs.filter(action_type=action_filter)

    if severity_filter:
        audit_logs = audit_logs.filter(severity=severity_filter)

    if table_filter:
        audit_logs = audit_logs.filter(table_name=table_filter)

    # Get filter options
    available_actions = AuditLog.objects.values_list('action_type', flat=True).distinct()
    available_tables = AuditLog.objects.exclude(table_name='').values_list('table_name', flat=True).distinct()

    context = {
        'audit_logs': audit_logs,
        'available_actions': available_actions,
        'available_tables': available_tables,
        'severity_choices': AuditLog.SEVERITY_LEVELS,
        'current_filters': {
            'action_type': action_filter,
            'severity': severity_filter,
            'table_name': table_filter,
        }
    }

    return render(request, 'core/dashboard/admin_audit_logs.html', context)


@login_required
def notification_preferences_view(request):
    """View and update notification preferences"""
    from .notification_manager import NotificationManager

    preferences = NotificationManager.get_user_preferences(request.user)

    if request.method == 'POST':
        # Update preferences
        preferences.email_enabled = request.POST.get('email_enabled') == 'on'
        preferences.email_order_updates = request.POST.get('email_order_updates') == 'on'
        preferences.email_import_updates = request.POST.get('email_import_updates') == 'on'
        preferences.email_inquiry_responses = request.POST.get('email_inquiry_responses') == 'on'
        preferences.email_marketing = request.POST.get('email_marketing') == 'on'
        preferences.email_security_alerts = request.POST.get('email_security_alerts') == 'on'

        preferences.sms_enabled = request.POST.get('sms_enabled') == 'on'
        preferences.sms_order_updates = request.POST.get('sms_order_updates') == 'on'
        preferences.sms_import_updates = request.POST.get('sms_import_updates') == 'on'
        preferences.sms_security_alerts = request.POST.get('sms_security_alerts') == 'on'

        preferences.push_enabled = request.POST.get('push_enabled') == 'on'
        preferences.push_order_updates = request.POST.get('push_order_updates') == 'on'
        preferences.push_import_updates = request.POST.get('push_import_updates') == 'on'
        preferences.push_inquiry_responses = request.POST.get('push_inquiry_responses') == 'on'
        preferences.push_marketing = request.POST.get('push_marketing') == 'on'

        preferences.in_app_enabled = request.POST.get('in_app_enabled') == 'on'
        preferences.in_app_order_updates = request.POST.get('in_app_order_updates') == 'on'
        preferences.in_app_import_updates = request.POST.get('in_app_import_updates') == 'on'
        preferences.in_app_inquiry_responses = request.POST.get('in_app_inquiry_responses') == 'on'
        preferences.in_app_system_updates = request.POST.get('in_app_system_updates') == 'on'

        preferences.digest_frequency = request.POST.get('digest_frequency', 'immediate')
        preferences.quiet_hours_enabled = request.POST.get('quiet_hours_enabled') == 'on'

        if preferences.quiet_hours_enabled:
            quiet_start = request.POST.get('quiet_hours_start')
            quiet_end = request.POST.get('quiet_hours_end')
            if quiet_start:
                preferences.quiet_hours_start = quiet_start
            if quiet_end:
                preferences.quiet_hours_end = quiet_end

        preferences.save()
        messages.success(request, 'Notification preferences updated successfully!')

        if request.headers.get('HX-Request'):
            return JsonResponse({'success': True, 'message': 'Preferences updated!'})

        return redirect('core:notification_preferences')

    context = {
        'preferences': preferences,
    }

    return render(request, 'core/dashboard/notification_preferences.html', context)


@login_required
def admin_notification_queue_view(request):
    """Admin view for notification queue management"""
    if not request.user.role == 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get notification queue
    queue_items = NotificationQueue.objects.all()[:100]

    # Apply filters
    status_filter = request.GET.get('status')
    channel_filter = request.GET.get('channel')

    if status_filter:
        queue_items = queue_items.filter(status=status_filter)

    if channel_filter:
        queue_items = queue_items.filter(channel=channel_filter)

    # Get statistics
    stats = {
        'total': NotificationQueue.objects.count(),
        'pending': NotificationQueue.objects.filter(status='pending').count(),
        'processing': NotificationQueue.objects.filter(status='processing').count(),
        'sent': NotificationQueue.objects.filter(status='sent').count(),
        'failed': NotificationQueue.objects.filter(status='failed').count(),
    }

    context = {
        'queue_items': queue_items,
        'stats': stats,
        'status_choices': NotificationQueue.STATUS_CHOICES,
        'channel_choices': NotificationQueue.CHANNEL_CHOICES,
        'current_filters': {
            'status': status_filter,
            'channel': channel_filter,
        }
    }

    return render(request, 'core/dashboard/admin_notification_queue.html', context)


# ===== ADD NEW FUNCTIONALITY =====

@login_required
def add_new_modal(request):
    """Display Add New modal content via HTMX"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    context = {
        'user': request.user,
    }

    return render(request, 'core/dashboard/partials/add_new_modal.html', context)


# ===== NEW ADMIN PAGES =====

@login_required
def admin_import_requests_view(request):
    """Admin import requests management view with enhanced filtering and search"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Base queryset with optimized select_related
    import_requests = ImportRequest.objects.select_related('customer').all().order_by('-created_at')

    # Apply filters
    import_requests = apply_import_request_filters(request, import_requests)

    # Calculate stats (unfiltered for dashboard overview)
    total_imports = ImportRequest.objects.count()
    pending_imports = ImportRequest.objects.filter(status='pending').count()
    in_transit_imports = ImportRequest.objects.filter(status__in=['processing', 'shipped']).count()
    completed_imports = ImportRequest.objects.filter(status='completed').count()

    # Get unique values for filter dropdowns
    brands = ImportRequest.objects.values_list('brand', flat=True).distinct().order_by('brand')
    countries = ImportRequest.objects.values_list('origin_country', flat=True).distinct().order_by('origin_country')

    context = {
        'import_requests': import_requests[:20],  # Limit to 20 for performance
        'total_imports': total_imports,
        'pending_imports': pending_imports,
        'in_transit_imports': in_transit_imports,
        'completed_imports': completed_imports,
        'status_choices': ImportRequest.STATUS_CHOICES,
        'brands': brands,
        'countries': countries,
        # Preserve filter values
        'current_status': request.GET.get('status', ''),
        'current_search': request.GET.get('search', ''),
        'current_brand': request.GET.get('brand', ''),
        'current_country': request.GET.get('country', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
    }

    return render(request, 'core/dashboard/admin_import_requests.html', context)


@login_required
def admin_import_requests_table_partial(request):
    """Return just the import requests table for HTMX updates with filtering"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Base queryset with optimized select_related
    import_requests = ImportRequest.objects.select_related('customer').all().order_by('-created_at')

    # Apply the same filters as the main view
    import_requests = apply_import_request_filters(request, import_requests)

    context = {
        'import_requests': import_requests[:20],  # Limit to 20 for performance
    }

    return render(request, 'core/dashboard/partials/admin_import_requests_table.html', context)


@login_required
def admin_import_requests_export(request):
    """Export import requests to CSV with current filters applied"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Get filtered queryset
    import_requests = ImportRequest.objects.select_related('customer').all().order_by('-created_at')
    import_requests = apply_import_request_filters(request, import_requests)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="import_requests.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Order ID', 'Customer Name', 'Customer Email', 'Brand', 'Model', 'Year',
        'Origin Country', 'Status', 'Budget Min', 'Budget Max', 'Estimated Cost',
        'Created Date', 'Updated Date', 'Tracking Number'
    ])

    # Write data
    for request_obj in import_requests:
        writer.writerow([
            f"#{request_obj.id:05d}",
            request_obj.customer.get_full_name() or request_obj.customer.username,
            request_obj.customer.email,
            request_obj.brand,
            request_obj.model,
            request_obj.year,
            request_obj.origin_country,
            request_obj.get_status_display(),
            request_obj.budget_min,
            request_obj.budget_max,
            request_obj.estimated_cost or '',
            request_obj.created_at.strftime('%Y-%m-%d %H:%M'),
            request_obj.updated_at.strftime('%Y-%m-%d %H:%M'),
            request_obj.tracking_number or ''
        ])

    return response


@login_required
def admin_import_requests_refresh(request):
    """Refresh import requests table while maintaining current filters"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Simply return the updated table partial with current filters
    return admin_import_requests_table_partial(request)


def apply_tracking_filters(request, queryset):
    """Apply filters to tracking management queryset"""
    # Status filter
    status_filter = request.GET.get('status')
    if status_filter:
        queryset = queryset.filter(status=status_filter)

    # Search filter
    search = request.GET.get('search')
    if search:
        queryset = queryset.filter(
            Q(order_number__icontains=search) |
            Q(customer__username__icontains=search) |
            Q(customer__email__icontains=search) |
            Q(customer__first_name__icontains=search) |
            Q(customer__last_name__icontains=search) |
            Q(brand__icontains=search) |
            Q(model__icontains=search) |
            Q(chassis_number__icontains=search) |
            Q(bill_of_lading__icontains=search) |
            Q(vessel_name__icontains=search)
        )

    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            queryset = queryset.filter(created_at__date__lte=to_date)
        except ValueError:
            pass

    return queryset


@login_required
def admin_tracking_management_table_partial(request):
    """Return just the tracking management table for HTMX updates with pagination"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Get import orders with tracking information
    import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')

    # Apply filters
    import_orders = apply_tracking_filters(request, import_orders)

    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(import_orders, 20)  # 20 items per page

    try:
        import_orders_page = paginator.page(page)
    except PageNotAnInteger:
        import_orders_page = paginator.page(1)
    except EmptyPage:
        import_orders_page = paginator.page(paginator.num_pages)

    context = {
        'import_orders': import_orders_page,
        'paginator': paginator,
        'page_obj': import_orders_page,
        'is_paginated': paginator.num_pages > 1,
        'current_filter': request.GET.get('status', ''),
        'current_search': request.GET.get('search', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
    }

    return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', context)


@login_required
def admin_tracking_management_export_csv(request):
    """Export tracking management data to CSV with current filters applied"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Get filtered queryset
    import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')
    import_orders = apply_tracking_filters(request, import_orders)

    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="import_orders_tracking.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Order Number', 'Customer Name', 'Customer Email', 'Brand', 'Model', 'Year',
        'Color', 'Engine Size', 'Fuel Type', 'Origin Country', 'Status', 'Payment Status',
        'Total Cost', 'Paid Amount', 'Balance Due', 'Chassis Number', 'Bill of Lading',
        'Vessel Name', 'Departure Port', 'Arrival Port', 'Estimated Arrival', 'Actual Arrival',
        'Progress %', 'Created Date', 'Updated Date'
    ])

    # Write data rows
    for order in import_orders:
        writer.writerow([
            order.order_number,
            order.customer.get_full_name() or order.customer.username,
            order.customer.email,
            order.brand,
            order.model,
            order.year,
            order.color,
            order.engine_size,
            order.fuel_type,
            order.origin_country,
            order.get_status_display(),
            order.get_payment_status_display(),
            order.total_cost or 0,
            order.paid_amount or 0,
            order.balance_due or 0,
            order.chassis_number or '',
            order.bill_of_lading or '',
            order.vessel_name or '',
            order.departure_port or '',
            order.arrival_port or '',
            order.estimated_arrival_date.strftime('%Y-%m-%d') if order.estimated_arrival_date else '',
            order.actual_arrival_date.strftime('%Y-%m-%d') if order.actual_arrival_date else '',
            f"{order.get_progress_percentage()}%",
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])

    return response


@login_required
def admin_tracking_management_export_excel(request):
    """Export tracking management data to Excel with current filters applied"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return admin_tracking_management_export_csv(request)

    # Get filtered queryset
    import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')
    import_orders = apply_tracking_filters(request, import_orders)

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Orders Tracking"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    headers = [
        'Order Number', 'Customer Name', 'Customer Email', 'Brand', 'Model', 'Year',
        'Color', 'Engine Size', 'Fuel Type', 'Origin Country', 'Status', 'Payment Status',
        'Total Cost', 'Paid Amount', 'Balance Due', 'Chassis Number', 'Bill of Lading',
        'Vessel Name', 'Departure Port', 'Arrival Port', 'Estimated Arrival', 'Actual Arrival',
        'Progress %', 'Created Date', 'Updated Date'
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row, order in enumerate(import_orders, 2):
        data = [
            order.order_number,
            order.customer.get_full_name() or order.customer.username,
            order.customer.email,
            order.brand,
            order.model,
            order.year,
            order.color,
            order.engine_size,
            order.fuel_type,
            order.origin_country,
            order.get_status_display(),
            order.get_payment_status_display(),
            order.total_cost or 0,
            order.paid_amount or 0,
            order.balance_due or 0,
            order.chassis_number or '',
            order.bill_of_lading or '',
            order.vessel_name or '',
            order.departure_port or '',
            order.arrival_port or '',
            order.estimated_arrival_date.strftime('%Y-%m-%d') if order.estimated_arrival_date else '',
            order.actual_arrival_date.strftime('%Y-%m-%d') if order.actual_arrival_date else '',
            f"{order.get_progress_percentage()}%",
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ]

        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="import_orders_tracking.xlsx"'

    wb.save(response)
    return response


# ===== IMPORT ORDER MANAGEMENT VIEWS =====

@login_required
def admin_import_order_add_modal(request):
    """Show add import order modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Get all customers for the dropdown
    customers = User.objects.filter(role='customer').order_by('first_name', 'last_name', 'username')

    context = {
        'customers': customers,
    }

    return render(request, 'core/modals/admin_import_order_add.html', context)


@login_required
def admin_import_order_add(request):
    """Handle adding new import order"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            # Get form data
            customer_id = request.POST.get('customer')
            brand = request.POST.get('brand')
            model = request.POST.get('model')
            year = request.POST.get('year')
            color = request.POST.get('color', '')
            engine_size = request.POST.get('engine_size', '')
            fuel_type = request.POST.get('fuel_type', '')
            total_cost = request.POST.get('total_cost')
            initial_status = request.POST.get('initial_status', 'confirmed')
            admin_notes = request.POST.get('admin_notes', '')

            # Validate required fields
            if not all([customer_id, brand, model, year, total_cost]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                    'import_orders': ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
                })

            # Get customer
            try:
                customer = User.objects.get(id=customer_id, role='customer')
            except User.DoesNotExist:
                messages.error(request, 'Selected customer not found.')
                return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                    'import_orders': ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
                })

            # Generate unique order number
            import uuid
            order_number = f"IO{uuid.uuid4().hex[:8].upper()}"
            while ImportOrder.objects.filter(order_number=order_number).exists():
                order_number = f"IO{uuid.uuid4().hex[:8].upper()}"

            # Create import order
            import_order = ImportOrder.objects.create(
                order_number=order_number,
                customer=customer,
                brand=brand,
                model=model,
                year=int(year),
                color=color,
                engine_size=engine_size,
                fuel_type=fuel_type,
                total_cost=float(total_cost),
                status=initial_status,
                admin_notes=admin_notes,
                created_by=request.user
            )

            # Create initial status history entry
            ImportOrderStatusHistory.objects.create(
                import_order=import_order,
                previous_status='',
                new_status=initial_status,
                changed_by=request.user,
                change_reason=f'Initial order creation with status: {initial_status}',
                admin_notes=admin_notes,
                customer_notification_sent=False
            )

            messages.success(request, f'Import order {order_number} created successfully!')

            # Return updated table
            import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
            return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                'import_orders': import_orders
            })

        except Exception as e:
            messages.error(request, f'Error creating import order: {str(e)}')
            import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
            return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                'import_orders': import_orders
            })

    return HttpResponse('Method not allowed', status=405)


@login_required
def admin_import_order_edit_modal(request, order_id):
    """Show edit import order modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    try:
        import_order = get_object_or_404(ImportOrder, id=order_id)

        # Get status choices for the dropdown
        status_choices = ImportOrder.STATUS_CHOICES

        context = {
            'import_order': import_order,
            'status_choices': status_choices,
        }

        return render(request, 'core/modals/admin_import_order_edit.html', context)

    except ImportOrder.DoesNotExist:
        return HttpResponse('Import order not found', status=404)


@login_required
def admin_import_order_edit(request, order_id):
    """Handle editing import order"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            import_order = get_object_or_404(ImportOrder, id=order_id)

            # Get form data
            brand = request.POST.get('brand')
            model = request.POST.get('model')
            year = request.POST.get('year')
            color = request.POST.get('color', '')
            engine_size = request.POST.get('engine_size', '')
            fuel_type = request.POST.get('fuel_type', '')
            total_cost = request.POST.get('total_cost')
            status = request.POST.get('status')
            chassis_number = request.POST.get('chassis_number', '')
            bill_of_lading = request.POST.get('bill_of_lading', '')
            vessel_name = request.POST.get('vessel_name', '')
            estimated_arrival = request.POST.get('estimated_arrival')
            admin_notes = request.POST.get('admin_notes', '')

            # Validate required fields
            if not all([brand, model, year, total_cost, status]):
                messages.error(request, 'Please fill in all required fields.')
                return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                    'import_orders': ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
                })

            # Store previous status for history
            previous_status = import_order.status

            # Update import order
            import_order.brand = brand
            import_order.model = model
            import_order.year = int(year)
            import_order.color = color
            import_order.engine_size = engine_size
            import_order.fuel_type = fuel_type
            import_order.total_cost = float(total_cost)
            import_order.status = status
            import_order.chassis_number = chassis_number
            import_order.bill_of_lading = bill_of_lading
            import_order.vessel_name = vessel_name
            import_order.admin_notes = admin_notes

            if estimated_arrival:
                from datetime import datetime
                import_order.estimated_arrival = datetime.strptime(estimated_arrival, '%Y-%m-%d').date()

            import_order.save()

            # Create status history entry if status changed
            if previous_status != status:
                ImportOrderStatusHistory.objects.create(
                    import_order=import_order,
                    previous_status=previous_status,
                    new_status=status,
                    changed_by=request.user,
                    change_reason=f'Status updated via edit modal from {previous_status} to {status}',
                    admin_notes=admin_notes,
                    customer_notification_sent=False
                )

            messages.success(request, f'Import order {import_order.order_number} updated successfully!')

            # Return updated table
            import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
            return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                'import_orders': import_orders
            })

        except Exception as e:
            messages.error(request, f'Error updating import order: {str(e)}')
            import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')[:20]
            return render(request, 'core/dashboard/partials/admin_tracking_management_table.html', {
                'import_orders': import_orders
            })

    return HttpResponse('Method not allowed', status=405)


# ===== TRACKING MANAGEMENT MODAL VIEWS (Placeholders) =====

@login_required
def admin_tracking_status_modal(request, order_id):
    """Show tracking status update modal (placeholder)"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Placeholder - return simple modal
    return HttpResponse('<div class="modal">Status update modal for order {}</div>'.format(order_id))


@login_required
def admin_tracking_timeline_modal(request, order_id):
    """Show tracking timeline modal with complete status history"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    try:
        import_order = get_object_or_404(ImportOrder, id=order_id)

        # Get status history ordered by creation date (newest first)
        status_history = import_order.status_history.all().order_by('-created_at')

        context = {
            'import_order': import_order,
            'status_history': status_history,
        }

        return render(request, 'core/modals/admin_import_order_timeline.html', context)

    except ImportOrder.DoesNotExist:
        return HttpResponse('Import order not found', status=404)


@login_required
def admin_tracking_location_modal(request, order_id):
    """Show tracking location modal with geolocation features"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    try:
        import_order = get_object_or_404(ImportOrder, id=order_id)

        context = {
            'import_order': import_order,
        }

        return render(request, 'core/modals/admin_import_order_location.html', context)

    except ImportOrder.DoesNotExist:
        return HttpResponse('Import order not found', status=404)


# ===== GPS TRACKING ADMIN VIEWS =====

@login_required
def admin_gps_tracking_management(request):
    """Main GPS tracking management dashboard for admins"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Get import orders with tracking enabled
    tracked_orders = ImportOrder.objects.filter(
        tracking_enabled=True
    ).exclude(status__in=['delivered', 'cancelled']).select_related('customer').order_by('-created_at')

    # Get recent location updates
    from .models import LocationTrackingHistory
    recent_updates = LocationTrackingHistory.objects.select_related(
        'import_order', 'import_order__customer'
    ).order_by('-recorded_at')[:20]

    # Statistics
    total_tracked_orders = tracked_orders.count()
    orders_with_current_location = tracked_orders.filter(
        current_latitude__isnull=False,
        current_longitude__isnull=False
    ).count()

    context = {
        'tracked_orders': tracked_orders,
        'recent_updates': recent_updates,
        'total_tracked_orders': total_tracked_orders,
        'orders_with_current_location': orders_with_current_location,
        'location_coverage_percentage': int((orders_with_current_location / max(total_tracked_orders, 1)) * 100),
    }

    return render(request, 'core/dashboard/admin_gps_tracking.html', context)


@login_required
def admin_update_order_location(request, order_id):
    """Update GPS location for an import order"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_order = get_object_or_404(ImportOrder, id=order_id)

    if request.method == 'POST':
        try:
            latitude = float(request.POST.get('latitude'))
            longitude = float(request.POST.get('longitude'))
            location_name = request.POST.get('location_name', '')
            notes = request.POST.get('notes', '')

            # Update order location
            import_order.update_current_location(
                latitude=latitude,
                longitude=longitude,
                location_name=location_name,
                user=request.user
            )

            # Create detailed tracking history entry
            from .models import LocationTrackingHistory
            LocationTrackingHistory.objects.create(
                import_order=import_order,
                latitude=latitude,
                longitude=longitude,
                tracking_source='manual',
                status_at_time=import_order.status,
                notes=notes or f"Location updated by admin: {location_name}",
                recorded_at=timezone.now(),
                created_by=request.user
            )

            messages.success(request, f'Location updated successfully for order {import_order.order_number}')

            # Return updated tracking table
            tracked_orders = ImportOrder.objects.filter(
                tracking_enabled=True
            ).exclude(status__in=['delivered', 'cancelled']).select_related('customer').order_by('-created_at')

            return render(request, 'core/dashboard/partials/admin_gps_tracking_table.html', {
                'tracked_orders': tracked_orders
            })

        except (ValueError, TypeError) as e:
            messages.error(request, f'Invalid coordinates provided: {str(e)}')
            return HttpResponse('Invalid data', status=400)

    return HttpResponse('Method not allowed', status=405)


@login_required
def admin_location_management_modal(request, order_id):
    """Show comprehensive location management modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_order = get_object_or_404(ImportOrder, id=order_id)

    # Get all locations for this order
    locations = import_order.locations.all().order_by('-created_at')

    # Get tracking history
    tracking_history = import_order.tracking_history.all().order_by('-recorded_at')[:10]

    # Check if route exists
    route = getattr(import_order, 'route', None)
    waypoints = route.waypoints.all().order_by('sequence_order') if route else []

    context = {
        'import_order': import_order,
        'locations': locations,
        'tracking_history': tracking_history,
        'route': route,
        'waypoints': waypoints,
    }

    return render(request, 'core/modals/admin_location_management.html', context)


@login_required
def admin_create_route(request, order_id):
    """Create a new route for an import order"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_order = get_object_or_404(ImportOrder, id=order_id)

    if request.method == 'POST':
        try:
            from .models import ImportOrderRoute, ImportOrderLocation

            route_name = request.POST.get('route_name')
            route_type = request.POST.get('route_type')
            origin_lat = float(request.POST.get('origin_lat'))
            origin_lng = float(request.POST.get('origin_lng'))
            origin_name = request.POST.get('origin_name')
            dest_lat = float(request.POST.get('dest_lat'))
            dest_lng = float(request.POST.get('dest_lng'))
            dest_name = request.POST.get('dest_name')

            # Create origin location
            origin_location = ImportOrderLocation.objects.create(
                import_order=import_order,
                location_type='origin',
                name=origin_name,
                latitude=origin_lat,
                longitude=origin_lng,
                created_by=request.user
            )

            # Create destination location
            dest_location = ImportOrderLocation.objects.create(
                import_order=import_order,
                location_type='delivery_address',
                name=dest_name,
                latitude=dest_lat,
                longitude=dest_lng,
                created_by=request.user
            )

            # Create route
            route = ImportOrderRoute.objects.create(
                import_order=import_order,
                route_name=route_name,
                route_type=route_type,
                origin_location=origin_location,
                destination_location=dest_location,
                route_status='planned',
                created_by=request.user
            )

            messages.success(request, f'Route created successfully for order {import_order.order_number}')

            return JsonResponse({
                'success': True,
                'message': 'Route created successfully',
                'route_id': route.id
            })

        except (ValueError, TypeError) as e:
            return JsonResponse({
                'success': False,
                'message': f'Invalid data provided: {str(e)}'
            }, status=400)

    return HttpResponse('Method not allowed', status=405)


@login_required
def admin_add_waypoint(request, order_id):
    """Add a waypoint to an import order route"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_order = get_object_or_404(ImportOrder, id=order_id)

    if request.method == 'POST':
        try:
            from .models import RouteWaypoint, ImportOrderLocation

            if not hasattr(import_order, 'route'):
                return JsonResponse({
                    'success': False,
                    'message': 'No route exists for this order. Create a route first.'
                }, status=400)

            waypoint_name = request.POST.get('waypoint_name')
            waypoint_type = request.POST.get('waypoint_type')
            latitude = float(request.POST.get('latitude'))
            longitude = float(request.POST.get('longitude'))
            sequence_order = int(request.POST.get('sequence_order', 1))

            # Create location for waypoint
            location = ImportOrderLocation.objects.create(
                import_order=import_order,
                location_type='transit_port' if waypoint_type == 'transit' else waypoint_type,
                name=waypoint_name,
                latitude=latitude,
                longitude=longitude,
                is_waypoint=True,
                created_by=request.user
            )

            # Create waypoint
            waypoint = RouteWaypoint.objects.create(
                route=import_order.route,
                location=location,
                waypoint_type=waypoint_type,
                sequence_order=sequence_order,
                name=waypoint_name
            )

            messages.success(request, f'Waypoint "{waypoint_name}" added successfully')

            return JsonResponse({
                'success': True,
                'message': 'Waypoint added successfully',
                'waypoint_id': waypoint.id
            })

        except (ValueError, TypeError) as e:
            return JsonResponse({
                'success': False,
                'message': f'Invalid data provided: {str(e)}'
            }, status=400)

    return HttpResponse('Method not allowed', status=405)


@login_required
def admin_tracking_details_modal(request, order_id):
    """Show comprehensive tracking details modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    try:
        import_order = get_object_or_404(ImportOrder, id=order_id)

        context = {
            'import_order': import_order,
        }

        return render(request, 'core/modals/admin_import_order_view.html', context)

    except ImportOrder.DoesNotExist:
        return HttpResponse('Import order not found', status=404)





@login_required
def admin_queries_view(request):
    """Admin queries/inquiries management view"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    inquiries = Inquiry.objects.all().order_by('-created_at')

    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        inquiries = inquiries.filter(status=status_filter)

    # Calculate stats
    total_inquiries = Inquiry.objects.count()
    new_inquiries = Inquiry.objects.filter(status='new').count()
    in_progress_inquiries = Inquiry.objects.filter(status='in_progress').count()
    resolved_inquiries = Inquiry.objects.filter(status='resolved').count()

    context = {
        'inquiries': inquiries[:20],
        'total_inquiries': total_inquiries,
        'new_inquiries': new_inquiries,
        'in_progress_inquiries': in_progress_inquiries,
        'resolved_inquiries': resolved_inquiries,
        'current_filter': status_filter,
    }

    return render(request, 'core/dashboard/admin_queries.html', context)


@login_required
def admin_content_management_view(request):
    """Enhanced admin content management view with comprehensive stats"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from .models import StaticPage, ContentCategory, ContentTag, ContentSeries

    # Get comprehensive content statistics
    total_posts = BlogPost.objects.count()
    published_posts = BlogPost.objects.filter(is_published=True).count()
    draft_posts = BlogPost.objects.filter(is_published=False).count()
    featured_posts = BlogPost.objects.filter(is_featured=True).count()

    total_testimonials = Testimonial.objects.count()
    approved_testimonials = Testimonial.objects.filter(is_approved=True).count()
    featured_testimonials = Testimonial.objects.filter(is_featured=True).count()

    total_static_pages = StaticPage.objects.count()
    published_pages = StaticPage.objects.filter(status='published').count()

    total_categories = ContentCategory.objects.filter(is_active=True).count()
    total_tags = ContentTag.objects.filter(is_active=True).count()
    total_series = ContentSeries.objects.filter(is_active=True).count()

    # Content type breakdown
    content_types = BlogPost.objects.values('content_type').annotate(
        count=models.Count('id')
    ).order_by('-count')

    # Recent content for quick overview
    recent_posts = BlogPost.objects.select_related('author', 'category').order_by('-created_at')[:5]
    recent_testimonials = Testimonial.objects.select_related('customer').order_by('-created_at')[:5]
    recent_pages = StaticPage.objects.select_related('author').order_by('-updated_at')[:5]

    context = {
        # Statistics
        'total_posts': total_posts,
        'published_posts': published_posts,
        'draft_posts': draft_posts,
        'featured_posts': featured_posts,
        'total_testimonials': total_testimonials,
        'approved_testimonials': approved_testimonials,
        'featured_testimonials': featured_testimonials,
        'total_static_pages': total_static_pages,
        'published_pages': published_pages,
        'total_categories': total_categories,
        'total_tags': total_tags,
        'total_series': total_series,
        'content_types': content_types,

        # Recent content
        'recent_posts': recent_posts,
        'recent_testimonials': recent_testimonials,
        'recent_pages': recent_pages,
    }

    return render(request, 'core/dashboard/admin_content_management.html', context)


@login_required
def admin_system_settings_view(request):
    """Admin system settings view"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    settings = SystemSetting.objects.all().order_by('key')

    context = {
        'settings': settings,
    }

    return render(request, 'core/dashboard/admin_system_settings.html', context)


@login_required
def admin_spare_shop_view(request):
    """Enhanced admin spare shop management view with comprehensive analytics"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get search query
    search_query = request.GET.get('search', '').strip()

    # Base queryset with optimized select_related
    spare_parts = SparePart.objects.select_related(
        'category_new', 'supplier', 'vendor'
    ).all().order_by('-created_at')

    # Apply search filter
    if search_query:
        spare_parts = spare_parts.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(part_number__icontains=search_query) |
            Q(barcode__icontains=search_query) |
            Q(vendor__company_name__icontains=search_query)
        )

    # Filter by category
    category_filter = request.GET.get('category')
    if category_filter:
        spare_parts = spare_parts.filter(category_new_id=category_filter)

    # Filter by availability - fix the filter values to match template
    availability_filter = request.GET.get('availability')
    if availability_filter == 'available':
        spare_parts = spare_parts.filter(stock_quantity__gt=0, is_available=True)
    elif availability_filter == 'low_stock':
        spare_parts = spare_parts.filter(
            stock_quantity__lte=F('minimum_stock'),
            stock_quantity__gt=0,
            minimum_stock__isnull=False
        )
    elif availability_filter == 'out_of_stock':
        spare_parts = spare_parts.filter(stock_quantity=0)

    # Filter by brand (compatible_brands is a ManyToManyField)
    brand_filter = request.GET.get('brand')
    if brand_filter:
        try:
            spare_parts = spare_parts.filter(compatible_brands__name__icontains=brand_filter)
        except:
            # Field might not exist or have different structure, skip brand filtering
            pass

    # Filter by price range
    price_range_filter = request.GET.get('price_range')
    if price_range_filter:
        if price_range_filter == '0-1000':
            spare_parts = spare_parts.filter(price__lt=1000)
        elif price_range_filter == '1000-5000':
            spare_parts = spare_parts.filter(price__gte=1000, price__lt=5000)
        elif price_range_filter == '5000-10000':
            spare_parts = spare_parts.filter(price__gte=5000, price__lt=10000)
        elif price_range_filter == '10000-50000':
            spare_parts = spare_parts.filter(price__gte=10000, price__lt=50000)
        elif price_range_filter == '50000+':
            spare_parts = spare_parts.filter(price__gte=50000)

    # Filter by condition (if field exists)
    condition_filter = request.GET.get('condition')
    if condition_filter:
        try:
            spare_parts = spare_parts.filter(condition=condition_filter)
        except:
            # Field might not exist, skip condition filtering
            pass

    # Enhanced statistics with aggregations
    stats = SparePart.objects.aggregate(
        total_parts=Count('id'),
        in_stock_parts=Count('id', filter=Q(stock_quantity__gt=0, is_available=True)),
        low_stock_parts=Count('id', filter=Q(
            stock_quantity__lte=F('minimum_stock'),
            stock_quantity__gt=0,
            minimum_stock__isnull=False
        )),
        out_of_stock_parts=Count('id', filter=Q(stock_quantity=0)),
        total_stock_value=Sum(F('stock_quantity') * F('price')),
        avg_price=Avg('price')
    )

    # Get categories for filter
    categories = SparePartCategory.objects.filter(is_active=True).order_by('name')

    # Get recent orders with proper error handling
    recent_orders = []
    try:
        # Try different possible field names for order items
        try:
            recent_orders = Order.objects.filter(
                items__spare_part__isnull=False
            ).select_related('customer').distinct().order_by('-created_at')[:5]
        except:
            try:
                recent_orders = Order.objects.filter(
                    orderitem__spare_part__isnull=False
                ).select_related('customer').distinct().order_by('-created_at')[:5]
            except:
                # Just get recent orders without spare part filtering
                recent_orders = Order.objects.select_related('customer').order_by('-created_at')[:5]
    except (AttributeError, NameError):
        # Handle case where Order model doesn't exist or has different structure
        pass

    # Pagination for better performance
    from django.core.paginator import Paginator
    paginator = Paginator(spare_parts, 10)  # 10 rows per page as per user preference
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'spare_parts': page_obj,
        'total_parts': stats['total_parts'] or 0,
        'in_stock_parts': stats['in_stock_parts'] or 0,
        'low_stock_parts': stats['low_stock_parts'] or 0,
        'out_of_stock_parts': stats['out_of_stock_parts'] or 0,
        'total_stock_value': stats['total_stock_value'] or 0,
        'avg_price': stats['avg_price'] or 0,
        'categories': categories,
        'recent_orders': recent_orders,
        'current_category': category_filter,
        'current_availability': availability_filter,
        'current_brand': brand_filter,
        'current_price_range': price_range_filter,
        'current_condition': condition_filter,
        'search_query': search_query,
        'page_obj': page_obj,
    }

    # Return partial template for HTMX requests
    if request.headers.get('HX-Request'):
        return render(request, 'core/partials/admin_spare_shop_table.html', context)

    return render(request, 'core/dashboard/admin_spare_shop.html', context)


@login_required
def admin_spare_part_add_modal_view(request):
    """Modal view for adding new spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    categories = SparePartCategory.objects.filter(is_active=True).order_by('name')
    suppliers = Supplier.objects.filter(is_active=True).order_by('name')

    context = {
        'categories': categories,
        'suppliers': suppliers,
    }

    return render(request, 'core/modals/admin_spare_part_add.html', context)


@login_required
@require_http_methods(["GET", "POST"])
def admin_spare_part_add_view(request):
    """Handle adding new spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'GET':
        # Return add modal with all categories (parent and sub-categories)
        categories = SparePartCategory.objects.filter(is_active=True).order_by('parent__name', 'name')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        context = {
            'categories': categories,
            'suppliers': suppliers,
        }
        return render(request, 'core/modals/admin_spare_part_add.html', context)

    elif request.method == 'POST':
        print("Admin POST request received")
        print(f"POST data: {request.POST}")
        print(f"FILES data: {request.FILES}")
        print(f"Content type: {request.content_type}")
        print(f"HX-Request header: {request.headers.get('HX-Request')}")

        # Check specific fields
        print(f"Unit field: {request.POST.get('unit', 'NOT FOUND')}")
        print(f"Reorder quantity field: {request.POST.get('reorder_quantity', 'NOT FOUND')}")

        form = SparePartForm(request.POST, request.FILES)
        print(f"Form created: {form}")
        print(f"Form data: {form.data}")

        if form.is_valid():
            print("Admin form is valid")
            try:
                spare_part = form.save(commit=False)

                # Set category field based on category_new (required for model)
                if spare_part.category_new:
                    spare_part.category = spare_part.category_new.name
                elif not spare_part.category:
                    spare_part.category = 'General'

                # Admin parts don't need a vendor
                spare_part.vendor = None
                spare_part.is_available = True
                spare_part.save()

                # Create initial stock movement
                StockMovement.objects.create(
                    spare_part=spare_part,
                    movement_type='in',
                    reason='initial',
                    quantity=spare_part.stock_quantity,
                    quantity_before=0,
                    quantity_after=spare_part.stock_quantity,
                    notes=f'Initial stock for {spare_part.name}',
                    created_by=request.user
                )

                messages.success(request, f'Spare part "{spare_part.name}" added successfully.')

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': True,
                        'message': f'Spare part "{spare_part.name}" added successfully!',
                        'part_id': spare_part.id
                    })

                return redirect('core:admin_spare_shop')

            except Exception as e:
                error_message = f'Error saving spare part: {str(e)}'
                print(f"Admin error: {error_message}")

                if request.headers.get('HX-Request'):
                    return JsonResponse({
                        'success': False,
                        'message': error_message,
                        'errors': {'__all__': [error_message]}
                    })

                messages.error(request, error_message)
                return redirect('core:admin_spare_shop')
        else:
            # Form validation failed
            print("Admin form validation failed")
            print(f"Form errors: {form.errors}")

            if request.headers.get('HX-Request'):
                # Format errors for JSON response
                formatted_errors = {}
                for field, errors in form.errors.items():
                    if field == '__all__':
                        formatted_errors['__all__'] = [str(error) for error in errors]
                    else:
                        field_label = form.fields.get(field, {}).label or field.replace('_', ' ').title()
                        formatted_errors[field] = [str(error) for error in errors]

                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': formatted_errors
                })

            messages.error(request, 'Please correct the form errors.')
            return redirect('core:admin_spare_shop')


@login_required
def admin_spare_part_edit_modal_view(request, part_id):
    """Modal view for editing spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.get(id=part_id)
        # Get all categories (parent and sub-categories) for hardcoded display
        categories = SparePartCategory.objects.filter(is_active=True).order_by('parent__name', 'name')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        context = {
            'spare_part': spare_part,
            'categories': categories,
            'suppliers': suppliers,
        }

        return render(request, 'core/modals/admin_spare_part_edit.html', context)

    except SparePart.DoesNotExist:
        return JsonResponse({'error': 'Spare part not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def admin_spare_part_edit_view(request, part_id):
    """Handle editing spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.get(id=part_id)

        # Use the SparePartForm for validation
        from .forms import SparePartForm
        form = SparePartForm(request.POST, request.FILES, instance=spare_part)

        if form.is_valid():
            updated_part = form.save()

            # Create stock movement if stock quantity changed
            old_quantity = spare_part.stock_quantity
            new_quantity = updated_part.stock_quantity

            if old_quantity != new_quantity:
                from .models import StockMovement
                StockMovement.objects.create(
                    spare_part=updated_part,
                    movement_type='adjustment',
                    reason='manual_adjustment',
                    quantity=abs(new_quantity - old_quantity),
                    quantity_before=old_quantity,
                    quantity_after=new_quantity,
                    notes=f'Stock adjusted via edit form from {old_quantity} to {new_quantity}',
                    created_by=request.user
                )

            messages.success(request, f'Spare part "{updated_part.name}" updated successfully!')

            # Handle HTMX requests
            if request.headers.get('HX-Request'):
                # Return updated table content
                spare_parts = SparePart.objects.all().order_by('-created_at')[:10]
                context = {'spare_parts': spare_parts}
                return render(request, 'core/partials/admin_spare_shop_table.html', context)

            return redirect('core:admin_spare_shop')
        else:
            # Handle form errors
            if request.headers.get('HX-Request'):
                return JsonResponse({
                    'success': False,
                    'errors': form.errors
                })

            messages.error(request, 'Please correct the errors in the form.')
            return redirect('core:admin_spare_shop')

    except SparePart.DoesNotExist:
        error_msg = 'Spare part not found'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')
    except Exception as e:
        error_msg = f'Error updating spare part: {str(e)}'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')


@login_required
def admin_spare_part_restock_modal_view(request, part_id):
    """Modal view for restocking spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.get(id=part_id)
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')

        context = {
            'spare_part': spare_part,
            'suppliers': suppliers,
        }

        return render(request, 'core/modals/admin_spare_part_restock.html', context)

    except SparePart.DoesNotExist:
        return JsonResponse({'error': 'Spare part not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def admin_spare_part_restock_view(request, part_id):
    """Handle restocking spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.get(id=part_id)

        # Validate input
        quantity = request.POST.get('quantity')
        if not quantity:
            error_msg = 'Quantity is required'
            if request.headers.get('HX-Request'):
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('core:admin_spare_shop')

        try:
            quantity = int(quantity)
        except ValueError:
            error_msg = 'Quantity must be a valid number'
            if request.headers.get('HX-Request'):
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('core:admin_spare_shop')

        if quantity <= 0:
            error_msg = 'Quantity must be greater than 0'
            if request.headers.get('HX-Request'):
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('core:admin_spare_shop')

        # Store old quantity for stock movement
        old_quantity = spare_part.stock_quantity

        # Update stock quantity
        spare_part.stock_quantity += quantity
        spare_part.save()

        # Create stock movement record
        from .models import StockMovement
        movement_type = request.POST.get('movement_type', 'in')
        reason = request.POST.get('reason', 'purchase')
        unit_cost = request.POST.get('unit_cost')
        supplier_id = request.POST.get('supplier') if request.POST.get('supplier') else None
        notes = request.POST.get('notes', '')

        StockMovement.objects.create(
            spare_part=spare_part,
            movement_type=movement_type,
            reason=reason,
            quantity=quantity,
            quantity_before=old_quantity,
            quantity_after=spare_part.stock_quantity,
            unit_cost=unit_cost if unit_cost else None,
            supplier_id=supplier_id,
            notes=notes,
            created_by=request.user
        )

        success_msg = f'Added {quantity} units to "{spare_part.name}". New stock: {spare_part.stock_quantity}'

        # Handle HTMX requests
        if request.headers.get('HX-Request'):
            # Return updated table content
            spare_parts = SparePart.objects.all().order_by('-created_at')[:10]
            context = {'spare_parts': spare_parts}
            return render(request, 'core/partials/admin_spare_shop_table.html', context)

        messages.success(request, success_msg)
        return redirect('core:admin_spare_shop')

    except SparePart.DoesNotExist:
        error_msg = 'Spare part not found'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')
    except Exception as e:
        error_msg = f'Error restocking spare part: {str(e)}'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')


@login_required
def admin_spare_part_view_modal_view(request, part_id):
    """Modal view for viewing spare part details"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.select_related(
            'category_new', 'supplier', 'vendor'
        ).get(id=part_id)

        # Calculate profit margin if both prices are available
        profit_margin = None
        if spare_part.cost_price and spare_part.price and spare_part.cost_price > 0:
            profit_margin = round(((spare_part.price - spare_part.cost_price) / spare_part.cost_price) * 100, 2)

        context = {
            'spare_part': spare_part,
            'profit_margin': profit_margin,
        }

        return render(request, 'core/modals/admin_spare_part_view.html', context)

    except SparePart.DoesNotExist:
        return JsonResponse({'error': 'Spare part not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def admin_spare_part_delete_view(request, part_id):
    """Handle deleting spare part"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        spare_part = SparePart.objects.get(id=part_id)

        # Check if spare part has related orders
        from .models import OrderItem
        related_orders = OrderItem.objects.filter(spare_part=spare_part).exists()

        if related_orders:
            # Don't allow deletion if there are related orders
            error_msg = f'Cannot delete "{spare_part.name}" because it has associated orders. Consider marking it as inactive instead.'
            if request.headers.get('HX-Request'):
                return JsonResponse({'success': False, 'error': error_msg})
            messages.error(request, error_msg)
            return redirect('core:admin_spare_shop')

        part_name = spare_part.name

        # Create a final stock movement record before deletion
        from .models import StockMovement
        if spare_part.stock_quantity > 0:
            StockMovement.objects.create(
                spare_part=spare_part,
                movement_type='out',
                reason='deletion',
                quantity=spare_part.stock_quantity,
                quantity_before=spare_part.stock_quantity,
                quantity_after=0,
                notes=f'Stock cleared due to part deletion: {part_name}',
                created_by=request.user
            )

        # Delete the spare part (this will cascade to related stock movements)
        spare_part.delete()

        success_msg = f'Spare part "{part_name}" deleted successfully!'

        # Handle HTMX requests
        if request.headers.get('HX-Request'):
            # Return updated table content
            spare_parts = SparePart.objects.all().order_by('-created_at')[:10]
            context = {'spare_parts': spare_parts}
            return render(request, 'core/partials/admin_spare_shop_table.html', context)

        messages.success(request, success_msg)
        return redirect('core:admin_spare_shop')

    except SparePart.DoesNotExist:
        error_msg = 'Spare part not found'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')
    except Exception as e:
        error_msg = f'Error deleting spare part: {str(e)}'
        if request.headers.get('HX-Request'):
            return JsonResponse({'success': False, 'error': error_msg})
        messages.error(request, error_msg)
        return redirect('core:admin_spare_shop')


@login_required
def admin_tracking_management_view(request):
    """Admin tracking management view for 7-stage import workflow with pagination"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    # Get import orders with tracking information
    from django.db import transaction

    try:
        with transaction.atomic():
            import_orders = ImportOrder.objects.select_related('customer').prefetch_related('status_history').all().order_by('-created_at')

            # Apply filters
            import_orders = apply_tracking_filters(request, import_orders)

            # Pagination
            page = request.GET.get('page', 1)
            paginator = Paginator(import_orders, 20)  # 20 items per page

            try:
                import_orders_page = paginator.page(page)
            except PageNotAnInteger:
                import_orders_page = paginator.page(1)
            except EmptyPage:
                import_orders_page = paginator.page(paginator.num_pages)

    except Exception as e:
        # Handle database lock gracefully
        print(f"Database error in tracking management: {e}")
        messages.error(request, 'Database temporarily unavailable. Please try again.')
        return redirect('core:admin_dashboard')

    # Define workflow stages for the UI
    workflow_stages = [
        {'key': 'import_request', 'name': 'Import Request', 'icon': 'fas fa-file-import', 'color': 'blue'},
        {'key': 'auction_won', 'name': 'Auction Won', 'icon': 'fas fa-gavel', 'color': 'green'},
        {'key': 'shipped', 'name': 'Shipped', 'icon': 'fas fa-ship', 'color': 'indigo'},
        {'key': 'in_transit', 'name': 'In Transit', 'icon': 'fas fa-route', 'color': 'yellow'},
        {'key': 'arrived_docked', 'name': 'Arrived - Docked', 'icon': 'fas fa-anchor', 'color': 'purple'},
        {'key': 'under_clearance', 'name': 'Under Clearance', 'icon': 'fas fa-file-signature', 'color': 'orange'},
        {'key': 'registered', 'name': 'Registered', 'icon': 'fas fa-certificate', 'color': 'teal'},
        {'key': 'ready_for_dispatch', 'name': 'Ready for Dispatch', 'icon': 'fas fa-truck', 'color': 'pink'},
        {'key': 'delivered', 'name': 'Delivered', 'icon': 'fas fa-flag-checkered', 'color': 'green'},
    ]

    # Calculate tracking stats with optimized single query
    from django.db.models import Count, Case, When, IntegerField
    from django.db import transaction

    try:
        with transaction.atomic():
            # Use aggregation to get all stats in one query
            stats = ImportOrder.objects.aggregate(
                total_orders=Count('id'),
                import_request_orders=Count(Case(When(status='import_request', then=1), output_field=IntegerField())),
                in_transit_orders=Count(Case(When(status__in=['shipped', 'in_transit'], then=1), output_field=IntegerField())),
                arrived_orders=Count(Case(When(status='arrived_docked', then=1), output_field=IntegerField())),
                completed_orders=Count(Case(When(status='delivered', then=1), output_field=IntegerField()))
            )

            total_orders = stats['total_orders']
            import_request_orders = stats['import_request_orders']
            in_transit_orders = stats['in_transit_orders']
            arrived_orders = stats['arrived_orders']
            completed_orders = stats['completed_orders']

    except Exception as e:
        # Fallback to default values if database is locked
        print(f"Database error in tracking stats: {e}")
        total_orders = 0
        import_request_orders = 0
        in_transit_orders = 0
        arrived_orders = 0
        completed_orders = 0

    context = {
        'import_orders': import_orders_page,
        'paginator': paginator,
        'page_obj': import_orders_page,
        'is_paginated': paginator.num_pages > 1,
        'total_orders': total_orders,
        'import_request_orders': import_request_orders,
        'in_transit_orders': in_transit_orders,
        'arrived_orders': arrived_orders,
        'completed_orders': completed_orders,
        'workflow_stages': workflow_stages,
        'current_filter': request.GET.get('status', ''),
        'current_search': request.GET.get('search', ''),
        'current_date_from': request.GET.get('date_from', ''),
        'current_date_to': request.GET.get('date_to', ''),
    }

    return render(request, 'core/dashboard/admin_tracking_management.html', context)


@login_required
def update_tracking_status(request, order_id):
    """Update tracking status for an import order via HTMX"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            order = get_object_or_404(ImportOrder, id=order_id)
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')

            # Update order status
            old_status = order.status
            order.status = new_status
            order.save()

            # Create status history entry
            ImportOrderStatusHistory.objects.create(
                order=order,
                status=new_status,
                notes=notes,
                updated_by=request.user
            )

            # Create notification for customer
            create_notification(
                recipient=order.customer,
                title=f"Import Order Status Updated",
                message=f"Your import order #{order.id:05d} status has been updated to {new_status.replace('_', ' ').title()}",
                notification_type='info',
                action_url=f"/dashboard/orders/{order.id}/",
                action_text="View Order"
            )

            return JsonResponse({
                'success': True,
                'message': f'Status updated from {old_status} to {new_status}',
                'new_status': new_status
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# ===== IMPORT REQUEST MODAL VIEWS =====

@login_required
def admin_import_request_add_modal(request):
    """Show add import request modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    customers = User.objects.filter(role='customer').order_by('username')

    context = {
        'customers': customers,
    }

    return render(request, 'core/modals/admin_import_request_add.html', context)


@login_required
def admin_import_request_add(request):
    """Add new import request via HTMX"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    if request.method == 'POST':
        try:
            # Create new import request
            import_request = ImportRequest.objects.create(
                customer_id=request.POST.get('customer'),
                brand=request.POST.get('brand'),
                model=request.POST.get('model'),
                year=int(request.POST.get('year')),
                preferred_color=request.POST.get('preferred_color', ''),
                origin_country=request.POST.get('origin_country'),
                budget_min=float(request.POST.get('budget_min')),
                budget_max=float(request.POST.get('budget_max')),
                estimated_cost=float(request.POST.get('estimated_cost', 0)) or None,
                special_requirements=request.POST.get('special_requirements', ''),
                admin_notes=request.POST.get('admin_notes', ''),
                status='pending'
            )

            # Create notification for customer
            create_notification(
                recipient=import_request.customer,
                title="New Import Request Created",
                message=f"Your import request for {import_request.year} {import_request.brand} {import_request.model} has been created.",
                notification_type='info',
                action_url=f"/dashboard/import-requests/{import_request.id}/",
                action_text="View Request"
            )

            messages.success(request, f'Import request #{import_request.id:05d} created successfully.')

            # Return updated table
            return admin_import_requests_table_partial(request)

        except Exception as e:
            messages.error(request, f'Error creating import request: {str(e)}')
            return admin_import_requests_table_partial(request)

    return HttpResponse('Invalid request method', status=405)


@login_required
def admin_import_request_view_modal(request, request_id):
    """Show view import request modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_request = get_object_or_404(ImportRequest, id=request_id)

    context = {
        'import_request': import_request,
        'status_choices': ImportRequest.STATUS_CHOICES,
    }

    return render(request, 'core/modals/admin_import_request_view.html', context)


@login_required
def admin_import_request_edit_modal(request, request_id):
    """Show edit import request modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_request = get_object_or_404(ImportRequest, id=request_id)
    customers = User.objects.filter(role='customer').order_by('username')

    context = {
        'import_request': import_request,
        'customers': customers,
        'status_choices': ImportRequest.STATUS_CHOICES,
    }

    return render(request, 'core/modals/admin_import_request_edit.html', context)


@login_required
def admin_import_request_edit(request, request_id):
    """Edit import request via HTMX"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    if request.method == 'POST':
        try:
            import_request = get_object_or_404(ImportRequest, id=request_id)

            # Update import request
            import_request.customer_id = request.POST.get('customer')
            import_request.brand = request.POST.get('brand')
            import_request.model = request.POST.get('model')
            import_request.year = int(request.POST.get('year'))
            import_request.preferred_color = request.POST.get('preferred_color', '')
            import_request.origin_country = request.POST.get('origin_country')
            import_request.budget_min = float(request.POST.get('budget_min'))
            import_request.budget_max = float(request.POST.get('budget_max'))
            import_request.status = request.POST.get('status')

            estimated_cost = request.POST.get('estimated_cost')
            if estimated_cost:
                import_request.estimated_cost = float(estimated_cost)

            estimated_delivery = request.POST.get('estimated_delivery')
            if estimated_delivery:
                from datetime import datetime
                import_request.estimated_delivery = datetime.strptime(estimated_delivery, '%Y-%m-%d').date()

            import_request.tracking_number = request.POST.get('tracking_number', '')
            import_request.special_requirements = request.POST.get('special_requirements', '')
            import_request.admin_notes = request.POST.get('admin_notes', '')

            import_request.save()

            # Create notification for customer if status changed
            if 'status' in request.POST:
                create_notification(
                    recipient=import_request.customer,
                    title="Import Request Updated",
                    message=f"Your import request #{import_request.id:05d} has been updated. Status: {import_request.get_status_display()}",
                    notification_type='info',
                    action_url=f"/dashboard/import-requests/{import_request.id}/",
                    action_text="View Request"
                )

            messages.success(request, f'Import request #{import_request.id:05d} updated successfully.')

            # Return updated table
            return admin_import_requests_table_partial(request)

        except Exception as e:
            messages.error(request, f'Error updating import request: {str(e)}')
            return admin_import_requests_table_partial(request)

    return HttpResponse('Invalid request method', status=405)


@login_required
def admin_import_request_delete_modal(request, request_id):
    """Show delete import request confirmation modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_request = get_object_or_404(ImportRequest, id=request_id)

    context = {
        'import_request': import_request,
    }

    return render(request, 'core/modals/admin_import_request_delete.html', context)


@login_required
def admin_import_request_delete(request, request_id):
    """Delete import request via HTMX"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    # Handle both DELETE requests and POST requests with _method=DELETE
    if request.method == 'DELETE' or (request.method == 'POST' and request.POST.get('_method') == 'DELETE'):
        try:
            import_request = get_object_or_404(ImportRequest, id=request_id)
            customer = import_request.customer
            request_info = f"#{import_request.id:05d} - {import_request.year} {import_request.brand} {import_request.model}"

            # Delete the import request
            import_request.delete()

            # Create notification for customer
            create_notification(
                recipient=customer,
                title="Import Request Cancelled",
                message=f"Your import request {request_info} has been cancelled by admin.",
                notification_type='warning'
            )

            messages.success(request, f'Import request {request_info} deleted successfully.')

            # Return updated table
            return admin_import_requests_table_partial(request)

        except Exception as e:
            messages.error(request, f'Error deleting import request: {str(e)}')
            return admin_import_requests_table_partial(request)

    return HttpResponse('Invalid request method', status=405)


@login_required
def admin_import_request_track(request, request_id):
    """Move import request to tracking management"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    if request.method == 'POST':
        try:
            import_request = get_object_or_404(ImportRequest, id=request_id)

            # Check if already has an import order
            if hasattr(import_request, 'import_order'):
                messages.warning(request, f'Import request #{import_request.id:05d} is already being tracked.')
                return admin_import_requests_table_partial(request)

            # Create import order for tracking with comprehensive data transfer
            import_order = ImportOrder.objects.create(
                import_request=import_request,
                customer=import_request.customer,
                brand=import_request.brand,
                model=import_request.model,
                year=import_request.year,
                color=import_request.preferred_color or '',
                origin_country=import_request.origin_country,
                quotation_amount=import_request.estimated_cost or 0,
                total_cost=import_request.estimated_cost or 0,
                special_requirements=import_request.special_requirements,
                admin_notes=import_request.admin_notes,
                status='confirmed',  # Start with confirmed status in 7-stage workflow
                payment_status='pending'
            )

            # Keep import request as completed but mark it as being tracked
            import_request.tracking_number = import_order.order_number
            import_request.save()

            # Create initial status history entry with proper field names
            ImportOrderStatusHistory.objects.create(
                import_order=import_order,
                previous_status='',  # No previous status for initial entry
                new_status='confirmed',
                changed_by=request.user,
                change_reason='Import request moved to tracking management - 7-stage workflow initiated',
                admin_notes=f'Converted from Import Request #{import_request.id:05d}',
                customer_notification_sent=True,
                actual_date=timezone.now().date()
            )

            # Create notification for customer
            create_notification(
                recipient=import_request.customer,
                title="Import Request Now Being Tracked",
                message=f"Your import request #{import_request.id:05d} is now being tracked. Order number: {import_order.order_number}",
                notification_type='success',
                action_url=f"/import/tracking/{import_order.order_number}/",
                action_text="Track Order"
            )

            messages.success(request, f'Import request #{import_request.id:05d} moved to tracking management. Order number: {import_order.order_number}')

            # For HTMX requests, return a response that triggers both table updates
            if request.headers.get('HX-Request'):
                # Return the import requests table with HX-Trigger header to update tracking table
                response = admin_import_requests_table_partial(request)
                response['HX-Trigger'] = 'updateTrackingTable'
                return response
            else:
                # For non-HTMX requests, redirect to import requests page
                return redirect('core:admin_import_requests')

        except Exception as e:
            messages.error(request, f'Error moving to tracking: {str(e)}')
            return admin_import_requests_table_partial(request)

    return HttpResponse('Invalid request method', status=405)


@login_required
def admin_import_request_status_modal(request, request_id):
    """Show status update modal"""
    if request.user.role != 'admin':
        return HttpResponse('Access denied', status=403)

    import_request = get_object_or_404(ImportRequest, id=request_id)

    context = {
        'import_request': import_request,
    }

    return render(request, 'core/modals/admin_import_request_status_update.html', context)


@login_required
def admin_import_request_status_update(request, request_id):
    """Update import request status with workflow management"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            import_request = get_object_or_404(ImportRequest, id=request_id)
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')

            # Define 6-stage status workflow (removed shipped and arrived)
            status_workflow = {
                'pending': ['on_quotation', 'cancelled'],
                'on_quotation': ['processing', 'pending', 'cancelled'],
                'processing': ['fee_paid', 'on_quotation', 'cancelled'],
                'fee_paid': ['completed', 'processing', 'cancelled'],
                'completed': ['fee_paid'],  # Allow going back for corrections
                'cancelled': ['pending']   # Allow reactivation
            }

            # Validate status transition
            current_status = import_request.status
            if new_status not in status_workflow.get(current_status, []):
                messages.error(request, f'Invalid status transition from {current_status} to {new_status}')
                return admin_import_requests_view(request)

            # Update status
            old_status = import_request.status
            import_request.status = new_status
            import_request.save()

            # Create notification for customer
            status_messages = {
                'pending': 'Your import request is pending review.',
                'on_quotation': 'We are preparing a quotation for your import request.',
                'processing': 'Your import request is now being processed.',
                'fee_paid': 'Payment received. Your order is being prepared for import.',
                'completed': 'Your import request has been completed successfully.',
                'cancelled': 'Your import request has been cancelled.'
            }

            create_notification(
                recipient=import_request.customer,
                title=f"Import Request Status Updated",
                message=f"Request #{import_request.id:05d}: {status_messages.get(new_status, f'Status updated to {new_status}')}",
                notification_type='info' if new_status != 'cancelled' else 'warning',
                action_url=f"/dashboard/import-requests/{import_request.id}/",
                action_text="View Request"
            )

            messages.success(request, f'Status updated from {old_status} to {new_status}')

            # Return updated table HTML for HTMX
            return admin_import_requests_table_partial(request)

        except Exception as e:
            messages.error(request, f'Error updating status: {str(e)}')
            return admin_import_requests_table_partial(request)

    return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)


# ===== ENHANCED CONTENT MANAGEMENT VIEWS =====

@login_required
def admin_content_posts_tab(request):
    """HTMX endpoint for content posts tab"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    content_type = request.GET.get('content_type', '')
    status = request.GET.get('status', '')
    category_id = request.GET.get('category', '')

    # Base queryset
    posts = BlogPost.objects.select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if search_query:
        posts = posts.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query)
        )

    if content_type:
        posts = posts.filter(content_type=content_type)

    if status == 'published':
        posts = posts.filter(is_published=True)
    elif status == 'draft':
        posts = posts.filter(is_published=False)
    elif status == 'featured':
        posts = posts.filter(is_featured=True)

    if category_id:
        posts = posts.filter(category_id=category_id)

    # Order by creation date
    posts = posts.order_by('-created_at')

    # Pagination
    paginator = Paginator(posts, 20)  # 20 items per page
    page = request.GET.get('page', 1)

    try:
        posts_page = paginator.page(page)
    except PageNotAnInteger:
        posts_page = paginator.page(1)
    except EmptyPage:
        posts_page = paginator.page(paginator.num_pages)

    # Get filter options
    from .models import ContentCategory
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')
    content_types = BlogPost.CONTENT_TYPE_CHOICES

    context = {
        'posts': posts_page,
        'categories': categories,
        'content_types': content_types,
        'current_search': search_query,
        'current_content_type': content_type,
        'current_status': status,
        'current_category': category_id,
    }

    return render(request, 'core/dashboard/partials/content_posts_tab.html', context)


@login_required
def admin_content_categories_tab(request):
    """HTMX endpoint for content categories tab"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import ContentCategory

    # Get search parameter
    search_query = request.GET.get('search', '').strip()

    # Base queryset
    categories = ContentCategory.objects.annotate(
        posts_count=models.Count('posts')
    )

    # Apply search filter
    if search_query:
        categories = categories.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Order by sort order and name
    categories = categories.order_by('sort_order', 'name')

    # Pagination
    paginator = Paginator(categories, 20)
    page = request.GET.get('page', 1)

    try:
        categories_page = paginator.page(page)
    except PageNotAnInteger:
        categories_page = paginator.page(1)
    except EmptyPage:
        categories_page = paginator.page(paginator.num_pages)

    context = {
        'categories': categories_page,
        'current_search': search_query,
    }

    return render(request, 'core/dashboard/partials/content_categories_tab.html', context)


@login_required
def admin_content_testimonials_tab(request):
    """HTMX endpoint for testimonials tab"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    status = request.GET.get('status', '')
    rating = request.GET.get('rating', '')

    # Base queryset
    testimonials = Testimonial.objects.select_related('customer', 'car')

    # Apply filters
    if search_query:
        testimonials = testimonials.filter(
            Q(content__icontains=search_query) |
            Q(customer__username__icontains=search_query) |
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query)
        )

    if status == 'approved':
        testimonials = testimonials.filter(is_approved=True)
    elif status == 'pending':
        testimonials = testimonials.filter(is_approved=False)
    elif status == 'featured':
        testimonials = testimonials.filter(is_featured=True)

    if rating:
        testimonials = testimonials.filter(rating=rating)

    # Order by creation date
    testimonials = testimonials.order_by('-created_at')

    # Pagination
    paginator = Paginator(testimonials, 20)
    page = request.GET.get('page', 1)

    try:
        testimonials_page = paginator.page(page)
    except PageNotAnInteger:
        testimonials_page = paginator.page(1)
    except EmptyPage:
        testimonials_page = paginator.page(paginator.num_pages)

    context = {
        'testimonials': testimonials_page,
        'current_search': search_query,
        'current_status': status,
        'current_rating': rating,
        'rating_choices': [(i, i) for i in range(1, 6)],
    }

    return render(request, 'core/dashboard/partials/content_testimonials_tab.html', context)


@login_required
def admin_content_static_pages_tab(request):
    """HTMX endpoint for static pages tab"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import StaticPage

    # Get search and filter parameters
    search_query = request.GET.get('search', '').strip()
    page_type = request.GET.get('page_type', '')
    status = request.GET.get('status', '')

    # Base queryset
    pages = StaticPage.objects.select_related('author')

    # Apply filters
    if search_query:
        pages = pages.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query)
        )

    if page_type:
        pages = pages.filter(page_type=page_type)

    if status:
        pages = pages.filter(status=status)

    # Order by menu order and title
    pages = pages.order_by('menu_order', 'title')

    # Pagination
    paginator = Paginator(pages, 20)
    page_num = request.GET.get('page', 1)

    try:
        pages_page = paginator.page(page_num)
    except PageNotAnInteger:
        pages_page = paginator.page(1)
    except EmptyPage:
        pages_page = paginator.page(paginator.num_pages)

    context = {
        'pages': pages_page,
        'page_types': StaticPage.PAGE_TYPE_CHOICES,
        'status_choices': StaticPage.STATUS_CHOICES,
        'current_search': search_query,
        'current_page_type': page_type,
        'current_status': status,
    }

    return render(request, 'core/dashboard/partials/content_static_pages_tab.html', context)


@login_required
def admin_content_analytics_tab(request):
    """HTMX endpoint for content analytics tab"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import ContentAnalytics, ContentView
    from datetime import datetime, timedelta
    from django.db.models import Sum, Avg, Count

    # Get date range parameters
    days = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=days)

    # Top performing posts
    top_posts = BlogPost.objects.filter(
        is_published=True,
        published_at__gte=start_date
    ).annotate(
        total_views=Count('content_views'),
        total_likes=Count('content_likes'),
        total_comments=Count('comments'),
        total_bookmarks=Count('bookmarks')
    ).order_by('-total_views')[:10]

    # Content type performance
    content_type_stats = BlogPost.objects.filter(
        is_published=True
    ).values('content_type').annotate(
        count=Count('id'),
        avg_views=Avg('views_count'),
        total_views=Sum('views_count'),
        avg_likes=Avg('likes_count')
    ).order_by('-total_views')

    # Recent activity
    recent_views = ContentView.objects.select_related('post', 'user').order_by('-viewed_at')[:20]

    # Monthly trends (simplified)
    monthly_stats = BlogPost.objects.filter(
        published_at__gte=start_date
    ).extra(
        select={'month': "DATE_FORMAT(published_at, '%%Y-%%m')"}
    ).values('month').annotate(
        posts_count=Count('id'),
        total_views=Sum('views_count')
    ).order_by('month')

    context = {
        'top_posts': top_posts,
        'content_type_stats': content_type_stats,
        'recent_views': recent_views,
        'monthly_stats': monthly_stats,
        'date_range_days': days,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'core/dashboard/partials/content_analytics_tab.html', context)


# Content CRUD Operations

@login_required
def admin_content_create_modal(request):
    """Display create content modal"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import ContentCategory, ContentTag

    categories = ContentCategory.objects.filter(is_active=True).order_by('name')
    tags = ContentTag.objects.filter(is_active=True).order_by('name')
    content_types = BlogPost.CONTENT_TYPE_CHOICES

    context = {
        'categories': categories,
        'tags': tags,
        'content_types': content_types,
    }

    return render(request, 'core/modals/admin_content_create.html', context)


@login_required
def admin_content_create(request):
    """Create new content post"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        from django.utils.text import slugify
        from .models import ContentCategory, ContentTag

        try:
            # Get form data
            title = request.POST.get('title', '').strip()
            content = request.POST.get('content', '').strip()
            excerpt = request.POST.get('excerpt', '').strip()
            content_type = request.POST.get('content_type', 'article')
            category_id = request.POST.get('category')
            tag_ids = request.POST.getlist('tags')
            is_published = request.POST.get('is_published') == 'on'
            is_featured = request.POST.get('is_featured') == 'on'

            # Validation
            if not title or not content:
                return JsonResponse({'success': False, 'error': 'Title and content are required'})

            # Create slug
            slug = slugify(title)
            original_slug = slug
            counter = 1
            while BlogPost.objects.filter(slug=slug).exists():
                slug = f"{original_slug}-{counter}"
                counter += 1

            # Create post
            post = BlogPost.objects.create(
                title=title,
                slug=slug,
                content=content,
                excerpt=excerpt,
                content_type=content_type,
                author=request.user,
                is_published=is_published,
                is_featured=is_featured,
                published_at=timezone.now() if is_published else None
            )

            # Set category
            if category_id:
                try:
                    category = ContentCategory.objects.get(id=category_id)
                    post.category = category
                    post.save()
                except ContentCategory.DoesNotExist:
                    pass

            # Set tags
            if tag_ids:
                tags = ContentTag.objects.filter(id__in=tag_ids)
                post.tags.set(tags)

            return JsonResponse({
                'success': True,
                'message': 'Content created successfully',
                'post_id': post.id
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def admin_sales_management_view(request):
    """Admin sales management view"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta
    from .models import Order, OrderItem, Payment, SparePart

    # Calculate date ranges
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    # Sales statistics
    total_sales = Order.objects.filter(
        items__spare_part__isnull=False,
        status='completed'
    ).aggregate(
        total_amount=Sum('total_amount'),
        total_orders=Count('id', distinct=True)
    )

    weekly_sales = Order.objects.filter(
        items__spare_part__isnull=False,
        status='completed',
        created_at__gte=week_ago
    ).aggregate(
        total_amount=Sum('total_amount'),
        total_orders=Count('id', distinct=True)
    )

    monthly_sales = Order.objects.filter(
        items__spare_part__isnull=False,
        status='completed',
        created_at__gte=month_ago
    ).aggregate(
        total_amount=Sum('total_amount'),
        total_orders=Count('id', distinct=True)
    )

    # Recent sales
    recent_sales = Order.objects.filter(
        items__spare_part__isnull=False
    ).select_related('customer').prefetch_related('items__spare_part').order_by('-created_at')[:10]

    # Top selling parts
    top_selling_parts = OrderItem.objects.filter(
        spare_part__isnull=False,
        order__status='completed'
    ).values(
        'spare_part__name', 'spare_part__id'
    ).annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_sold')[:10]

    context = {
        'total_sales': total_sales,
        'weekly_sales': weekly_sales,
        'monthly_sales': monthly_sales,
        'recent_sales': recent_sales,
        'top_selling_parts': top_selling_parts,
    }

    return render(request, 'core/dashboard/admin_sales_management.html', context)


@login_required
def admin_order_management_view(request):
    """Admin order management view"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied.')
        return redirect('core:dashboard')

    from django.db.models import Sum, Count, Q
    from datetime import datetime, timedelta
    from .models import Order, OrderItem, Payment, SparePart

    # Filter parameters
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')

    # Base queryset for spare parts orders
    orders = Order.objects.filter(
        items__spare_part__isnull=False
    ).select_related('customer').prefetch_related('items__spare_part', 'payments').distinct()

    # Apply filters
    if status_filter:
        orders = orders.filter(status=status_filter)

    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer__first_name__icontains=search_query) |
            Q(customer__last_name__icontains=search_query) |
            Q(customer__email__icontains=search_query)
        )

    # Order statistics
    order_stats = {
        'total_orders': orders.count(),
        'pending_orders': orders.filter(status='pending').count(),
        'processing_orders': orders.filter(status='processing').count(),
        'completed_orders': orders.filter(status='completed').count(),
        'cancelled_orders': orders.filter(status='cancelled').count(),
    }

    # Recent orders (paginated)
    from django.core.paginator import Paginator
    paginator = Paginator(orders.order_by('-created_at'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'orders': page_obj,
        'order_stats': order_stats,
        'status_filter': status_filter,
        'search_query': search_query,
        'status_choices': Order.STATUS_CHOICES,
    }

    return render(request, 'core/dashboard/admin_order_management.html', context)


# Resource Management Views

@login_required
def admin_resource_management_view(request):
    """Main Resource Management page with tabbed interface"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('core:dashboard')

    # Get overview statistics
    total_content = BlogPost.objects.count()
    published_content = BlogPost.objects.filter(is_published=True).count()
    draft_content = BlogPost.objects.filter(is_published=False).count()
    featured_content = BlogPost.objects.filter(is_featured=True).count()

    # Content type breakdown
    content_types = BlogPost.objects.values('content_type').annotate(
        count=Count('id')
    ).order_by('content_type')

    # Recent activity
    recent_content = BlogPost.objects.select_related('author', 'category').order_by('-updated_at')[:10]

    # Top performing content (by views)
    top_content = BlogPost.objects.filter(is_published=True).order_by('-views_count')[:5]

    # Categories and tags
    categories = ContentCategory.objects.filter(is_active=True).annotate(
        post_count=Count('posts')
    ).order_by('-post_count')[:10]

    tags = ContentTag.objects.filter(is_active=True).annotate(
        post_count=Count('posts')
    ).order_by('-post_count')[:15]

    context = {
        'total_content': total_content,
        'published_content': published_content,
        'draft_content': draft_content,
        'featured_content': featured_content,
        'content_types': content_types,
        'recent_content': recent_content,
        'top_content': top_content,
        'categories': categories,
        'tags': tags,
    }

    return render(request, 'core/dashboard/admin_resource_management.html', context)


@login_required
def admin_resource_all_content_tab(request):
    """HTMX endpoint for All Content tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    content_type = request.GET.get('content_type', '')
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    queryset = BlogPost.objects.select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if content_type:
        queryset = queryset.filter(content_type=content_type)

    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query) |
            Q(tags__name__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    posts = paginator.get_page(page_number)

    # Get filter options
    content_types = BlogPost.CONTENT_TYPE_CHOICES
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'posts': posts,
        'content_types': content_types,
        'categories': categories,
        'current_content_type': content_type,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
    }

    return render(request, 'core/dashboard/partials/admin_resource_all_content_tab.html', context)


@login_required
def admin_resource_articles_tab(request):
    """HTMX endpoint for Articles tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset for articles
    queryset = BlogPost.objects.filter(content_type='article').select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    articles = paginator.get_page(page_number)

    # Get filter options
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'articles': articles,
        'categories': categories,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
        'content_type': 'article',
    }

    return render(request, 'core/dashboard/partials/admin_resource_articles_tab.html', context)


@login_required
def admin_resource_guides_tab(request):
    """HTMX endpoint for Guides tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset for guides
    queryset = BlogPost.objects.filter(content_type='guide').select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    guides = paginator.get_page(page_number)

    # Get filter options
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'guides': guides,
        'categories': categories,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
        'content_type': 'guide',
    }

    return render(request, 'core/dashboard/partials/admin_resource_guides_tab.html', context)


@login_required
def admin_resource_infographics_tab(request):
    """HTMX endpoint for Infographics tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset for infographics
    queryset = BlogPost.objects.filter(content_type='infographic').select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    infographics = paginator.get_page(page_number)

    # Get filter options
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'infographics': infographics,
        'categories': categories,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
        'content_type': 'infographic',
    }

    return render(request, 'core/dashboard/partials/admin_resource_infographics_tab.html', context)


@login_required
def admin_resource_opinions_tab(request):
    """HTMX endpoint for Opinions tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset for opinions
    queryset = BlogPost.objects.filter(content_type='opinion').select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    opinions = paginator.get_page(page_number)

    # Get filter options
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'opinions': opinions,
        'categories': categories,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
        'content_type': 'opinion',
    }

    return render(request, 'core/dashboard/partials/admin_resource_opinions_tab.html', context)


@login_required
def admin_resource_news_tab(request):
    """HTMX endpoint for News tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    # Get filter parameters
    status = request.GET.get('status', '')
    category = request.GET.get('category', '')
    search_query = request.GET.get('search', '')

    # Base queryset for news
    queryset = BlogPost.objects.filter(content_type='news').select_related('author', 'category').prefetch_related('tags')

    # Apply filters
    if status == 'published':
        queryset = queryset.filter(is_published=True)
    elif status == 'draft':
        queryset = queryset.filter(is_published=False)
    elif status == 'featured':
        queryset = queryset.filter(is_featured=True)

    if category:
        queryset = queryset.filter(category__slug=category)

    if search_query:
        queryset = queryset.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query) |
            Q(author__username__icontains=search_query)
        ).distinct()

    # Order by latest
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    news = paginator.get_page(page_number)

    # Get filter options
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')

    context = {
        'news': news,
        'categories': categories,
        'current_status': status,
        'current_category': category,
        'search_query': search_query,
        'content_type': 'news',
    }

    return render(request, 'core/dashboard/partials/admin_resource_news_tab.html', context)


# Resource CRUD Operations

@login_required
def admin_resource_create_modal(request):
    """HTMX endpoint for resource creation modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    content_type = request.GET.get('content_type', 'article')

    # Get categories and tags for form
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')
    tags = ContentTag.objects.filter(is_active=True).order_by('name')

    context = {
        'content_type': content_type,
        'categories': categories,
        'tags': tags,
        'content_type_choices': BlogPost.CONTENT_TYPE_CHOICES,
        'difficulty_choices': BlogPost.DIFFICULTY_CHOICES,
    }

    return render(request, 'core/modals/admin_resource_create.html', context)


@login_required
def admin_resource_create(request):
    """Handle resource creation"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            # Create new blog post
            post = BlogPost()
            post.author = request.user
            post.title = request.POST.get('title', '')
            post.content = request.POST.get('content', '')
            post.excerpt = request.POST.get('excerpt', '')
            post.content_type = request.POST.get('content_type', 'article')
            post.difficulty_level = request.POST.get('difficulty_level', '')
            post.estimated_read_time = int(request.POST.get('estimated_read_time', 5))
            post.meta_description = request.POST.get('meta_description', '')
            post.meta_keywords = request.POST.get('meta_keywords', '')
            post.video_url = request.POST.get('video_url', '')
            post.is_published = request.POST.get('is_published') == 'on'
            post.is_featured = request.POST.get('is_featured') == 'on'

            # Generate slug
            from django.utils.text import slugify
            base_slug = slugify(post.title)
            slug = base_slug
            counter = 1
            while BlogPost.objects.filter(slug=slug).exists():
                slug = f"{base_slug}-{counter}"
                counter += 1
            post.slug = slug

            # Handle category
            category_id = request.POST.get('category')
            if category_id:
                try:
                    post.category = ContentCategory.objects.get(id=category_id)
                except ContentCategory.DoesNotExist:
                    pass

            # Handle featured image
            if 'featured_image' in request.FILES:
                post.featured_image = request.FILES['featured_image']
                post.featured_image_alt = request.POST.get('featured_image_alt', '')

            # Set published date if publishing
            if post.is_published and not post.published_at:
                post.published_at = timezone.now()

            post.save()

            # Handle tags
            tag_ids = request.POST.getlist('tags')
            if tag_ids:
                tags = ContentTag.objects.filter(id__in=tag_ids)
                post.tags.set(tags)

            return JsonResponse({
                'success': True,
                'message': f'{post.content_type.title()} created successfully!',
                'redirect': reverse('core:admin_resource_management')
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error creating content: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def admin_resource_edit_modal(request, post_id):
    """HTMX endpoint for resource edit modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    post = get_object_or_404(BlogPost, id=post_id)

    # Get categories and tags for form
    categories = ContentCategory.objects.filter(is_active=True).order_by('name')
    tags = ContentTag.objects.filter(is_active=True).order_by('name')

    context = {
        'post': post,
        'categories': categories,
        'tags': tags,
        'content_type_choices': BlogPost.CONTENT_TYPE_CHOICES,
        'difficulty_choices': BlogPost.DIFFICULTY_CHOICES,
    }

    return render(request, 'core/modals/admin_resource_edit.html', context)


@login_required
def admin_resource_edit(request, post_id):
    """Handle resource editing"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    post = get_object_or_404(BlogPost, id=post_id)

    if request.method == 'POST':
        try:
            # Update blog post
            post.title = request.POST.get('title', post.title)
            post.content = request.POST.get('content', post.content)
            post.excerpt = request.POST.get('excerpt', post.excerpt)
            post.content_type = request.POST.get('content_type', post.content_type)
            post.difficulty_level = request.POST.get('difficulty_level', post.difficulty_level)
            post.estimated_read_time = int(request.POST.get('estimated_read_time', post.estimated_read_time))
            post.meta_description = request.POST.get('meta_description', post.meta_description)
            post.meta_keywords = request.POST.get('meta_keywords', post.meta_keywords)
            post.video_url = request.POST.get('video_url', post.video_url)
            post.is_published = request.POST.get('is_published') == 'on'
            post.is_featured = request.POST.get('is_featured') == 'on'
            post.featured_image_alt = request.POST.get('featured_image_alt', post.featured_image_alt)

            # Handle category
            category_id = request.POST.get('category')
            if category_id:
                try:
                    post.category = ContentCategory.objects.get(id=category_id)
                except ContentCategory.DoesNotExist:
                    post.category = None
            else:
                post.category = None

            # Handle featured image
            if 'featured_image' in request.FILES:
                post.featured_image = request.FILES['featured_image']

            # Set published date if publishing for the first time
            if post.is_published and not post.published_at:
                post.published_at = timezone.now()

            post.save()

            # Handle tags
            tag_ids = request.POST.getlist('tags')
            if tag_ids:
                tags = ContentTag.objects.filter(id__in=tag_ids)
                post.tags.set(tags)
            else:
                post.tags.clear()

            return JsonResponse({
                'success': True,
                'message': f'{post.content_type.title()} updated successfully!',
                'redirect': reverse('core:admin_resource_management')
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error updating content: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def admin_resource_delete(request, post_id):
    """Handle resource deletion"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            post = get_object_or_404(BlogPost, id=post_id)
            content_type = post.content_type
            title = post.title

            post.delete()

            return JsonResponse({
                'success': True,
                'message': f'{content_type.title()} "{title}" deleted successfully!'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error deleting content: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def admin_resource_toggle_featured(request, post_id):
    """Toggle featured status of a resource"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            post = get_object_or_404(BlogPost, id=post_id)
            post.is_featured = not post.is_featured
            post.save()

            status = 'featured' if post.is_featured else 'unfeatured'
            return JsonResponse({
                'success': True,
                'message': f'Content {status} successfully!',
                'is_featured': post.is_featured
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error updating featured status: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def admin_resource_toggle_published(request, post_id):
    """Toggle published status of a resource"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        try:
            post = get_object_or_404(BlogPost, id=post_id)
            post.is_published = not post.is_published

            # Set published date if publishing for the first time
            if post.is_published and not post.published_at:
                post.published_at = timezone.now()

            post.save()

            status = 'published' if post.is_published else 'unpublished'
            return JsonResponse({
                'success': True,
                'message': f'Content {status} successfully!',
                'is_published': post.is_published
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error updating published status: {str(e)}'
            })

    return JsonResponse({'success': False, 'message': 'Invalid request method'})


@login_required
def admin_resource_statistics(request):
    """HTMX endpoint for refreshing resource management statistics"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        # Get current statistics
        total_content = BlogPost.objects.count()
        published_content = BlogPost.objects.filter(is_published=True).count()
        draft_content = BlogPost.objects.filter(is_published=False).count()
        featured_content = BlogPost.objects.filter(is_featured=True).count()

        stats = {
            'total_content': total_content,
            'published_content': published_content,
            'draft_content': draft_content,
            'featured_content': featured_content,
        }

        return JsonResponse({
            'success': True,
            'stats': stats
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def admin_resource_export(request):
    """Export resource management data to Excel"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.utils import timezone
        import io

        # Create workbook
        wb = openpyxl.Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets for different content types
        sheets_data = {
            'All Content': BlogPost.objects.all(),
            'Articles': BlogPost.objects.filter(content_type='article'),
            'Guides': BlogPost.objects.filter(content_type='guide'),
            'Infographics': BlogPost.objects.filter(content_type='infographic'),
            'Opinions': BlogPost.objects.filter(content_type='opinion'),
            'News': BlogPost.objects.filter(content_type='news'),
        }

        for sheet_name, queryset in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)

            # Headers
            headers = [
                'ID', 'Title', 'Content Type', 'Author', 'Category',
                'Status', 'Featured', 'Views', 'Likes', 'Comments',
                'Created Date', 'Updated Date', 'Tags'
            ]

            # Style headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Data rows
            for row, post in enumerate(queryset.select_related('author', 'category').prefetch_related('tags'), 2):
                ws.cell(row=row, column=1, value=post.id)
                ws.cell(row=row, column=2, value=post.title)
                ws.cell(row=row, column=3, value=post.get_content_type_display())
                ws.cell(row=row, column=4, value=post.author.get_full_name() if post.author else 'N/A')
                ws.cell(row=row, column=5, value=post.category.name if post.category else 'N/A')
                ws.cell(row=row, column=6, value='Published' if post.is_published else 'Draft')
                ws.cell(row=row, column=7, value='Yes' if post.is_featured else 'No')
                ws.cell(row=row, column=8, value=post.views_count or 0)
                ws.cell(row=row, column=9, value=post.likes_count or 0)
                ws.cell(row=row, column=10, value=post.comment_count or 0)
                ws.cell(row=row, column=11, value=post.created_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=12, value=post.updated_at.strftime('%Y-%m-%d %H:%M'))
                ws.cell(row=row, column=13, value=', '.join([tag.name for tag in post.tags.all()]))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f'gurumisha_resources_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    except ImportError:
        return JsonResponse({
            'error': 'openpyxl library not installed. Please install it to use export functionality.'
        }, status=500)
    except Exception as e:
        return JsonResponse({
            'error': f'Export failed: {str(e)}'
        }, status=500)


@login_required
def admin_resource_toggle_publish(request, post_id):
    """Toggle publish status of a resource"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        post = get_object_or_404(BlogPost, id=post_id)

        # Toggle publish status
        post.is_published = not post.is_published
        post.save()

        # Return the updated tab content
        return admin_resource_all_content_tab(request)

    except Exception as e:
        return JsonResponse({
            'error': f'Failed to toggle publish status: {str(e)}'
        }, status=500)


@login_required
def admin_resource_search(request):
    """HTMX endpoint for resource search across all tabs"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    search_query = request.GET.get('search', '')
    content_type = request.GET.get('content_type', '')
    tab = request.GET.get('tab', 'all')

    if not search_query:
        return JsonResponse({'success': False, 'message': 'No search query provided'})

    # Base queryset
    queryset = BlogPost.objects.select_related('author', 'category').prefetch_related('tags')

    # Filter by content type if specified
    if content_type:
        queryset = queryset.filter(content_type=content_type)

    # Apply search filter
    queryset = queryset.filter(
        Q(title__icontains=search_query) |
        Q(content__icontains=search_query) |
        Q(excerpt__icontains=search_query) |
        Q(author__username__icontains=search_query) |
        Q(tags__name__icontains=search_query)
    ).distinct()

    # Order by relevance (title matches first, then content)
    queryset = queryset.order_by('-updated_at')

    # Pagination
    paginator = Paginator(queryset, 20)
    page_number = request.GET.get('page', 1)
    results = paginator.get_page(page_number)

    # Determine which template to use based on tab
    template_map = {
        'all': 'core/dashboard/partials/admin_resource_all_content_tab.html',
        'articles': 'core/dashboard/partials/admin_resource_articles_tab.html',
        'guides': 'core/dashboard/partials/admin_resource_guides_tab.html',
        'infographics': 'core/dashboard/partials/admin_resource_infographics_tab.html',
        'opinions': 'core/dashboard/partials/admin_resource_opinions_tab.html',
        'news': 'core/dashboard/partials/admin_resource_news_tab.html',
    }

    template = template_map.get(tab, template_map['all'])

    context = {
        'posts' if tab == 'all' else tab: results,
        'search_query': search_query,
        'content_type': content_type,
    }

    return render(request, template, context)


@login_required
def admin_mpesa_config_modal_view(request):
    """M-Pesa configuration modal view"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    context = {
        # Add M-Pesa configuration context here
    }

    return render(request, 'core/modals/admin_mpesa_config.html', context)


# Message Management Views

@login_required
def admin_message_management_view(request):
    """Main Message Management page with comprehensive messaging system"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('core:dashboard')

    from .models import Message, MessageAnalytics
    from django.db.models import Count, Sum, Avg
    from datetime import timedelta

    # Get overview statistics
    total_messages = Message.objects.count()
    active_messages = Message.objects.filter(status='active').count()
    scheduled_messages = Message.objects.filter(status='scheduled').count()
    draft_messages = Message.objects.filter(status='draft').count()

    # Message type breakdown
    message_types = Message.objects.values('message_type').annotate(
        count=Count('id')
    ).order_by('message_type')

    # Recent performance metrics (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_analytics = MessageAnalytics.objects.filter(
        date__gte=thirty_days_ago.date()
    ).aggregate(
        total_displays=Sum('total_displays'),
        total_clicks=Sum('total_clicks'),
        total_dismissals=Sum('total_dismissals'),
    )

    # Calculate CTR manually
    total_displays = recent_analytics.get('total_displays', 0) or 0
    total_clicks = recent_analytics.get('total_clicks', 0) or 0
    avg_ctr = (total_clicks / total_displays * 100) if total_displays > 0 else 0
    recent_analytics['avg_ctr'] = round(avg_ctr, 2)

    # Top performing messages
    top_messages = Message.objects.filter(
        status='active',
        total_views__gt=0
    ).order_by('-total_clicks')[:5]

    # Recent messages
    recent_messages = Message.objects.order_by('-created_at')[:10]

    context = {
        'total_messages': total_messages,
        'active_messages': active_messages,
        'scheduled_messages': scheduled_messages,
        'draft_messages': draft_messages,
        'message_types': message_types,
        'recent_analytics': recent_analytics,
        'top_messages': top_messages,
        'recent_messages': recent_messages,
    }

    return render(request, 'core/dashboard/admin_message_management.html', context)


@login_required
def admin_message_list_tab(request):
    """HTMX endpoint for message list tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message
    from django.core.paginator import Paginator

    # Get filter parameters
    status_filter = request.GET.get('status', '')
    message_type_filter = request.GET.get('message_type', '')
    target_audience_filter = request.GET.get('target_audience', '')
    search_query = request.GET.get('search', '')

    # Base queryset
    messages_qs = Message.objects.select_related('created_by').order_by('-created_at')

    # Apply filters
    if status_filter and status_filter != 'all':
        messages_qs = messages_qs.filter(status=status_filter)

    if message_type_filter and message_type_filter != 'all':
        messages_qs = messages_qs.filter(message_type=message_type_filter)

    if target_audience_filter and target_audience_filter != 'all':
        messages_qs = messages_qs.filter(target_audience=target_audience_filter)

    if search_query:
        messages_qs = messages_qs.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(excerpt__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(messages_qs, 20)  # 20 messages per page
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'messages': page_obj,
        'page_obj': page_obj,
        'status_filter': status_filter,
        'message_type_filter': message_type_filter,
        'target_audience_filter': target_audience_filter,
        'search_query': search_query,
        'message_type_choices': Message.MESSAGE_TYPE_CHOICES,
        'target_audience_choices': Message.TARGET_AUDIENCE_CHOICES,
        'status_choices': Message.STATUS_CHOICES,
    }

    return render(request, 'core/dashboard/partials/admin_message_list_tab.html', context)


@login_required
def admin_message_create_modal(request):
    """HTMX endpoint for message creation modal"""
    # Check admin permissions
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message, MessageTemplate

    # Get available templates
    templates = MessageTemplate.objects.filter(is_active=True).order_by('category', 'name')

    context = {
        'message_type_choices': Message.MESSAGE_TYPE_CHOICES,
        'target_audience_choices': Message.TARGET_AUDIENCE_CHOICES,
        'priority_choices': Message.PRIORITY_CHOICES,
        'status_choices': Message.STATUS_CHOICES,
        'templates': templates,
    }

    return render(request, 'core/modals/admin_message_create.html', context)


@login_required
def admin_message_create(request):
    """Handle message creation"""
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    if request.method == 'POST':
        from .models import Message, MessageSchedule
        import json

        try:
            # Extract form data
            title = request.POST.get('title', '').strip()
            content = request.POST.get('content', '').strip()
            excerpt = request.POST.get('excerpt', '').strip()
            message_type = request.POST.get('message_type', 'announcement')
            target_audience = request.POST.get('target_audience', 'all')
            priority = int(request.POST.get('priority', 2))
            status = request.POST.get('status', 'draft')

            # Display settings
            show_as_popup = request.POST.get('show_as_popup') == 'on'
            show_as_banner = request.POST.get('show_as_banner') == 'on'
            show_in_dashboard = request.POST.get('show_in_dashboard') == 'on'

            # Styling
            background_color = request.POST.get('background_color', '#ffffff')
            text_color = request.POST.get('text_color', '#000000')
            icon_class = request.POST.get('icon_class', '')

            # Action button
            action_button_text = request.POST.get('action_button_text', '')
            action_button_url = request.POST.get('action_button_url', '')
            action_button_color = request.POST.get('action_button_color', '#dc2626')

            # Scheduling
            publication_date = request.POST.get('publication_date')
            expiration_date = request.POST.get('expiration_date')

            # Validation
            if not title:
                return JsonResponse({'success': False, 'error': 'Title is required'})

            if not content:
                return JsonResponse({'success': False, 'error': 'Content is required'})

            # Create message
            message = Message.objects.create(
                title=title,
                content=content,
                excerpt=excerpt,
                message_type=message_type,
                target_audience=target_audience,
                priority=priority,
                status=status,
                show_as_popup=show_as_popup,
                show_as_banner=show_as_banner,
                show_in_dashboard=show_in_dashboard,
                background_color=background_color,
                text_color=text_color,
                icon_class=icon_class,
                action_button_text=action_button_text,
                action_button_url=action_button_url,
                action_button_color=action_button_color,
                publication_date=publication_date if publication_date else None,
                expiration_date=expiration_date if expiration_date else None,
                created_by=request.user
            )

            # Handle file upload
            if 'featured_image' in request.FILES:
                message.featured_image = request.FILES['featured_image']
                message.featured_image_alt = request.POST.get('featured_image_alt', '')
                message.save()

            return JsonResponse({
                'success': True,
                'message': f'Message "{title}" created successfully!',
                'message_id': message.id
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating message: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def admin_message_edit_modal(request, message_id):
    """HTMX endpoint for message edit modal"""
    # Check admin permissions
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message, MessageTemplate

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    # Get available templates
    templates = MessageTemplate.objects.filter(is_active=True).order_by('category', 'name')

    context = {
        'message': message,
        'message_type_choices': Message.MESSAGE_TYPE_CHOICES,
        'target_audience_choices': Message.TARGET_AUDIENCE_CHOICES,
        'priority_choices': Message.PRIORITY_CHOICES,
        'status_choices': Message.STATUS_CHOICES,
        'templates': templates,
    }

    return render(request, 'core/modals/admin_message_edit.html', context)


@login_required
def admin_message_edit(request, message_id):
    """Handle message editing"""
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    if request.method == 'POST':
        try:
            # Extract form data
            message.title = request.POST.get('title', '').strip()
            message.content = request.POST.get('content', '').strip()
            message.excerpt = request.POST.get('excerpt', '').strip()
            message.message_type = request.POST.get('message_type', 'announcement')
            message.target_audience = request.POST.get('target_audience', 'all')
            message.priority = int(request.POST.get('priority', 2))
            message.status = request.POST.get('status', 'draft')

            # Display settings
            message.show_as_popup = request.POST.get('show_as_popup') == 'on'
            message.show_as_banner = request.POST.get('show_as_banner') == 'on'
            message.show_in_dashboard = request.POST.get('show_in_dashboard') == 'on'

            # Styling
            message.background_color = request.POST.get('background_color', '#ffffff')
            message.text_color = request.POST.get('text_color', '#000000')
            message.icon_class = request.POST.get('icon_class', '')

            # Action button
            message.action_button_text = request.POST.get('action_button_text', '')
            message.action_button_url = request.POST.get('action_button_url', '')
            message.action_button_color = request.POST.get('action_button_color', '#dc2626')

            # Scheduling
            publication_date = request.POST.get('publication_date')
            expiration_date = request.POST.get('expiration_date')
            message.publication_date = publication_date if publication_date else None
            message.expiration_date = expiration_date if expiration_date else None

            # Validation
            if not message.title:
                return JsonResponse({'success': False, 'error': 'Title is required'})

            if not message.content:
                return JsonResponse({'success': False, 'error': 'Content is required'})

            # Handle file upload
            if 'featured_image' in request.FILES:
                message.featured_image = request.FILES['featured_image']
                message.featured_image_alt = request.POST.get('featured_image_alt', '')

            message.save()

            return JsonResponse({
                'success': True,
                'message': f'Message "{message.title}" updated successfully!'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error updating message: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def admin_message_delete(request, message_id):
    """Handle message deletion"""
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    # Handle both DELETE requests and POST requests with _method=DELETE
    if request.method == 'DELETE' or (request.method == 'POST' and request.POST.get('_method') == 'DELETE'):
        message_title = message.title
        message.delete()

        return JsonResponse({
            'success': True,
            'message': f'Message "{message_title}" deleted successfully!'
        })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def admin_message_preview_modal(request, message_id):
    """HTMX endpoint for message preview modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    context = {
        'message': message,
    }

    return render(request, 'core/modals/admin_message_preview.html', context)


@login_required
def admin_message_search(request):
    """HTMX endpoint for message search"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message
    from django.core.paginator import Paginator

    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    message_type_filter = request.GET.get('message_type', '')
    target_audience_filter = request.GET.get('target_audience', '')

    if not search_query:
        return JsonResponse({'success': False, 'message': 'No search query provided'})

    # Base queryset
    messages_qs = Message.objects.select_related('created_by').order_by('-created_at')

    # Apply search
    messages_qs = messages_qs.filter(
        Q(title__icontains=search_query) |
        Q(content__icontains=search_query) |
        Q(excerpt__icontains=search_query)
    )

    # Apply filters
    if status_filter and status_filter != 'all':
        messages_qs = messages_qs.filter(status=status_filter)

    if message_type_filter and message_type_filter != 'all':
        messages_qs = messages_qs.filter(message_type=message_type_filter)

    if target_audience_filter and target_audience_filter != 'all':
        messages_qs = messages_qs.filter(target_audience=target_audience_filter)

    # Pagination
    paginator = Paginator(messages_qs, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'messages': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'message_type_filter': message_type_filter,
        'target_audience_filter': target_audience_filter,
    }

    return render(request, 'core/dashboard/partials/admin_message_list_tab.html', context)


@login_required
def admin_message_analytics_tab(request):
    """HTMX endpoint for message analytics tab"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message, MessageAnalytics
    from django.db.models import Sum, Avg, Count
    from datetime import timedelta

    # Get date range
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    # Overall analytics
    total_messages = Message.objects.count()
    active_messages = Message.objects.filter(status='active').count()

    # Performance metrics
    analytics_data = MessageAnalytics.objects.filter(
        date__gte=start_date.date()
    ).aggregate(
        total_displays=Sum('total_displays'),
        total_clicks=Sum('total_clicks'),
        total_dismissals=Sum('total_dismissals'),
        unique_users=Sum('unique_users_shown')
    )

    # Calculate rates
    total_displays = analytics_data['total_displays'] or 0
    total_clicks = analytics_data['total_clicks'] or 0
    total_dismissals = analytics_data['total_dismissals'] or 0

    click_through_rate = (total_clicks / total_displays * 100) if total_displays > 0 else 0
    dismissal_rate = (total_dismissals / total_displays * 100) if total_displays > 0 else 0

    # Top performing messages
    top_messages = Message.objects.filter(
        total_views__gt=0
    ).order_by('-total_clicks')[:10]

    # Daily analytics for chart
    daily_analytics = MessageAnalytics.objects.filter(
        date__gte=start_date.date()
    ).values('date').annotate(
        displays=Sum('total_displays'),
        clicks=Sum('total_clicks'),
        dismissals=Sum('total_dismissals')
    ).order_by('date')

    context = {
        'total_messages': total_messages,
        'active_messages': active_messages,
        'analytics_data': analytics_data,
        'click_through_rate': round(click_through_rate, 2),
        'dismissal_rate': round(dismissal_rate, 2),
        'top_messages': top_messages,
        'daily_analytics': list(daily_analytics),
        'days': days,
    }

    return render(request, 'core/dashboard/partials/admin_message_analytics_tab.html', context)


@htmx_login_required
def admin_message_count_htmx(request):
    """HTMX endpoint for message count badge in sidebar"""
    # Allow access for admin users and staff
    if not (request.user.is_staff or getattr(request.user, 'role', None) == 'admin'):
        return HttpResponse('0')  # Return 0 instead of 403 for non-admin users

    try:
        from .models import Message
        # Count active messages
        active_count = Message.objects.filter(status='active').count()
        return HttpResponse(str(active_count))
    except Exception:
        # Return 0 if Message model doesn't exist or any other error
        return HttpResponse('0')


# Frontend Message Display Views

@ajax_login_required
def user_messages_popup(request):
    """Get messages to display as popups for the current user"""
    from .models import Message
    from .services.message_service import MessageTargetingService, MessageAnalyticsService
    import logging

    logger = logging.getLogger(__name__)
    user = request.user

    # Get active popup messages
    messages_qs = Message.objects.filter(
        status='active',
        show_as_popup=True
    ).select_related('created_by').order_by('-priority', '-created_at')

    logger.debug(f"Found {messages_qs.count()} active popup messages for user {user.username}")

    # Find the first message that should be shown to this user
    message_to_show = None
    for message in messages_qs:
        if MessageTargetingService.should_show_message_to_user(message, user):
            message_to_show = message
            logger.debug(f"Message '{message.title}' will be shown to user {user.username}")
            break

    if message_to_show:
        # Record the view using the analytics service
        MessageAnalyticsService.record_message_view(message_to_show, user, request)

        context = {
            'message': message_to_show,
        }

        return render(request, 'core/messages/user_message_popup.html', context)

    logger.debug(f"No messages to show to user {user.username}")
    return JsonResponse({'no_messages': True})


@login_required
def user_message_action(request, message_id):
    """Handle user actions on messages (click, dismiss, etc.)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)

    from .models import Message
    from .services.message_service import MessageAnalyticsService

    try:
        message = Message.objects.get(id=message_id)
        action = request.POST.get('action', 'viewed')

        # Record the action using the analytics service
        MessageAnalyticsService.record_message_action(message, request.user, action, request)

        return JsonResponse({'success': True, 'action': action})

    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def user_dashboard_messages(request):
    """Get messages to display in user dashboard"""
    from .models import Message
    from .services.message_service import MessageTargetingService

    user = request.user

    # Get active messages that should be shown in dashboard
    messages_qs = Message.objects.filter(
        status='active',
        show_in_dashboard=True
    ).select_related('created_by').order_by('-priority', '-created_at')

    # Filter messages based on targeting rules
    valid_messages = []
    for message in messages_qs:
        if MessageTargetingService.should_show_message_to_user(message, user):
            valid_messages.append(message)
            if len(valid_messages) >= 5:  # Limit to top 5 messages
                break

    context = {
        'messages': valid_messages,
    }

    return render(request, 'core/messages/user_dashboard_messages.html', context)


@login_required
def admin_message_schedule_modal(request, message_id):
    """HTMX endpoint for message scheduling modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message, MessageSchedule

    try:
        message = Message.objects.get(id=message_id)
        schedule = getattr(message, 'schedule', None)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    context = {
        'message': message,
        'schedule': schedule,
        'frequency_choices': MessageSchedule.FREQUENCY_CHOICES,
        'weekday_choices': MessageSchedule.WEEKDAY_CHOICES,
    }

    return render(request, 'core/modals/admin_message_schedule.html', context)


@login_required
def admin_message_schedule_save(request, message_id):
    """Handle message schedule creation/update"""
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message
    from .services.message_service import MessageSchedulingService

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    if request.method == 'POST':
        try:
            # Extract schedule data
            schedule_data = {
                'frequency': request.POST.get('frequency', 'once'),
                'send_time': request.POST.get('send_time'),
                'timezone': request.POST.get('timezone', 'UTC'),
                'weekdays': request.POST.getlist('weekdays'),
                'day_of_month': request.POST.get('day_of_month'),
                'max_occurrences': request.POST.get('max_occurrences'),
                'end_date': request.POST.get('end_date'),
            }

            # Clean up empty values
            for key, value in list(schedule_data.items()):
                if value == '' or value is None:
                    schedule_data[key] = None
                elif key == 'weekdays':
                    schedule_data[key] = [int(day) for day in value if day.isdigit()]
                elif key in ['max_occurrences', 'day_of_month'] and value:
                    schedule_data[key] = int(value)

            # Create or update schedule
            if hasattr(message, 'schedule'):
                schedule = MessageSchedulingService.update_schedule(message.schedule, schedule_data)
                action = 'updated'
            else:
                schedule = MessageSchedulingService.create_schedule(message, schedule_data)
                action = 'created'

            return JsonResponse({
                'success': True,
                'message': f'Message schedule {action} successfully!'
            })

        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error saving schedule: {str(e)}'
            })

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def admin_message_targeting_modal(request, message_id):
    """HTMX endpoint for message targeting modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message, MessageTarget

    try:
        message = Message.objects.get(id=message_id)
        targeting_rules = message.targeting_rules.all()
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    context = {
        'message': message,
        'targeting_rules': targeting_rules,
        'field_choices': MessageTarget.FIELD_CHOICES,
        'condition_choices': MessageTarget.CONDITION_CHOICES,
    }

    return render(request, 'core/modals/admin_message_targeting.html', context)


@login_required
def admin_message_performance_modal(request, message_id):
    """HTMX endpoint for message performance modal"""
    # Check admin permissions
    if not request.user.is_staff and request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    from .models import Message
    from .services.message_service import MessageAnalyticsService

    try:
        message = Message.objects.get(id=message_id)
    except Message.DoesNotExist:
        return JsonResponse({'error': 'Message not found'}, status=404)

    # Get performance data for different time periods
    performance_7d = MessageAnalyticsService.get_message_performance(message, 7)
    performance_30d = MessageAnalyticsService.get_message_performance(message, 30)
    performance_90d = MessageAnalyticsService.get_message_performance(message, 90)

    context = {
        'message': message,
        'performance_7d': performance_7d,
        'performance_30d': performance_30d,
        'performance_90d': performance_90d,
    }

    return render(request, 'core/modals/admin_message_performance.html', context)


# Vendor Spare Parts HTMX Views

@login_required
@require_http_methods(["GET"])
def vendor_spare_parts_table_htmx(request):
    """HTMX endpoint for vendor spare parts table with real-time filtering"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    # Get filter parameters
    search = request.GET.get('search', '').strip()
    category = request.GET.get('category', '')
    stock_status = request.GET.get('stock_status', '')

    # Base queryset
    spare_parts = SparePart.objects.filter(vendor=vendor).select_related(
        'supplier', 'category_new'
    )

    # Apply filters
    if search:
        spare_parts = spare_parts.filter(
            Q(name__icontains=search) |
            Q(sku__icontains=search) |
            Q(part_number__icontains=search) |
            Q(barcode__icontains=search)
        )

    if category:
        spare_parts = spare_parts.filter(category_new_id=category)

    if stock_status == 'low':
        spare_parts = [part for part in spare_parts if part.is_low_stock]
    elif stock_status == 'out':
        spare_parts = spare_parts.filter(stock_quantity=0)
    elif stock_status == 'available':
        spare_parts = spare_parts.filter(stock_quantity__gt=0)

    # Pagination (10 rows per page as per user preference)
    paginator = Paginator(spare_parts, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'spare_parts': page_obj,
        'total_results': paginator.count,
    }

    return render(request, 'core/partials/vendor_spare_parts_table.html', context)


@login_required
@require_http_methods(["GET"])
def vendor_spare_parts_search_htmx(request):
    """HTMX endpoint for vendor spare parts live search"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    search_query = request.GET.get('search', '').strip()

    # Base queryset for vendor's spare parts
    spare_parts = SparePart.objects.filter(vendor=vendor).select_related('supplier', 'category_new')

    # Apply search filter only if search query is provided and has at least 2 characters
    if search_query and len(search_query) >= 2:
        spare_parts = spare_parts.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(part_number__icontains=search_query) |
            Q(barcode__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    # Pagination
    paginator = Paginator(spare_parts.order_by('-created_at'), 12)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'spare_parts': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
        'total_results': paginator.count,
    }

    return render(request, 'core/partials/vendor_spare_parts_grid.html', context)


@login_required
@require_http_methods(["GET"])
def vendor_spare_parts_stats_htmx(request):
    """HTMX endpoint for vendor spare parts statistics"""
    if request.user.role != 'vendor':
        return JsonResponse({'error': 'Access denied'}, status=403)

    try:
        vendor = request.user.vendor
    except Vendor.DoesNotExist:
        return JsonResponse({'error': 'Vendor profile not found'}, status=404)

    # Calculate statistics
    total_parts = SparePart.objects.filter(vendor=vendor, is_available=True).count()
    in_stock_parts = SparePart.objects.filter(
        vendor=vendor,
        is_available=True,
        stock_quantity__gt=10
    ).count()
    low_stock_parts = SparePart.objects.filter(
        vendor=vendor,
        is_available=True,
        stock_quantity__lte=10,
        stock_quantity__gt=0
    ).count()
    out_of_stock_parts = SparePart.objects.filter(
        vendor=vendor,
        is_available=True,
        stock_quantity=0
    ).count()

    # Calculate total inventory value
    total_value = SparePart.objects.filter(
        vendor=vendor,
        is_available=True
    ).aggregate(
        total=Sum(F('stock_quantity') * F('cost_price'))
    )['total'] or 0

    stats = {
        'total_parts': total_parts,
        'in_stock_parts': in_stock_parts,
        'low_stock_parts': low_stock_parts,
        'out_of_stock_parts': out_of_stock_parts,
        'total_value': total_value
    }

    return render(request, 'core/partials/vendor_spare_parts_stats.html', {'stats': stats})


# Admin Spare Shop Statistics HTMX Views

@login_required
@require_http_methods(["GET"])
def admin_spare_shop_stats_total(request):
    """HTMX endpoint for admin spare shop total parts statistics"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    total_parts = SparePart.objects.filter(is_available=True).count()
    return HttpResponse(str(total_parts))


@login_required
@require_http_methods(["GET"])
def admin_spare_shop_stats_in_stock(request):
    """HTMX endpoint for admin spare shop in-stock parts statistics"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    in_stock_parts = SparePart.objects.filter(
        is_available=True,
        stock_quantity__gt=10
    ).count()
    return HttpResponse(str(in_stock_parts))


@login_required
@require_http_methods(["GET"])
def admin_spare_shop_stats_out_of_stock(request):
    """HTMX endpoint for admin spare shop out-of-stock parts statistics"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    out_of_stock_parts = SparePart.objects.filter(
        is_available=True,
        stock_quantity=0
    ).count()
    return HttpResponse(str(out_of_stock_parts))


@login_required
@require_http_methods(["GET"])
def admin_spare_shop_stats_low_stock(request):
    """HTMX endpoint for admin spare shop low-stock parts statistics"""
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Access denied'}, status=403)

    low_stock_parts = SparePart.objects.filter(
        is_available=True,
        stock_quantity__lte=10,
        stock_quantity__gt=0
    ).count()
    return HttpResponse(str(low_stock_parts))
